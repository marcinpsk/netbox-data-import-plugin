# SPDX-License-Identifier: Apache-2.0
# Copyright (C) 2025 Marcin Zieba <marcinpsk@gmail.com>
"""Import engine: parse Excel files and run (or preview) imports into NetBox.

Public API
----------
parse_file(file_obj, profile)  ->  list[dict]
run_import(rows, profile, context, dry_run=True)  ->  ImportResult
"""

from __future__ import annotations

from collections.abc import Callable
from dataclasses import replace
import logging
import re
from copy import copy
from functools import partial
from dataclasses import dataclass, field
from typing import Literal
from io import BytesIO

from django.core.exceptions import ValidationError
from django.db import DatabaseError, IntegrityError, transaction
from django.utils.text import slugify
import openpyxl

from .contact_resolution import ContactResolutionRequired, PrimaryContactResolver
from .device_field_review import DeviceFieldReviewer
from .catalog import CANDIDATE_TARGET_PREFIX, has_implemented_module
from .models import DeviceImportSource, ImportProfile
from .object_permissions import (
    ObjectPermissionDenied as _ObjectPermissionDenied,
    enforce_saved_object_permission as _enforce_saved_object_permission,
)

logger = logging.getLogger(__name__)


class ParseError(Exception):
    """Raised when the source file cannot be parsed."""


class _DeviceBindingConflict(IntegrityError):
    """Raised when one source binding would replace another device identity."""


_CandidateResolutionRequired = ContactResolutionRequired


# ---------------------------------------------------------------------------
# Result dataclasses
# ---------------------------------------------------------------------------


@dataclass
class RowResult:
    """Holds the result of processing a single source row."""

    row_number: int
    source_id: str
    name: str
    action: Literal["create", "update", "skip", "error", "ignore"]
    object_type: Literal["rack", "device", "manufacturer", "device_type", ""]
    detail: str
    netbox_url: str = ""
    rack_name: str = ""
    # Contextual metadata used by the preview template for inline quick-fix actions
    extra_data: dict = field(default_factory=dict)

    def to_dict(self) -> dict:
        """Serialize this result to a plain dict."""
        return self.__dict__.copy()

    @classmethod
    def from_dict(cls, d: dict) -> "RowResult":
        """Deserialize a RowResult from a plain dict."""
        d = dict(d)
        d.setdefault("extra_data", {})
        return cls(**d)


@dataclass
class ImportResult:
    """Aggregates all RowResult objects and summary counts for an import run."""

    rows: list[RowResult] = field(default_factory=list)
    counts: dict = field(default_factory=dict)
    has_errors: bool = False

    def _recompute_counts(self):
        c: dict = {}
        for r in self.rows:
            if r.action == "error":
                c["errors"] = c.get("errors", 0) + 1
            elif r.action == "skip":
                c["skipped"] = c.get("skipped", 0) + 1
            elif r.action == "ignore":
                c["ignored"] = c.get("ignored", 0) + 1
            elif r.action in ("create", "update"):
                key = f"{r.object_type}s_{r.action}d"
                c[key] = c.get(key, 0) + 1
        self.counts = c
        self.has_errors = c.get("errors", 0) > 0

    def to_session_dict(self) -> dict:
        """Serialize this result to a session-safe dict."""
        # Store parsed rows so the execute step can re-use them
        return {
            "rows": [r.to_dict() for r in self.rows],
            "counts": self.counts,
            "has_errors": self.has_errors,
        }

    @classmethod
    def from_session_dict(cls, d: dict) -> "ImportResult":
        """Deserialize an ImportResult from a session-stored dict."""
        result = cls()
        result.rows = [RowResult.from_dict(r) for r in d.get("rows", [])]
        result.counts = d.get("counts", {})
        result.has_errors = d.get("has_errors", False)
        return result

    @property
    def rack_groups(self) -> dict:
        """Return rows grouped by rack name for the rack view template.

        Devices within each group are sorted by u_position (ascending, with
        devices that have no position placed last).  Rack rows with an empty
        name (caused by cabinet source rows that have no RACK column value) are
        excluded so they don't produce a confusing unnamed card.
        """
        groups: dict = {}
        for row in self.rows:
            if row.object_type == "rack":
                if not row.name:
                    continue
                if row.name not in groups:
                    groups[row.name] = {"rack_row": row, "devices": []}
                else:
                    groups[row.name]["rack_row"] = row
            elif row.object_type == "device":
                rack = row.rack_name or "(No rack)"
                if rack not in groups:
                    groups[rack] = {"rack_row": None, "devices": []}
                groups[rack]["devices"].append(row)
        # Sort each rack's device list by u_position ascending (None → last)
        for group in groups.values():
            group["devices"].sort(
                key=lambda r: (
                    r.extra_data.get("u_position") is None,
                    r.extra_data.get("u_position") or 0,
                )
            )
        return groups


@dataclass
class ImportContext:
    """Shared execution context passed through all import pipeline stages.

    Holds profile/site/tenant/flags so internal helpers do not need
    to accept those as separate positional arguments.  ``rack_map`` is
    populated by pass 2 and consumed by pass 3.
    """

    profile: ImportProfile
    site: object
    location: object | None
    tenant: object | None
    dry_run: bool
    result: ImportResult
    rack_map: dict = field(default_factory=dict)
    pending_device_roles: set = field(default_factory=set)
    user: object | None = None
    # Tracks ``(rack_name, position, face) -> (row_number, device_name)`` for
    # rows that claim a rack position in the current file, so within-file
    # rack-position duplicates are flagged before they would cause an
    # IntegrityError on save.
    claimed_positions: dict = field(default_factory=dict)
    # Maps ``(normalized rack identity, position)`` to claimed faces. This
    # avoids scanning every position claim for full-depth devices.
    claimed_position_faces: dict = field(default_factory=dict)
    # Memoizes ``DeviceType.u_height == 0`` lookups by ``(mfg_slug, dt_slug)``
    # to avoid an N+1 query in preview for large imports.
    zero_u_cache: dict = field(default_factory=dict)
    device_type_cache: dict = field(default_factory=dict)
    device_role_cache: dict = field(default_factory=dict)
    # Captures the previewed identity action for each row. Execute mode must
    # not change a previewed create into an update, or update another object.
    expected_intents: dict = field(default_factory=dict)
    duplicate_source_ids: frozenset = frozenset()
    duplicate_source_id_rows: dict = field(default_factory=dict)
    duplicate_serials: frozenset = frozenset()
    # Maps a duplicated serial to every source row number that carries it.
    duplicate_serial_rows: dict = field(default_factory=dict)
    duplicate_asset_tags: frozenset = frozenset()
    duplicate_asset_tag_rows: dict = field(default_factory=dict)
    duplicate_rack_names: frozenset = frozenset()
    ignored_source_ids: frozenset = frozenset()
    claimed_device_ids: dict = field(default_factory=dict)
    slug_conflicts_by_row: dict = field(default_factory=dict)
    reserved_device_names: set = field(default_factory=set)
    candidate_source_columns: dict[str, frozenset[str]] = field(default_factory=dict)
    progress_callback: Callable[[int, int], None] | None = None
    field_reviewer: DeviceFieldReviewer | None = None
    device_type_identity: _DeviceTypeIdentityResolver | None = None
    effective_duplicate_identity: dict[int, dict[str, str | None]] = field(default_factory=dict)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_NONE_LIKE = frozenset({"none", "nan", "null", "n/a", "#n/a"})


def _duplicate_value_detail(label: str, value: str, other_rows: list[int]) -> str:
    """Name a duplicated identity value and every other source row that carries it."""
    where = ", ".join(f"row {number}" for number in other_rows)
    return f"Duplicate {label} '{value}' appears more than once in this import" + (
        f", also on {where}." if where else "."
    )


def _str_val(v) -> str:
    """Safely convert a cell value to a stripped string.

    None, NaN (pandas), and sentinel strings like "None"/"nan"/"null" are
    returned as an empty string so callers never see the literal text "None".
    """
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in _NONE_LIKE else s


def _coerce_int(value, default=None):
    """Return a source value as an int, or the default when it is not a finite number."""
    try:
        return int(float(value))
    except (TypeError, ValueError, OverflowError):
        return default


def _coerce_position(value, default=None):
    """Return a finite source rack position without losing half-U precision."""
    from decimal import Decimal, InvalidOperation
    import math

    if value is None or value == "":
        return default
    try:
        position = Decimal(str(value).strip())
    except (InvalidOperation, TypeError, ValueError):
        return default
    if not position.is_finite():
        return default
    if not math.isfinite(float(position)):
        return default
    if position == position.to_integral_value():
        return int(position)
    return position


def _has_below_rack_position(row) -> bool:
    """Return True when a source row is explicitly below rack unit 1."""
    position = _coerce_position(row.get("u_position"))
    return position is not None and position < 1


def _is_writing_device_row(row, crm, ignored_source_ids) -> bool:
    """Return True when a row is eligible to create or update a device."""
    source_id = _str_val(row.get("source_id"))
    return bool(
        crm
        and not crm.creates_rack
        and not crm.ignore
        and crm.role_slug
        and source_id not in ignored_source_ids
        and not _has_below_rack_position(row)
    )


def _effective_device_name(row) -> str:
    """Return the imported name, including the asset-tag fallback."""
    device_name = _str_val(row.get("device_name"))
    if device_name:
        return device_name
    return _str_val(row.get("asset_tag"))[:50]


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------


def _build_header_index_map(ws) -> dict[str, int]:
    """Build a header-name → column-index map from the first worksheet row.

    First occurrence wins when duplicate headers exist.
    """
    raw_headers: dict[str, int] = {}
    for idx, cell in enumerate(ws[1]):
        if cell.value is not None:
            header = str(cell.value).strip()
            if header not in raw_headers:
                raw_headers[header] = idx
    return raw_headers


def _apply_transform_rules(row_dict: dict, raw_row, raw_headers: dict, transform_rules) -> None:
    """Apply column transform rules in-place to *row_dict*."""
    for rule in transform_rules:
        idx = raw_headers.get(rule.source_column)
        if idx is None:
            continue
        raw_value = raw_row[idx] if idx < len(raw_row) else None
        if raw_value is None:
            continue
        raw_str = str(raw_value).strip()
        try:
            m = re.fullmatch(rule.pattern, raw_str)
        except re.error as exc:
            raise ParseError(
                f"Invalid regex pattern '{rule.pattern}' in transform rule for column "
                f"'{rule.source_column}' (value: {raw_str!r}): {exc}"
            ) from exc
        if m and rule.group_1_target and len(m.groups()) >= 1:
            row_dict[rule.group_1_target] = m.group(1)
        if m and rule.group_2_target and len(m.groups()) >= 2:
            row_dict[rule.group_2_target] = m.group(2)


def _collect_unmapped_values(row, raw_headers, unmapped_cols, unused_stats, return_stats, capture_extra):
    """Return extra dict for a single row and update unused_stats in-place."""
    extra: dict[str, str] = {}
    for col in unmapped_cols:
        idx = raw_headers[col]
        raw_val = row[idx] if idx < len(row) else None
        str_val = _str_val(raw_val)
        if not str_val:
            continue
        extra[col] = str_val
        if return_stats:
            entry = unused_stats.setdefault(col, {"count": 0, "samples": []})
            entry["count"] += 1
            if len(entry["samples"]) < 5:
                entry["samples"].append(str_val)
    return extra


def _promote_extra_json_fields(row_dict: dict) -> None:
    """Move any extra_json:<key> entries from row_dict into row_dict["_extra_columns"]."""
    for k in [k for k in list(row_dict) if isinstance(k, str) and k.startswith("extra_json:")]:
        json_key = k[len("extra_json:") :]
        val = row_dict.pop(k)
        if val not in (None, ""):
            row_dict.setdefault("_extra_columns", {})[json_key] = val


_NUMERIC_TARGET_FIELDS: frozenset[str] = frozenset({"u_position", "u_height"})


def _cmp_for_field(field: str, val) -> str:
    """Return the comparison key for *val* appropriate to *field* type."""
    if field in _NUMERIC_TARGET_FIELDS:
        return _normalize_for_compare(val)
    return "" if val is None else str(val).strip()


def _build_grouped_col_map(profile: ImportProfile) -> dict[str, list[str]]:
    """Return a target-field keyed map of all mapped source columns."""
    grouped: dict[str, list[str]] = {}
    for cm in profile.column_mappings.all():
        grouped.setdefault(cm.target_field, []).append(cm.source_column)
    return grouped


def _build_source_to_targets_map(profile: ImportProfile) -> dict[str, list[str]]:
    """Return a source-column keyed map of all target fields that column feeds."""
    rev: dict[str, list[str]] = {}
    for cm in profile.column_mappings.all():
        rev.setdefault(cm.source_column, []).append(cm.target_field)
    return rev


def _apply_one_resolution(row_dict: dict, res, source_to_targets: dict[str, list[str]]) -> None:
    """Apply a single SourceResolution to a row dict.

    The user's split modal lets them assign each split part to a target field
    OR ignore it (no field selected).  Ignored parts must result in the
    corresponding target field being cleared — otherwise the original
    pre-split value (e.g. ``"TEST-ASSET-002 - temporary device ..."``) silently
    persists in the device_name field.

    Logic:
      1. Apply explicit ``resolved_fields`` overrides.
      2. For each target_field that the resolution's source_column originally
         feeds: if it's not in ``resolved_fields`` and its current value still
         equals the resolution's ``original_value`` (i.e. no other column
         contributed a different value via multi-source merge), clear it.
    """
    row_dict.update(res.resolved_fields)
    _clear_resolved_conflicts(row_dict, res.resolved_fields)

    # The split modal's data-source-column attribute stores the target field
    # name (e.g. "device_name"), not the original CSV column header — so we
    # treat the resolution's source_column as a candidate target_field too.
    candidate_targets = list(source_to_targets.get(res.source_column, []))
    if res.source_column not in candidate_targets:
        candidate_targets.append(res.source_column)
    for target_field in candidate_targets:
        if target_field in res.resolved_fields:
            continue
        current = row_dict.get(target_field)
        if current is None:
            continue
        if str(current) == str(res.original_value):
            row_dict[target_field] = None


def _merge_row_values(
    row_num: int,
    raw_row,
    raw_headers: dict[str, int],
    grouped_col_map: dict[str, list[str]],
) -> dict[str, object]:
    """Build a parsed row dict using multi-source merge semantics."""
    row_dict: dict[str, object] = {"_row_number": row_num}
    for target_field, source_cols in grouped_col_map.items():
        values: dict[str, object] = {}
        for source_col in source_cols:
            idx = raw_headers.get(source_col)
            if idx is None:
                continue
            value = raw_row[idx] if idx < len(raw_row) else None
            if isinstance(value, str):
                value = value.strip()
            if value is not None and str(value).strip():
                values[source_col] = value

        if not values:
            continue
        if target_field.startswith(CANDIDATE_TARGET_PREFIX):
            candidate_target = target_field.removeprefix(CANDIDATE_TARGET_PREFIX)
            row_dict.setdefault("_candidate_values", {})[candidate_target] = {
                source_column: str(value) for source_column, value in values.items()
            }
            continue
        if len({_cmp_for_field(target_field, v) for v in values.values()}) == 1:
            row_dict[target_field] = next(iter(values.values()))
        else:
            row_dict[target_field] = None
            row_dict.setdefault("_conflicts", {})[target_field] = {k: str(v) for k, v in values.items()}
    return row_dict


def _clear_resolved_conflicts(row_dict: dict[str, object], resolved_fields: dict) -> None:
    """Remove conflicts for fields overridden by saved resolutions."""
    for resolved_field in resolved_fields:
        row_dict.get("_conflicts", {}).pop(resolved_field, None)
    if not row_dict.get("_conflicts"):
        row_dict.pop("_conflicts", None)


def apply_column_mappings(rows: list[dict], profile: ImportProfile) -> list[dict]:
    """Re-apply the profile's column mappings to already-parsed session rows.

    Used after a quick-add column mapping so that in-session row dicts reflect
    the new mapping without requiring the source file to be re-uploaded.
    Handles multi-source merge: if a newly-mapped source column conflicts with an
    already-mapped value for the same target field, a _conflicts entry is recorded.
    """
    grouped = _build_grouped_col_map(profile)

    for row in rows:
        extra_columns = row.get("_extra_columns", {})
        for target_field, source_cols in grouped.items():
            unmapped = {}
            for sc in source_cols:
                if sc in row:
                    val = row.pop(sc)
                elif sc in extra_columns:
                    val = extra_columns.pop(sc)
                else:
                    continue
                if val is not None and str(val).strip():
                    unmapped[sc] = val

            if not unmapped:
                continue

            if target_field.startswith(CANDIDATE_TARGET_PREFIX):
                candidate_target = target_field.removeprefix(CANDIDATE_TARGET_PREFIX)
                candidates = row.setdefault("_candidate_values", {}).setdefault(candidate_target, {})
                candidates.update({source_column: str(value) for source_column, value in unmapped.items()})
                continue

            existing = row.get(target_field)
            existing_nonempty = existing is not None and str(existing).strip() != ""

            if not existing_nonempty:
                unique = {_cmp_for_field(target_field, v) for v in unmapped.values()}
                if len(unique) == 1:
                    row[target_field] = next(iter(unmapped.values()))
                else:
                    row[target_field] = None
                    row.setdefault("_conflicts", {})[target_field] = {k: str(v) for k, v in unmapped.items()}
            else:
                cmp_existing = _cmp_for_field(target_field, existing)
                cmp_unmapped = {_cmp_for_field(target_field, v) for v in unmapped.values()}
                if len({cmp_existing} | cmp_unmapped) > 1:
                    all_candidates = {target_field: str(existing), **{k: str(v) for k, v in unmapped.items()}}
                    row[target_field] = None
                    row.setdefault("_conflicts", {})[target_field] = all_candidates

        if not extra_columns:
            row.pop("_extra_columns", None)
        _promote_extra_json_fields(row)

    return rows


def parse_file(file_obj, profile: ImportProfile, return_stats: bool = False):
    """Read the Excel file and return a list of row-dicts keyed by target_field name.

    When return_stats=True, returns a tuple (rows, unused_stats) where unused_stats
    is a dict mapping unmapped source column names to {"count": int, "samples": list[str]}.

    Raises ParseError if the file or sheet is invalid.
    """
    if not has_implemented_module(profile.output_kinds):
        raise ParseError(f"This release has no Target Module for the '{profile.source_adapter}' source adapter.")

    try:
        content = file_obj.read()
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise ParseError(f"Cannot open Excel file: {exc}") from exc

    if profile.adapter_settings.sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise ParseError(f"Sheet '{profile.adapter_settings.sheet_name}' not found. Available sheets: {available}")

    ws = wb[profile.adapter_settings.sheet_name]
    raw_headers = _build_header_index_map(ws)

    # Build grouped source-column map from profile
    col_map = _build_grouped_col_map(profile)

    # Unmapped columns: present in the sheet but not in any mapping
    all_mapped_sources = {src for srcs in col_map.values() for src in srcs}
    unmapped_cols = [col for col in raw_headers if col not in all_mapped_sources]

    # Pre-fetch transform rules for efficiency
    transform_rules = list(profile.column_transform_rules.all())

    unused_stats: dict[str, dict] = {}
    capture_extra = profile.adapter_settings.capture_extra_data

    rows = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip fully empty rows
        if all(v is None for v in row):
            continue

        row_dict = _merge_row_values(row_num, row, raw_headers, col_map)

        # Promote explicit extra_json: mappings into _extra_columns
        _promote_extra_json_fields(row_dict)

        _apply_transform_rules(row_dict, row, raw_headers, transform_rules)

        if return_stats or capture_extra:
            extra = _collect_unmapped_values(row, raw_headers, unmapped_cols, unused_stats, return_stats, capture_extra)
            if capture_extra and extra:
                row_dict.setdefault("_extra_columns", {}).update(extra)

        rows.append(row_dict)

    if return_stats:
        return rows, unused_stats
    return rows


def derive_effective_rows(rows: list[dict], profile) -> list[dict]:
    """Return *rows* with every saved SourceResolution applied, leaving *rows* untouched.

    `rows` must be the pristine parsed rows. Applying a resolution only ever sets fields, so a
    derivation that starts from an earlier result cannot express a target field the operator has
    since dropped. Every caller derives, and none stores what it derived.
    """
    resolutions_by_source_id: dict[str, list] = {}
    # Two resolutions can share a source_id, and a later one overrides the fields an earlier one
    # set. Meta.ordering stops at source_id, so source_column completes the order the callers
    # compare against each other.
    for res in profile.source_resolutions.order_by("source_id", "source_column"):
        resolutions_by_source_id.setdefault(str(res.source_id), []).append(res)

    if not resolutions_by_source_id:
        return rows

    source_to_targets = _build_source_to_targets_map(profile)

    result = []
    for row in rows:
        source_id = _str_val(row.get("source_id"))
        if source_id and source_id in resolutions_by_source_id:
            row = dict(row)  # shallow copy — don't mutate the session dict
            if "_conflicts" in row:
                row["_conflicts"] = dict(row["_conflicts"])
            for res in resolutions_by_source_id[source_id]:
                _apply_one_resolution(row, res, source_to_targets)
        result.append(row)
    return result


# Bounded so a word cannot leave a shorter valid address: `2001:db8::1backup` -> `2001:db8::1bac`.
_IP_TOKEN = re.compile(r"(?<![0-9A-Za-z])[0-9A-Fa-f:.]+(?:/\d{1,3})?(?![0-9A-Za-z])")


def _normalized_ip(token: str) -> str | None:
    """Return *token* as 'address/prefix', or None when it is not one address."""
    import ipaddress

    try:
        if "/" in token:
            return str(ipaddress.ip_interface(token))
        addr = ipaddress.ip_address(token)
    except ValueError:
        return None
    return f"{addr}/32" if addr.version == 4 else f"{addr}/128"


def _parse_ip_with_prefix(raw_value: str) -> str | None:
    """Return the one address a source value names, as 'address/prefix', or None.

    Sources export an address inside a label or with a separator appended, so the whole value is
    tried first and the addresses spelled inside it only after that.
    """
    raw = str(raw_value).strip()
    if not raw:
        return None
    whole = _normalized_ip(raw)
    if whole is not None:
        return whole
    for token in _IP_TOKEN.findall(raw):
        found = _normalized_ip(token)
        if found is not None:
            return found
    return None


def _normalize_mapping_text(value: str) -> str:
    r"""Normalize whitespace and decode JavaScript-style \uXXXX escapes."""
    value = re.sub(r"\\u([0-9a-fA-F]{4})", lambda match: chr(int(match.group(1), 16)), value)
    return " ".join(value.split())


class _DeviceTypeIdentityResolver:
    """Resolve all profile Device Type identities from two batch-loaded indexes."""

    def __init__(self, device_type_mappings, manufacturer_mappings):
        self.device_type_mappings = tuple(device_type_mappings)
        self.manufacturer_mappings = tuple(manufacturer_mappings)
        self._device_types_exact = {}
        self._device_types_by_make = {}
        for mapping in self.device_type_mappings:
            self._device_types_exact.setdefault((mapping.source_make, mapping.source_model), mapping)
            self._device_types_by_make.setdefault(mapping.source_make.lower(), []).append(mapping)
        self._manufacturers_exact = {}
        for mapping in self.manufacturer_mappings:
            self._manufacturers_exact.setdefault(mapping.source_make, mapping)
        self.mapped_source_makes = frozenset(self._manufacturers_exact)

    @classmethod
    def for_profile(cls, profile):
        """Load both mapping tables once for one import run."""
        return cls(
            profile.device_type_mappings.all(),
            profile.manufacturer_mappings.all(),
        )

    def resolve(self, make: str, model: str) -> tuple[str, str, bool]:
        """Return manufacturer slug, Device Type slug, and explicit status."""
        mapping = self._device_types_exact.get((make, model))
        if mapping is None:
            mapping = next(
                (
                    candidate
                    for candidate in self._device_types_by_make.get(make.lower(), ())
                    if _normalize_mapping_text(candidate.source_model) == model
                ),
                None,
            )
        if mapping is not None:
            return mapping.netbox_manufacturer_slug, mapping.netbox_device_type_slug, True

        manufacturer_mapping = self._manufacturers_exact.get(make)
        if manufacturer_mapping is None:
            manufacturer_mapping = next(
                (
                    candidate
                    for candidate in self.manufacturer_mappings
                    if _normalize_mapping_text(candidate.source_make) == make
                ),
                None,
            )
        manufacturer_slug = (
            manufacturer_mapping.netbox_manufacturer_slug if manufacturer_mapping is not None else slugify(make)[:50]
        )
        return manufacturer_slug, slugify(f"{make}-{model}")[:50], False


def _resolve_device_type_slugs(
    make: str,
    model: str,
    profile: ImportProfile,
    resolver: _DeviceTypeIdentityResolver | None = None,
) -> tuple[str, str, bool]:
    """Resolve one Device Type identity through a shared batch index."""
    return (resolver or _DeviceTypeIdentityResolver.for_profile(profile)).resolve(make, model)


# ---------------------------------------------------------------------------
# Main import runner — pass helpers
# ---------------------------------------------------------------------------

# Value-translation maps (shared across passes)
_STATUS_MAP = {
    "live": "active",
    "production": "active",
    "planned": "planned",
    "staged": "staged",
    "failed": "failed",
    "offline": "offline",
    "decommissioning": "decommissioning",
}


def _get_translation_maps():
    """Return (SIDE_MAP, AIRFLOW_MAP, STATUS_MAP) with lazy-imported choice values."""
    from dcim.choices import DeviceAirflowChoices, DeviceFaceChoices

    side = {
        "front": DeviceFaceChoices.FACE_FRONT,
        "back": DeviceFaceChoices.FACE_REAR,
        "rear": DeviceFaceChoices.FACE_REAR,
    }
    airflow = {
        "front to back": DeviceAirflowChoices.AIRFLOW_FRONT_TO_REAR,
        "back to front": DeviceAirflowChoices.AIRFLOW_REAR_TO_FRONT,
        "passive": DeviceAirflowChoices.AIRFLOW_PASSIVE,
    }
    return side, airflow, _STATUS_MAP


def _perm_denied_row(perm: str, row: dict, name: str, object_type: str) -> RowResult:
    """Return a permission-denied RowResult for a write operation the user may not perform."""
    return RowResult(
        row_number=row["_row_number"],
        source_id=_str_val(row.get("source_id")),
        name=name,
        action="error",
        object_type=object_type,
        detail=f"Permission denied: {perm}",
    )


def _ensure_manufacturer(
    mfg_slug,
    make,
    seen_manufacturers,
    ctx,
    row,
    Manufacturer,
    *,
    explicit_identity=False,
):
    """Create (or preview-create) a manufacturer if not yet seen."""
    if ctx.dry_run and mfg_slug in seen_manufacturers:
        return
    seen_manufacturers.add(mfg_slug)
    if not ctx.dry_run:
        if ctx.profile.adapter_settings.create_missing_device_types:
            if not Manufacturer.objects.filter(slug=mfg_slug).exists():
                if ctx.user is not None and not ctx.user.has_perm("dcim.add_manufacturer"):
                    ctx.result.rows.append(_perm_denied_row("dcim.add_manufacturer", row, make, "manufacturer"))
                    return
            manufacturer, _ = Manufacturer.objects.get_or_create(slug=mfg_slug, defaults={"name": make})
            if not explicit_identity and _identity_text(manufacturer.name) != _identity_text(make):
                ctx.slug_conflicts_by_row[row["_row_number"]] = (
                    f"Manufacturer '{make}' derives slug '{mfg_slug}', which now belongs to "
                    f"'{manufacturer.name}' in NetBox. Refresh the preview and add an explicit manufacturer mapping."
                )
    elif (
        not Manufacturer.objects.filter(slug=mfg_slug).exists()
        and ctx.profile.adapter_settings.create_missing_device_types
    ):
        if ctx.user is not None and not ctx.user.has_perm("dcim.add_manufacturer"):
            ctx.result.rows.append(_perm_denied_row("dcim.add_manufacturer", row, make, "manufacturer"))
        else:
            ctx.result.rows.append(
                RowResult(
                    row_number=row["_row_number"],
                    source_id=str(row.get("source_id", "")),
                    name=make,
                    action="create",
                    object_type="manufacturer",
                    detail=f"Would create manufacturer '{make}' (slug: {mfg_slug})",
                    extra_data={"source_make": make, "mfg_slug": mfg_slug},
                )
            )


def _ensure_device_type(
    mfg_slug,
    dt_slug,
    make,
    model,
    u_height,
    seen_device_types,
    ctx,
    row,
    Manufacturer,
    DeviceType,
    *,
    explicit_identity=False,
):
    """Create (or preview-create) a device type if not yet seen."""
    dt_key = (mfg_slug, dt_slug)
    if ctx.dry_run and dt_key in seen_device_types:
        return
    seen_device_types.add(dt_key)
    if not ctx.dry_run:
        if ctx.profile.adapter_settings.create_missing_device_types:
            if not DeviceType.objects.filter(manufacturer__slug=mfg_slug, slug=dt_slug).exists():
                if not Manufacturer.objects.filter(slug=mfg_slug).exists():
                    if ctx.user is not None and not ctx.user.has_perm("dcim.add_manufacturer"):
                        ctx.result.rows.append(
                            _perm_denied_row("dcim.add_manufacturer", row, f"{make} / {model}", "device_type")
                        )
                        return
                if ctx.user is not None and not ctx.user.has_perm("dcim.add_devicetype"):
                    ctx.result.rows.append(
                        _perm_denied_row("dcim.add_devicetype", row, f"{make} / {model}", "device_type")
                    )
                    return
            mfg, _ = Manufacturer.objects.get_or_create(slug=mfg_slug, defaults={"name": make})
            if not explicit_identity and _identity_text(mfg.name) != _identity_text(make):
                ctx.slug_conflicts_by_row[row["_row_number"]] = (
                    f"Manufacturer '{make}' derives slug '{mfg_slug}', which now belongs to "
                    f"'{mfg.name}' in NetBox. Refresh the preview and add an explicit manufacturer mapping."
                )
                return
            device_type, _ = DeviceType.objects.get_or_create(
                manufacturer=mfg, slug=dt_slug, defaults={"model": model, "u_height": u_height}
            )
            if not explicit_identity and _identity_text(device_type.model) != _identity_text(model):
                ctx.slug_conflicts_by_row[row["_row_number"]] = (
                    f"Device type '{make} / {model}' derives slug '{mfg_slug}/{dt_slug}', which now belongs to "
                    f"'{mfg.name} / {device_type.model}' in NetBox. Refresh the preview and add an explicit mapping."
                )
        return
    exists = DeviceType.objects.filter(manufacturer__slug=mfg_slug, slug=dt_slug).exists()
    if exists:
        return
    if ctx.profile.adapter_settings.create_missing_device_types:
        if ctx.user is not None and not ctx.user.has_perm("dcim.add_devicetype"):
            ctx.result.rows.append(_perm_denied_row("dcim.add_devicetype", row, f"{make} / {model}", "device_type"))
        else:
            ctx.result.rows.append(
                RowResult(
                    row_number=row["_row_number"],
                    source_id=str(row.get("source_id", "")),
                    name=f"{make} / {model}",
                    action="create",
                    object_type="device_type",
                    detail=f"Would create device type '{model}' under '{make}'",
                    extra_data={
                        "source_make": make,
                        "source_model": model,
                        "mfg_slug": mfg_slug,
                        "dt_slug": dt_slug,
                        "u_height": u_height,
                    },
                )
            )
    else:
        ctx.result.rows.append(
            RowResult(
                row_number=row["_row_number"],
                source_id=str(row.get("source_id", "")),
                name=f"{make} / {model}",
                action="error",
                object_type="device_type",
                detail=f"Device type not found: {make} / {model} — add a mapping or enable 'Create missing device types'",
                extra_data={
                    "source_make": make,
                    "source_model": model,
                    "mfg_slug": mfg_slug,
                    "dt_slug": dt_slug,
                    "u_height": u_height,
                },
            )
        )


def _ensure_device_role(crm, seen_roles, ctx, DeviceRole):
    """Create a device role if not yet seen, respecting user permissions."""
    if not (crm and crm.role_slug and crm.role_slug not in seen_roles):
        return
    seen_roles.add(crm.role_slug)
    if ctx.dry_run:
        if ctx.user is None or ctx.user.has_perm("dcim.add_devicerole"):
            ctx.pending_device_roles.add(crm.role_slug)
    else:
        if ctx.user is not None and not ctx.user.has_perm("dcim.add_devicerole"):
            return
        DeviceRole.objects.get_or_create(
            slug=crm.role_slug,
            defaults={"name": crm.role_slug.replace("-", " ").title(), "color": "9e9e9e"},
        )


def _pass1_device_review(row, ctx, crm, Device):
    """Return the matched row review needed before relation side effects."""
    if ctx.field_reviewer is None:
        return None
    source_id = _str_val(row.get("source_id"))
    if not source_id:
        return None
    review_device_ids = ctx.field_reviewer.review_device_ids(source_id)
    if not review_device_ids:
        return None
    make = " ".join((_str_val(row.get("make")) or "Unknown").split())
    model = " ".join((_str_val(row.get("model")) or "Unknown").split())
    mfg_slug, dt_slug, _explicit_identity = _resolve_device_type_slugs(
        make,
        model,
        ctx.profile,
        ctx.device_type_identity,
    )
    device_name = _effective_device_name(row)
    matched_device, _match_method = _find_existing_device(
        ctx.profile,
        source_id,
        ctx.site,
        device_name,
        _str_val(row.get("serial")),
        (_str_val(row.get("asset_tag")) or "")[:50],
        Device,
        tenant=ctx.tenant,
        device_queryset=_device_queryset_for_user(Device, ctx.user, "view") if ctx.user is not None else None,
        review_device_ids=review_device_ids,
    )
    if matched_device is None:
        return None
    proposal = {"device_type": (mfg_slug, dt_slug, make, model)}
    if crm and crm.role_slug:
        proposal["role"] = crm.role_slug
    return ctx.field_reviewer.review(source_id, matched_device, proposal)


def _pass1_ensure_types(rows, ctx, class_role_map):
    """Pass 1: ensure Manufacturer, DeviceType, and DeviceRole objects exist."""
    from dcim.models import Device, DeviceRole, DeviceType, Manufacturer

    seen_manufacturers: set[str] = set()
    seen_device_types: set[tuple] = set()
    seen_roles: set[str] = set()

    for row in rows:
        if row.get("_row_number") in ctx.slug_conflicts_by_row:
            continue
        device_class = _str_val(row.get("device_class"))
        crm = class_role_map.get(device_class)
        if not _is_writing_device_row(row, crm, ctx.ignored_source_ids):
            continue

        make = " ".join((_str_val(row.get("make")) or "Unknown").split())
        model = " ".join((_str_val(row.get("model")) or "Unknown").split())
        u_height_raw = row.get("u_height", 1)
        u_height = max(1, _coerce_int(u_height_raw, 1))

        mfg_slug, dt_slug, explicit_identity = _resolve_device_type_slugs(
            make,
            model,
            ctx.profile,
            ctx.device_type_identity,
        )
        review = _pass1_device_review(row, ctx, crm, Device)
        if review is None or "device_type" not in review.ignored:
            _ensure_manufacturer(
                mfg_slug,
                make,
                seen_manufacturers,
                ctx,
                row,
                Manufacturer,
                explicit_identity=explicit_identity,
            )
            if row.get("_row_number") in ctx.slug_conflicts_by_row:
                continue
            _ensure_device_type(
                mfg_slug,
                dt_slug,
                make,
                model,
                u_height,
                seen_device_types,
                ctx,
                row,
                Manufacturer,
                DeviceType,
                explicit_identity=explicit_identity,
            )
        if review is None or "role" not in review.ignored:
            _ensure_device_role(crm, seen_roles, ctx, DeviceRole)


def _rack_query(Rack, ctx, rack_name):
    """Return the rack query for the selected NetBox location."""
    location_filter = {"location": ctx.location} if ctx.location is not None else {"location__isnull": True}
    return Rack.objects.filter(site=ctx.site, name__iexact=rack_name, **location_filter)


def _get_unique_rack(Rack, ctx, rack_name, *, lock=False, permission_action=None):
    """Return one rack by active location, plus an ambiguity flag."""
    queryset = _rack_query(Rack, ctx, rack_name)
    if lock:
        queryset = queryset.select_for_update(of=("self",))
    racks = list(queryset[:2])
    if len(racks) > 1:
        return None, True
    rack = racks[0] if racks else None
    if (
        rack is not None
        and ctx.user is not None
        and permission_action is not None
        and not Rack.objects.restrict(ctx.user, permission_action).filter(pk=rack.pk).exists()
    ):
        return None, False
    return rack, False


def _tenant_filter(tenant):
    """Return an ORM filter for one tenant identity, including no tenant."""
    return {"tenant": tenant} if tenant is not None else {"tenant__isnull": True}


def _rack_identity_label(rack_name, location):
    """Return a rack label that includes location when one is selected."""
    if not rack_name:
        return ""
    return f"{location} / {rack_name}" if location is not None else rack_name


def _device_rack_identity_label(device):
    """Return the location-aware rack identity for a NetBox device."""
    if not device.rack_id:
        return ""
    return _rack_identity_label(device.rack.name, device.rack.location)


def placement_sync_is_noop(device, rack_name, position, face) -> bool:
    """Return whether the placement quick action would write nothing to *device*.

    Mirrors `views._set_rack_placement`: it always sets the rack, sets position and face only when
    the source supplies them, and clears both for a zero-U device type. A full import writes the
    position unconditionally, so it can still clear a position this row omits. The two are not the
    same question, which is why this one never claims that the placement matches.
    """
    device_type = getattr(device, "device_type", None)
    if device_type is not None and device_type.u_height == 0:
        if device.position is not None or device.face:
            return False
        position = face = None
    device_rack_name = device.rack.name if device.rack_id else ""
    if _identity_text(device_rack_name) != _identity_text(rack_name):
        return False
    if position is not None and _normalize_for_compare(device.position) != _normalize_for_compare(position):
        return False
    if face and (device.face or "") != face:
        return False
    return True


def _device_placement_differs(device, source_location_id, rack_name, position, face):
    """Return whether a source placement differs from a NetBox device placement."""
    device_rack_name = device.rack.name if device.rack_id else ""
    device_rack_location_id = device.rack.location_id if device.rack_id else None
    return (
        device.location_id != source_location_id
        or (device.rack_id is not None and device_rack_location_id != source_location_id)
        or _identity_text(device_rack_name) != _identity_text(rack_name)
        or _normalize_for_compare(device.position) != _normalize_for_compare(position)
        or (face is not None and (device.face or None) != face)
    )


def _rack_location_conflict(rack, location):
    """Return a clear error when a reviewed rack and location cannot coexist."""
    if rack is None:
        return None
    rack_location_id = getattr(rack, "location_id", None)
    location_id = getattr(location, "pk", None)
    if rack_location_id == location_id:
        return None
    rack_label = str(rack.location) if rack_location_id else "(no location)"
    location_label = str(location) if location_id else "(no location)"
    return (
        f"Rack '{rack.name}' belongs to location '{rack_label}', but the effective device "
        f"location is '{location_label}'. Review both rack and location together."
    )


def _identity_state_error(row, source_id, name, object_type, detail, rack_name=""):
    """Return an error when current identity differs from the preview."""
    return RowResult(
        row_number=row["_row_number"],
        source_id=source_id,
        name=name,
        action="error",
        object_type=object_type,
        detail=detail,
        rack_name=rack_name,
        extra_data={"identity_state_changed": True},
    )


def _ambiguous_rack_row(row, source_id, name, rack_name, ctx, object_type):
    """Return an error when a rack name does not identify one rack."""
    label = _rack_identity_label(rack_name, ctx.location)
    return RowResult(
        row_number=row["_row_number"],
        source_id=source_id,
        name=name,
        action="error",
        object_type=object_type,
        detail=f"Multiple racks named '{label}' exist. Select a location or make the rack names unique.",
        rack_name=rack_name,
        extra_data={"identity_conflict": "ambiguous_rack"},
    )


def _device_identity_state(device):
    """Return the device state that an import update can overwrite."""
    return {
        "name": device.name,
        "site_id": device.site_id,
        "location_id": device.location_id,
        "tenant_id": device.tenant_id,
        "rack_id": device.rack_id,
        "position": _normalize_for_compare(device.position),
        "face": device.face or "",
        "device_type_id": device.device_type_id,
        "role_id": device.role_id,
        "status": device.status,
        "serial": device.serial or "",
        "asset_tag": device.asset_tag or "",
        "airflow": device.airflow or "",
    }


def _rack_identity_state(rack):
    """Return the rack state that an import update can overwrite."""
    return {
        "name": rack.name,
        "site_id": rack.site_id,
        "location_id": rack.location_id,
        "tenant_id": rack.tenant_id,
        "u_height": rack.u_height,
        "serial": rack.serial or "",
        "rack_type_id": rack.rack_type_id,
    }


def _intent_matches(ctx, row, object_type, action, object_id=None, current_state=None):
    """Return True when an execute-time identity matches its preview intent."""
    expected = ctx.expected_intents.get((row.get("_row_number"), object_type))
    if expected is None:
        return True
    if expected.get("action") != action:
        return False
    expected_id = expected.get("object_id")
    if expected_id not in (None, "") and expected_id != object_id:
        return False
    expected_state = expected.get("object_state")
    return expected_state is None or expected_state == current_state


def _set_rack_import_fields(rack, u_height, serial, rack_type, ctx):
    """Apply the fields controlled by a rack import."""
    rack.u_height = u_height
    rack.serial = serial or rack.serial
    rack.rack_type = rack_type
    if ctx.location:
        rack.location = ctx.location
    if ctx.tenant:
        rack.tenant = ctx.tenant


# The five fields _set_rack_import_fields writes, so a row that changes none of them writes nothing.
_RACK_IMPORT_FIELDS = ("u_height", "serial", "rack_type_id", "location_id", "tenant_id")


def _ip_already_assigned(device, ip_field, ip_str) -> bool:
    """Return whether the device already carries exactly this address on *ip_field*.

    The writer only assigns after finding an interface of this device that already carries the
    address, and it resolves the IPAddress by that interface's VRF. Anything else it would either
    create or record as unassigned, so only that exact state counts as settled.
    """
    import ipaddress

    from ipam.models import IPAddress

    current = getattr(device, ip_field, None)
    if current is None:
        return False
    try:
        if ipaddress.ip_interface(str(current.address)) != ipaddress.ip_interface(str(ip_str)):
            return False
    except ValueError:
        return False
    same_address = list(IPAddress.objects.filter(address=str(current.address)).values_list("pk", flat=True)[:2])
    if same_address != [current.pk]:
        return False
    interface = current.assigned_object
    return (
        interface is not None
        and getattr(interface, "device_id", None) == device.pk
        and getattr(interface, "vrf_id", None) == current.vrf_id
    )


def _device_binding_is_current(profile, source_id, device, asset_tag) -> bool:
    """Return whether the source-to-device binding this row would write already exists."""
    if not source_id:
        return True
    return profile.device_matches.filter(
        source_id=source_id,
        netbox_device_id=device.pk,
        device_name=device.name,
        source_asset_tag=asset_tag or "",
    ).exists()


def _import_record_is_current(device, profile, source_id, extra_columns) -> bool:
    """Return whether the plugin's import record already holds what this row would store.

    A row that supplies no unassigned address expects an empty map, so a stored one is a change.
    """
    stored = DeviceImportSource.objects.filter(device_id=device.pk).first()
    return stored is not None and (
        stored.profile_id == profile.pk
        and stored.source_id == (source_id or "")
        and stored.extra_columns == (extra_columns or {})
        and not stored.unassigned_ips
    )


def _matched_device_writes_nothing(
    device, review, contact_review, ip_fields, profile, source_id, asset_tag, *, zero_u=False
) -> bool:
    """Return whether updating this matched Device would leave every stored value as it stands.

    The execute guard compares the writer's action to the previewed one, so both sides decide here.
    Every input is read-only, and anything this cannot prove counts as a write.
    """
    # `review` is None when no reviewer loaded, which leaves the field comparison unavailable.
    if review is None or review.differing:
        return False
    # A zero-U type has its position and face cleared whatever the row says, and a row that
    # omits either value is never compared against the stored one.
    if zero_u and (device.position is not None or device.face):
        return False
    plan = contact_review.plan if contact_review is not None else None
    if plan is not None and not (plan["contact_action"] == "reuse" and plan["assignment_action"] == "unchanged"):
        return False
    if any(not _ip_already_assigned(device, ip_field, ip_str) for ip_field, ip_str in (ip_fields or {}).items()):
        return False
    custom_field = profile.adapter_settings.custom_field_name
    if custom_field and source_id and device.custom_field_data.get(custom_field) != source_id:
        return False
    if not _device_binding_is_current(profile, source_id, device, asset_tag):
        return False
    extra_columns = contact_review.extra_columns if contact_review is not None else {}
    return _import_record_is_current(device, profile, source_id, extra_columns)


def _existing_rack_detail(rack_name, action, candidate) -> str:
    """Return the reason an existing rack row reports its action."""
    if action == "update":
        return f"Rack '{rack_name}' already exists"
    if candidate is None:
        return f"Rack '{rack_name}' already exists (update_existing=False)"
    return f"Rack '{rack_name}' already exists and this row changes nothing"


def _rack_import_candidate(rack, u_height, serial, rack_type, ctx):
    """Return the rack as this row would leave it, or None when the profile does not update."""
    if not ctx.profile.adapter_settings.update_existing:
        return None
    candidate = copy(rack)
    _set_rack_import_fields(candidate, u_height, serial, rack_type, ctx)
    return candidate


def _existing_rack_action(rack, candidate) -> str:
    """Return the action an existing rack takes.

    The execute guard compares the writer's action to the previewed one, so both sides decide here.
    """
    if candidate is None:
        return "skip"
    if all(getattr(rack, field) == getattr(candidate, field) for field in _RACK_IMPORT_FIELDS):
        return "skip"
    return "update"


def _build_rack_candidate(Rack, ctx, rack_name, u_height, serial, rack_type):
    """Return an unsaved rack carrying the fields an import controls."""
    return Rack(
        site=ctx.site,
        location=ctx.location,
        name=rack_name,
        tenant=ctx.tenant,
        u_height=u_height,
        serial=serial,
        rack_type=rack_type,
    )


def _rack_validation_error_row(row, source_id, rack_name, exc, operation):
    """Return one rack model validation error."""
    if isinstance(exc, ValidationError) and hasattr(exc, "message_dict"):
        message = "; ".join(f"{field}: {', '.join(errors)}" for field, errors in exc.message_dict.items())
    elif isinstance(exc, ValidationError):
        message = "; ".join(exc.messages)
    else:
        message = str(exc).split("\n")[0]
    return RowResult(
        row_number=row["_row_number"],
        source_id=source_id,
        name=rack_name,
        action="error",
        object_type="rack",
        detail=f"Cannot {operation} rack '{rack_name}': validation failed: {message}",
        rack_name=rack_name,
        extra_data={"identity_conflict": "rack_validation_failed"},
    )


def _write_rack_to_db(rack_name, u_height, serial, source_id, row, ctx, Rack, rack_type=None):
    """Write or update a rack in the database and record the result."""
    rack, ambiguous = _get_unique_rack(Rack, ctx, rack_name, lock=True)
    if ambiguous:
        ctx.result.rows.append(_ambiguous_rack_row(row, source_id, rack_name, rack_name, ctx, "rack"))
        return
    if rack is not None:
        candidate = _rack_import_candidate(rack, u_height, serial, rack_type, ctx)
        action = _existing_rack_action(rack, candidate)
        if not _intent_matches(ctx, row, "rack", action, rack.pk, _rack_identity_state(rack)):
            ctx.result.rows.append(
                _identity_state_error(
                    row,
                    source_id,
                    rack_name,
                    "rack",
                    f"Rack identity changed after preview for '{rack_name}'. Refresh the preview before importing.",
                )
            )
            return
        if action == "update":
            if ctx.user is not None and not Rack.objects.restrict(ctx.user, "change").filter(pk=rack.pk).exists():
                ctx.result.rows.append(_perm_denied_row("dcim.change_rack", row, rack_name, "rack"))
                return
            _set_rack_import_fields(rack, u_height, serial, rack_type, ctx)
            try:
                with transaction.atomic():
                    rack.full_clean()
                    rack.save()
                    _enforce_saved_object_permission(rack, ctx.user, "change")
            except _ObjectPermissionDenied:
                ctx.result.rows.append(_perm_denied_row("dcim.change_rack", row, rack_name, "rack"))
                return
            except (DatabaseError, ValidationError) as exc:
                ctx.result.rows.append(_rack_validation_error_row(row, source_id, rack_name, exc, "update"))
                return
            ctx.rack_map[_identity_text(rack_name)] = rack
            ctx.result.rows.append(
                RowResult(
                    row_number=row["_row_number"],
                    source_id=source_id,
                    name=rack_name,
                    action="update",
                    object_type="rack",
                    detail=f"Updated rack '{rack_name}'",
                    netbox_url=rack.get_absolute_url(),
                )
            )
        else:
            ctx.rack_map[_identity_text(rack_name)] = rack
            ctx.result.rows.append(
                RowResult(
                    row_number=row["_row_number"],
                    source_id=source_id,
                    name=rack_name,
                    action="skip",
                    object_type="rack",
                    detail=_existing_rack_detail(rack_name, action, candidate),
                )
            )
    else:
        if not _intent_matches(ctx, row, "rack", "create"):
            ctx.result.rows.append(
                _identity_state_error(
                    row,
                    source_id,
                    rack_name,
                    "rack",
                    f"Rack identity changed after preview for '{rack_name}'. Refresh the preview before importing.",
                )
            )
            return
        if ctx.user is not None and not ctx.user.has_perm("dcim.add_rack"):
            ctx.result.rows.append(_perm_denied_row("dcim.add_rack", row, rack_name, "rack"))
            return
        try:
            with transaction.atomic():
                rack = _build_rack_candidate(Rack, ctx, rack_name, u_height, serial, rack_type)
                rack.full_clean()
                rack.save()
                _store_source_id(rack, ctx.profile, source_id)
                _enforce_saved_object_permission(rack, ctx.user, "add")
        except _ObjectPermissionDenied:
            ctx.result.rows.append(_perm_denied_row("dcim.add_rack", row, rack_name, "rack"))
            return
        except (DatabaseError, ValidationError) as exc:
            ctx.result.rows.append(_rack_validation_error_row(row, source_id, rack_name, exc, "create"))
            return
        ctx.rack_map[_identity_text(rack_name)] = rack
        ctx.result.rows.append(
            RowResult(
                row_number=row["_row_number"],
                source_id=source_id,
                name=rack_name,
                action="create",
                object_type="rack",
                detail=f"Created rack '{rack_name}' ({u_height}U)",
                netbox_url=rack.get_absolute_url(),
            )
        )


def _pass2_process_racks(rows, ctx, class_role_map):
    """Pass 2: create or update Rack objects. Populates ctx.rack_map in place."""
    from dcim.models import Rack

    for row in rows:
        device_class = _str_val(row.get("device_class"))
        crm = class_role_map.get(device_class)
        if not (crm and crm.creates_rack):
            continue

        rack_name = _str_val(row.get("rack_name")) or _str_val(row.get("device_name"))
        source_id = _str_val(row.get("source_id"))
        u_height_raw = row.get("u_height", 42)
        serial = _str_val(row.get("serial"))

        u_height = max(1, _coerce_int(u_height_raw, 42))

        if not rack_name:
            ctx.result.rows.append(
                RowResult(
                    row_number=row["_row_number"],
                    source_id=source_id,
                    name="",
                    action="error",
                    object_type="rack",
                    detail="Missing rack name",
                )
            )
            continue

        if source_id in ctx.ignored_source_ids:
            ctx.result.rows.append(
                RowResult(
                    row_number=row["_row_number"],
                    source_id=source_id,
                    name=rack_name,
                    action="ignore",
                    object_type="rack",
                    detail="Ignored rack",
                    rack_name=rack_name,
                    extra_data={"ignore_kind": "individual"},
                )
            )
            continue

        if _identity_text(rack_name) in ctx.duplicate_rack_names:
            ctx.result.rows.append(
                RowResult(
                    row_number=row["_row_number"],
                    source_id=source_id,
                    name=rack_name,
                    action="error",
                    object_type="rack",
                    detail=f"Duplicate rack name '{rack_name}' appears more than once in this import.",
                    rack_name=rack_name,
                    extra_data={"identity_conflict": "duplicate_rack"},
                )
            )
            continue

        if source_id and source_id in ctx.duplicate_source_ids:
            other_rows = [
                number for number in ctx.duplicate_source_id_rows.get(source_id, []) if number != row["_row_number"]
            ]
            ctx.result.rows.append(
                RowResult(
                    row_number=row["_row_number"],
                    source_id=source_id,
                    name=rack_name,
                    action="error",
                    object_type="rack",
                    detail=_duplicate_value_detail("source ID", source_id, other_rows),
                    extra_data={
                        "identity_conflict": "duplicate_source_id",
                        "duplicate_source_id_rows": other_rows,
                    },
                )
            )
            continue

        if ctx.dry_run:
            rack_type_label = f", type={crm.rack_type}" if crm.rack_type_id else ""
            rack, ambiguous = _get_unique_rack(Rack, ctx, rack_name, permission_action="view")
            if ambiguous:
                ctx.result.rows.append(_ambiguous_rack_row(row, source_id, rack_name, rack_name, ctx, "rack"))
                continue
            if rack is not None:
                candidate = _rack_import_candidate(rack, u_height, serial, crm.rack_type, ctx)
                if candidate is not None:
                    try:
                        candidate.full_clean()
                    except ValidationError as exc:
                        ctx.result.rows.append(_rack_validation_error_row(row, source_id, rack_name, exc, "update"))
                        continue
                action = _existing_rack_action(rack, candidate)
                detail = _existing_rack_detail(rack_name, action, candidate)
            else:
                candidate = _build_rack_candidate(Rack, ctx, rack_name, u_height, serial, crm.rack_type)
                try:
                    candidate.full_clean()
                except ValidationError as exc:
                    ctx.result.rows.append(_rack_validation_error_row(row, source_id, rack_name, exc, "create"))
                    continue
                rack = None
                action = "create"
                detail = f"Would create rack '{rack_name}' ({u_height}U{rack_type_label}) at site '{ctx.site}'"
            # Keep the unsaved candidate so device rows can be placement-checked against
            # a rack this import has not created yet. `rack` stays None, so the row never
            # reports a NetBox rack ID that does not exist.
            ctx.rack_map[_identity_text(rack_name)] = rack if rack is not None else candidate
            ctx.result.rows.append(
                RowResult(
                    row_number=row["_row_number"],
                    source_id=source_id,
                    name=rack_name,
                    action=action,
                    object_type="rack",
                    detail=detail,
                    extra_data={
                        "source_class": device_class,
                        "rack_type_set": bool(crm.rack_type_id),
                        "rack_type_id": crm.rack_type_id or "",
                        "rack_type_name": str(crm.rack_type) if crm.rack_type_id else "",
                        **({"netbox_rack_id": rack.pk} if rack is not None else {}),
                        **({"_identity_state": _rack_identity_state(rack)} if rack is not None else {}),
                        **({"writes_nothing": True} if action == "skip" and candidate is not None else {}),
                    },
                )
            )
        else:
            _write_rack_to_db(rack_name, u_height, serial, source_id, row, ctx, Rack, rack_type=crm.rack_type)


def _find_existing_device(  # noqa: C901
    profile,
    source_id,
    site,
    device_name,
    serial,
    asset_tag,
    Device,
    ambiguous_names: frozenset = frozenset(),
    *,
    tenant=None,
    device_queryset=None,
    source_match=None,
    source_match_locked=False,
    review_device_ids=frozenset(),
):
    """Look up a pre-existing NetBox device by source-ID link, serial, asset_tag, or name.

    Returns (device, match_method) or (None, None).
    Matching priority: source-ID link → serial → asset_tag → name.
    The name-based fallback is skipped when *device_name* is in *ambiguous_names*
    (i.e. two or more rows in the current import share the same name), since
    matching all of them to the same NetBox device would be incorrect.
    When *site* is provided the name lookup is scoped to that site and tenant.
    Strong identifiers remain global so callers can report an unsafe cross-site
    match instead of silently creating a second identity.
    The optional device queryset controls visibility only. Hidden devices still
    participate in global ambiguity checks.
    """
    devices = Device.objects.select_related(
        "device_type__manufacturer",
        "role",
        "tenant",
        "location",
        "rack__location",
        "site",
    )
    visible_devices = device_queryset
    existing_match = source_match
    if source_id and not source_match_locked:
        existing_match = profile.device_matches.filter(source_id=source_id).first()
    matched_device = None
    match_method = None
    if existing_match:
        try:
            matched_device = devices.get(pk=existing_match.netbox_device_id)
            match_method = "source ID link"
        except Device.DoesNotExist:
            pass

    if matched_device is None and source_id:
        source_matches = list(
            devices.filter(
                data_import_source__profile=profile,
                data_import_source__source_id=source_id,
            )[:2]
        )
        if len(source_matches) == 1:
            matched_device = source_matches[0]
            match_method = "stored source ID"
        elif len(source_matches) > 1:
            logger.warning(
                "Ambiguous stored source ID match for profile=%s source_id=%r; skipping auto-match",
                profile.pk,
                source_id,
            )
            return None, "ambiguous stored source ID"

    if matched_device is None and len(review_device_ids) == 1:
        review_device_id = next(iter(review_device_ids))
        matched_device = devices.filter(pk=review_device_id).first()
        if matched_device is not None:
            match_method = "field review"
    elif matched_device is None and len(review_device_ids) > 1:
        return None, "ambiguous field review"

    if matched_device is None and serial:
        try:
            matched_device = devices.get(serial=serial)
            match_method = "serial"
        except Device.DoesNotExist:
            pass
        except Device.MultipleObjectsReturned:
            logger.warning("Ambiguous serial match for serial=%r; skipping auto-match", serial)
            return None, "ambiguous serial"

    if matched_device is None and asset_tag:
        try:
            matched_device = devices.get(asset_tag__iexact=asset_tag)
            match_method = "asset tag"
        except Device.DoesNotExist:
            pass
        except Device.MultipleObjectsReturned:
            logger.warning("Ambiguous asset_tag match for asset_tag=%r; skipping auto-match", asset_tag)
            return None, "ambiguous asset tag"

    if matched_device is None and device_name and _identity_text(device_name) not in ambiguous_names:
        name_filter = {"name__iexact": device_name}
        if site is not None:
            name_filter["site"] = site
            name_filter.update(_tenant_filter(tenant))
        try:
            matched_device = devices.get(**name_filter)
            match_method = "name"
        except Device.DoesNotExist:
            pass
        except Device.MultipleObjectsReturned:
            logger.warning("Ambiguous name match for device_name=%r; skipping auto-match", device_name)

    if (
        matched_device is not None
        and visible_devices is not None
        and not visible_devices.filter(pk=matched_device.pk).exists()
    ):
        return None, "inaccessible device"

    return matched_device, match_method


def _ambiguous_source_id_row(row, source_id, device_name, rack_name):
    """Return an error for duplicate stored source identity metadata."""
    return RowResult(
        row_number=row["_row_number"],
        source_id=source_id,
        name=device_name,
        action="error",
        object_type="device",
        detail=(
            f"Source ID '{source_id}' is stored on multiple NetBox devices. "
            "Remove the duplicate metadata or link one device explicitly."
        ),
        rack_name=rack_name,
        extra_data={"identity_conflict": "ambiguous_source_id"},
    )


def _ambiguous_field_review_row(row, source_id, device_name, rack_name):
    """Return an error when one source row has multiple stale review devices."""
    return RowResult(
        row_number=row["_row_number"],
        source_id=source_id,
        name=device_name,
        action="error",
        object_type="device",
        detail=(
            f"Source ID '{source_id}' has field reviews for multiple NetBox devices. "
            "Unignore stale reviews or link one device explicitly before importing."
        ),
        rack_name=rack_name,
        extra_data={"identity_conflict": "ambiguous_field_review"},
    )


def _ambiguous_asset_tag_row(row, source_id, device_name, rack_name, asset_tag):
    """Return an error when an asset tag identifies multiple devices."""
    return RowResult(
        row_number=row["_row_number"],
        source_id=source_id,
        name=device_name,
        action="error",
        object_type="device",
        detail=f"Asset tag '{asset_tag}' matches multiple NetBox devices. Make the asset tags unique before import.",
        rack_name=rack_name,
        extra_data={"identity_conflict": "ambiguous_asset_tag"},
    )


def _ambiguous_serial_row(row, source_id, device_name, rack_name, serial):
    """Return an error when a serial identifies multiple devices."""
    return RowResult(
        row_number=row["_row_number"],
        source_id=source_id,
        name=device_name,
        action="error",
        object_type="device",
        detail=f"Serial '{serial}' matches multiple NetBox devices. Make the serials unique before import.",
        rack_name=rack_name,
        extra_data={"identity_conflict": "ambiguous_serial"},
    )


def _device_binding_conflict(profile, source_id, device):
    """Return the source ID already bound to *device*, if it differs."""
    match = profile.device_matches.filter(netbox_device_id=device.pk).exclude(source_id=source_id).first()
    return match.source_id if match else None


def _bind_device_source(profile, source_id, device, asset_tag=""):
    """Persist the confirmed one-to-one source-to-device identity binding."""
    if not source_id:
        return
    conflict_source_id = _device_binding_conflict(profile, source_id, device)
    if conflict_source_id:
        raise _DeviceBindingConflict(f"Device is already bound to source ID '{conflict_source_id}'")
    existing_match = profile.device_matches.select_for_update().filter(source_id=source_id).first()
    if existing_match is not None:
        if existing_match.netbox_device_id != device.pk:
            raise _DeviceBindingConflict(f"Source ID is already bound to device #{existing_match.netbox_device_id}")
        existing_match.device_name = device.name
        existing_match.source_asset_tag = asset_tag or ""
        existing_match.save(update_fields=["device_name", "source_asset_tag"])
        return
    try:
        profile.device_matches.create(
            source_id=source_id,
            netbox_device_id=device.pk,
            device_name=device.name,
            source_asset_tag=asset_tag or "",
        )
    except IntegrityError as exc:
        raise _DeviceBindingConflict("The source-to-device binding changed during import") from exc


def _device_queryset_for_user(Device, user, action):
    """Return devices allowed by NetBox object permissions."""
    if user is None:
        return Device.objects.all()
    return Device.objects.restrict(user, action)


def _reserve_device_names(rows, ctx, class_role_map, Device):
    """Reserve case-insensitive source and NetBox names before suggestions."""
    existing_names = (
        _device_queryset_for_user(Device, ctx.user, "view")
        .filter(site=ctx.site, **_tenant_filter(ctx.tenant))
        .values_list("name", flat=True)
    )
    ctx.reserved_device_names.update(_identity_text(name) for name in existing_names)
    for row in rows:
        crm = class_role_map.get(_str_val(row.get("device_class")))
        if not _is_writing_device_row(row, crm, ctx.ignored_source_ids):
            continue
        name = _effective_device_name(row)
        if name:
            ctx.reserved_device_names.add(_identity_text(name))


def _effective_duplicate_identity_values(rows, ctx, class_role_map, ambiguous_names, Device):
    """Return one review-aware identity write plan per source row."""
    values = {}
    device_queryset = _device_queryset_for_user(Device, ctx.user, "view") if ctx.user is not None else None
    for row in rows:
        crm = class_role_map.get(_str_val(row.get("device_class")))
        if not _is_writing_device_row(row, crm, ctx.ignored_source_ids):
            continue
        source_id = _str_val(row.get("source_id"))
        device_name = _effective_device_name(row)
        serial = _str_val(row.get("serial"))
        asset_tag = (_str_val(row.get("asset_tag")) or "")[:50]
        plan = {"serial": serial or None, "asset_tag": asset_tag or None}
        matched_device, _match_method = _find_existing_device(
            ctx.profile,
            source_id,
            ctx.site,
            device_name,
            serial,
            asset_tag,
            Device,
            ambiguous_names,
            tenant=ctx.tenant,
            device_queryset=device_queryset,
            review_device_ids=(
                ctx.field_reviewer.review_device_ids(source_id) if ctx.field_reviewer is not None else frozenset()
            ),
        )
        if matched_device is not None and ctx.field_reviewer is not None:
            review = ctx.field_reviewer.review(
                source_id,
                matched_device,
                {"serial": serial, "asset_tag": asset_tag},
            )
            effective = review.effective_proposal
            if "serial" not in review.ignored:
                plan["serial"] = _str_val(effective.get("serial", serial)) or None
            else:
                plan["serial"] = None
            if "asset_tag" not in review.ignored:
                plan["asset_tag"] = (_str_val(effective.get("asset_tag", asset_tag)) or "")[:50] or None
            else:
                plan["asset_tag"] = None
        values[row.get("_row_number")] = plan
    return values


def _suggest_unique_device_name(row, ctx):
    """Build and reserve one deterministic, case-insensitive device name."""
    name = _effective_device_name(row)
    rack_name = _str_val(row.get("rack_name")) or "NO-RACK"
    position = _normalize_for_compare(row.get("u_position")) or "NO-U"
    base = f"{name}-{rack_name}-U{position}" if position != "NO-U" else f"{name}-{rack_name}"
    candidate = base[:64]
    if _identity_text(candidate) in ctx.reserved_device_names:
        source_suffix = _str_val(row.get("source_id")) or str(row.get("_row_number", "ROW"))
        source_suffix = re.sub(r"[^A-Za-z0-9_.-]+", "-", source_suffix).strip("-") or "ROW"
        candidate = f"{base[: max(1, 63 - len(source_suffix))]}-{source_suffix}"[:64]
    counter = 2
    unique_candidate = candidate
    while _identity_text(unique_candidate) in ctx.reserved_device_names:
        suffix = f"-{counter}"
        unique_candidate = f"{candidate[: 64 - len(suffix)]}{suffix}"
        counter += 1
    ctx.reserved_device_names.add(_identity_text(unique_candidate))
    return unique_candidate


def _normalize_for_compare(val) -> str:
    """Normalize a value for field-diff comparison.

    Whole-number floats (e.g. 35.0, "35.0") are normalized to their integer
    string form ("35") to avoid false diffs caused by type differences between
    the source file and what NetBox returns.
    """
    if val is None:
        return ""
    try:
        f = float(val)
        if f == int(f):
            return str(int(f))
        return str(f)
    except (TypeError, ValueError, OverflowError):
        return str(val).strip()


def _result_position(value):
    """Return a JSON-safe number so whole and half-U positions stay sortable together."""
    if value is None:
        return None
    number = float(value)
    return int(number) if number.is_integer() else number


_NOT_PROVIDED = object()


def _compute_field_diff(  # noqa: C901
    matched_device,
    device_name,
    serial,
    asset_tag,
    device_face,
    device_airflow,
    device_status,
    u_height,
    u_position,
    *,
    rack_name=_NOT_PROVIDED,
    device_type_key=_NOT_PROVIDED,
    role_slug=_NOT_PROVIDED,
    tenant=_NOT_PROVIDED,
    location=_NOT_PROVIDED,
    ip_fields=None,
):
    """Return a dict of fields that differ between the XLS row and the existing NetBox device."""
    proposal = {
        "device_name": device_name,
        "status": device_status,
        "serial": serial or "",
        "asset_tag": asset_tag or "",
        "u_position": u_position,
        "u_height": u_height,
    }
    display_overrides = {}
    if device_face is not None:
        proposal["face"] = device_face
    if device_airflow is not None:
        proposal["airflow"] = device_airflow
    if rack_name is not _NOT_PROVIDED:
        proposal["rack_name"] = rack_name.rsplit(" / ", 1)[-1] if rack_name else ""
        proposal["_rack_location_id"] = location.pk if location is not _NOT_PROVIDED and location is not None else None
        display_overrides["rack_name"] = _str_val(rack_name)
    if device_type_key is not _NOT_PROVIDED:
        proposal["device_type"] = device_type_key
    if role_slug is not _NOT_PROVIDED:
        proposal["role"] = role_slug
    if tenant is not _NOT_PROVIDED:
        proposal["tenant"] = tenant
    if location is not _NOT_PROVIDED:
        proposal["location"] = location
    proposal.update(ip_fields or {})
    return DeviceFieldReviewer.field_diff(
        matched_device,
        proposal,
        include_informational=True,
        display_overrides=display_overrides,
    )


def _review_device_proposal(
    ctx,
    source_id,
    matched_device,
    *,
    rack_name,
    device_name,
    serial,
    asset_tag,
    device_face,
    device_airflow,
    device_status,
    u_height,
    u_position,
    mfg_slug,
    dt_slug,
    make,
    model,
    role_slug=_NOT_PROVIDED,
    ip_fields=None,
):
    """Review one matched Device proposal and return its effective write values."""
    if ctx.field_reviewer is None:
        return (
            None,
            ctx,
            {
                "rack_name": rack_name,
                "device_name": device_name,
                "serial": serial or "",
                "asset_tag": asset_tag or "",
                "face": device_face,
                "airflow": device_airflow,
                "status": device_status,
                "u_height": u_height,
                "u_position": u_position,
                "device_type": (mfg_slug, dt_slug, make, model),
                "role": role_slug,
                "tenant": ctx.tenant,
                "location": ctx.location,
                **(ip_fields or {}),
            },
        )
    proposal = {
        "rack_name": rack_name,
        "_rack_location_id": ctx.location.pk if ctx.location is not None else None,
        "device_name": device_name,
        "serial": serial or "",
        "asset_tag": asset_tag or "",
        "face": device_face,
        "airflow": device_airflow,
        "status": device_status,
        "u_height": u_height,
        "u_position": u_position,
        "device_type": (mfg_slug, dt_slug, make, model),
        "tenant": ctx.tenant,
        "location": ctx.location,
        **(ip_fields or {}),
    }
    if role_slug is not _NOT_PROVIDED:
        proposal["role"] = role_slug
    review = ctx.field_reviewer.review(
        source_id,
        matched_device,
        proposal,
        display_overrides={"rack_name": _rack_identity_label(rack_name, ctx.location)},
    )
    effective = review.effective_proposal
    effective_ctx = replace(
        ctx,
        location=effective.get("location", ctx.location),
        tenant=effective.get("tenant", ctx.tenant),
    )
    return review, effective_ctx, effective


def _annotate_ip_sync_targets(matched_device, field_diff) -> None:
    """Say where each differing address would land, so the row states it before it is clicked."""
    from . import ip_assignment

    if matched_device is None or not field_diff:
        return
    for ip_field in ip_assignment.IP_FIELD_FAMILY:
        values = field_diff.get(ip_field)
        if not isinstance(values, dict) or not values.get("file"):
            continue
        try:
            values["ip_target"] = ip_assignment.resolve(matched_device, ip_field, values["file"]).placement
        except ip_assignment.IPAssignmentError as exc:
            values["ip_target"] = str(exc)


def _reviewed_rack(review, matched_device):
    """Return the matched rack when a current rack difference is ignored."""
    if review is not None and "rack_name" in review.ignored:
        return matched_device.rack if matched_device.rack_id else None
    return _NOT_PROVIDED


def _is_zero_u_device_type(mfg_slug, dt_slug, dt_exists, DeviceType, cache=None):
    """Return True when the resolved DeviceType has u_height == 0.

    ``cache`` is an optional dict (typically ``ctx.zero_u_cache``) that
    memoizes results by ``(mfg_slug, dt_slug)`` so a large import doesn't
    issue one query per row.
    """
    if not dt_exists:
        return False
    key = (mfg_slug, dt_slug)
    if cache is not None and key in cache:
        return cache[key]
    dt_obj = DeviceType.objects.filter(manufacturer__slug=mfg_slug, slug=dt_slug).only("u_height").first()
    result = dt_obj is not None and dt_obj.u_height == 0
    if cache is not None:
        cache[key] = result
    return result


def _cached_device_type(ctx, DeviceType, mfg_slug, dt_slug):
    """Return one DeviceType lookup result for this import run."""
    key = (mfg_slug, dt_slug)
    if key not in ctx.device_type_cache:
        ctx.device_type_cache[key] = DeviceType.objects.filter(
            manufacturer__slug=mfg_slug,
            slug=dt_slug,
        ).first()
    return ctx.device_type_cache[key]


def _cached_device_role(ctx, DeviceRole, role_slug):
    """Return one DeviceRole lookup result for this import run."""
    if role_slug not in ctx.device_role_cache:
        ctx.device_role_cache[role_slug] = DeviceRole.objects.filter(slug=role_slug).first()
    return ctx.device_role_cache[role_slug]


def _zero_u_overrides(device_type, position, face, ignored_fields=()):
    """Drop position/face when device_type is zero-U (e.g. vertical PDU).

    A reviewed field has already been restored to its current NetBox value.
    Keep that value even when another, non-ignored proposal selects a zero-U
    type. The later Device validation can reject an incompatible combination,
    but the review must never silently rewrite the explicitly ignored field.
    """
    if device_type is not None and device_type.u_height == 0:
        if "u_position" not in ignored_fields:
            position = None
        if "face" not in ignored_fields:
            face = None
    return position, face


ZERO_U_DROPPED_FIELDS = ("u_position", "face")


def _zero_u_review(review, device, zero_u):
    """Return *review* with the fields a zero-U type drops moved out of the differences.

    `_zero_u_overrides` clears the position and the face, so neither reaches NetBox. Reporting
    them as differences offers a sync NetBox refuses and keeps a row that writes nothing off the
    no-op count. A device that still holds either value is left alone, because clearing it is a
    write this cannot describe.
    """
    if review is None or not zero_u or device.position is not None or device.face:
        return review
    moved = {name: review.differing[name] for name in ZERO_U_DROPPED_FIELDS if name in review.differing}
    if not moved:
        return review
    return replace(
        review,
        differing={name: values for name, values in review.differing.items() if name not in moved},
        informational={**review.informational, **moved},
    )


def _zero_u_review_conflict(device_type, position, face, ignored_fields=()):
    """Return reviewed fields that cannot coexist with a zero-U type."""
    if device_type is None or device_type.u_height != 0:
        return ()
    conflicts = []
    if "u_position" in ignored_fields and position is not None:
        conflicts.append("u_position")
    if "face" in ignored_fields and face:
        conflicts.append("face")
    return tuple(conflicts)


def _rack_position_slots(position, height):
    """Return half-U slots occupied by a proposed device placement."""
    from decimal import Decimal

    try:
        start = Decimal(str(position))
    except (TypeError, ValueError, ArithmeticError):
        return ()
    try:
        slot_count = max(1, int(height) * 2)
    except (TypeError, ValueError, OverflowError):
        slot_count = 2
    step = Decimal("0.5")
    return tuple(start + step * offset for offset in range(slot_count))


def _check_rack_position_conflict(
    rack_name,
    position,
    device_face,
    ctx,
    row_number=None,
    device_name=None,
    u_height=1,
):
    """Detect within-file rack position conflicts (incl. multi-U range overlaps).

    A device occupying ``u_height`` units starting at ``position`` covers the
    range ``position .. position + u_height - 1``.  Two rows overlap if any
    slot in their ranges collides on the same ``(rack, face)``.

    Claims are keyed by ``(normalized rack identity, slot, face)`` and store
    ``(row_number, device_name)``. Device identity claims are checked before
    rack placement claims.

    Returns ``(message, conflicting_row_number)`` on the first real overlap,
    otherwise registers each uncovered slot and returns ``None``.
    """
    if not rack_name or position is None:
        return None
    try:
        height = max(1, int(u_height))
    except (TypeError, ValueError):
        height = 1
    effective_face = device_face or ""
    face_label = f" ({effective_face})" if effective_face else ""
    rack_key = _identity_text(rack_name)
    slots = _rack_position_slots(position, height)
    for slot in slots:
        if effective_face:
            claimed_keys = [
                key for key in ((rack_key, slot, effective_face), (rack_key, slot, "")) if key in ctx.claimed_positions
            ]
        else:
            claimed_keys = [
                (rack_key, slot, claimed_face) for claimed_face in ctx.claimed_position_faces.get((rack_key, slot), ())
            ]
        for claimed_key in claimed_keys:
            prev_row, prev_name = ctx.claimed_positions[claimed_key]
            slot_label = _normalize_for_compare(slot)
            if prev_row is not None:
                other = f"row {prev_row}"
                if prev_name:
                    other = f"{other} ('{prev_name}')"
                return (
                    f"Rack position conflict: {rack_name} U{slot_label}{face_label} also claimed by {other}",
                    prev_row,
                )
            return (
                f"Rack position conflict: {rack_name} U{slot_label}{face_label} already claimed by another row in this file",
                None,
            )
    # First claim wins: only register slots that aren't already claimed.
    for slot in slots:
        pos_key = (rack_key, slot, effective_face)
        if pos_key not in ctx.claimed_positions:
            ctx.claimed_positions[pos_key] = (row_number, device_name)
            ctx.claimed_position_faces.setdefault((rack_key, slot), set()).add(effective_face)
    return None


def _claim_rack_slots_for_preview(
    action,
    detail,
    rack_name,
    position,
    device_face,
    ctx,
    row,
    device_name,
    u_height,
):
    """Reserve the row's target U-range for create/update intents during preview.

    Rows that will actually write to the rack — both creates and updates that
    move a matched device into a new ``(rack, face, position, u_height)`` — must
    claim their target slots so later overlapping rows are flagged here instead
    of failing later as an ``IntegrityError`` on write. Skips and ignores make
    no claim.

    Returns ``(action, detail, conflict_row_number, identity_conflict)``.
    """
    if action not in ("create", "update"):
        return action, detail, None, None
    conflict = _check_rack_position_conflict(
        rack_name,
        position,
        device_face,
        ctx,
        row_number=row.get("_row_number"),
        device_name=device_name,
        u_height=u_height,
    )
    if conflict:
        new_detail, conflict_row_number = conflict
        return "error", new_detail, conflict_row_number, "rack_position_occupied"
    return action, detail, None, None


def _resolve_preview_rack(row, ctx, Rack, source_id, device_name, make, model, serial, asset_tag):
    """Resolve the target rack for preview and return an error row when missing."""
    rack_name = _str_val(row.get("rack_name"))
    if not rack_name:
        return None, "(no rack)", None
    cached_rack = ctx.rack_map.get(_identity_text(rack_name))
    if isinstance(cached_rack, Rack):
        return cached_rack, rack_name, None
    if cached_rack:
        return None, rack_name, None
    target_rack, ambiguous = _get_unique_rack(Rack, ctx, rack_name, permission_action="view")
    if ambiguous:
        return None, "", _ambiguous_rack_row(row, source_id, device_name, rack_name, ctx, "device")
    if target_rack is not None:
        ctx.rack_map[_identity_text(rack_name)] = target_rack
        return target_rack, rack_name, None
    error = RowResult(
        row_number=row["_row_number"],
        source_id=source_id,
        name=device_name,
        action="error",
        object_type="device",
        detail=(
            f"Rack '{_rack_identity_label(rack_name, ctx.location)}' was not found. "
            "Select the correct import location or fix the rack name."
        ),
        rack_name=rack_name,
        extra_data={
            "identity_conflict": "rack_not_found",
            "source_make": make,
            "source_model": model,
            "source_serial": serial or "",
            "asset_tag": asset_tag or "",
        },
    )
    return None, "", error


def _name_placement_conflict_row(
    row,
    ctx,
    matched_device,
    source_id,
    device_name,
    rack_name,
    position,
    face,
    serial,
    asset_tag,
):
    """Return an error when a name-only match targets a different placement."""
    netbox_rack = _device_rack_identity_label(matched_device) or "(none)"
    source_rack = _rack_identity_label(rack_name, ctx.location) or "(none)"
    netbox_location = str(matched_device.location) if matched_device.location_id else "(none)"
    source_location = str(ctx.location) if ctx.location is not None else "(none)"
    netbox_position = f" U{matched_device.position}" if matched_device.position is not None else ""
    source_position = f" U{position}" if position is not None else ""
    netbox_face = f" {matched_device.face}" if matched_device.face else ""
    source_face = f" {face}" if face else ""
    placement_differs = _device_placement_differs(
        matched_device,
        ctx.location.pk if ctx.location is not None else None,
        rack_name,
        position,
        face,
    )
    if not placement_differs:
        return None
    return RowResult(
        row_number=row["_row_number"],
        source_id=source_id,
        name=device_name,
        action="error",
        object_type="device",
        detail=(
            f"Device '{matched_device.name}' already exists in NetBox at location '{netbox_location}', "
            f"rack '{netbox_rack}'{netbox_position}{netbox_face}. Source placement is location "
            f"'{source_location}', rack '{source_rack}'{source_position}{source_face}. "
            "Confirm an explicit link to merge it, or use the suggested unique name."
        ),
        netbox_url=matched_device.get_absolute_url(),
        rack_name=rack_name,
        extra_data={
            "identity_conflict": "name_placement_conflict",
            "suggested_name": _suggest_unique_device_name(row, ctx),
            "source_rack_name": source_rack,
            "source_location": source_location,
            "source_position": _normalize_for_compare(position),
            "source_face": face or "",
            "netbox_device_id": matched_device.pk,
            "netbox_device_name": matched_device.name,
            "netbox_rack_name": netbox_rack,
            "netbox_location": netbox_location,
            "netbox_position": _normalize_for_compare(matched_device.position),
            "netbox_face": matched_device.face or "",
            "source_serial": serial or "",
            "asset_tag": asset_tag or "",
        },
    )


def _validate_preview_placement(
    action,
    detail,
    rack_name,
    target_rack,
    position,
    face,
    device_type,
    proposed_u_height,
    proposed_full_depth,
    matched_device,
    ctx,
):
    """Validate required placement fields and existing rack capacity."""
    effective_face = face
    if effective_face is None and matched_device is not None and action == "update":
        effective_face = matched_device.face or None
    if action not in ("create", "update"):
        return action, detail, None, effective_face
    if position is not None and not rack_name:
        return "error", "A rack is required when a device has a rack position.", "rack_required", effective_face
    if position is not None and effective_face is None:
        return (
            "error",
            "A rack face is required when a device has a rack position.",
            "rack_face_required",
            effective_face,
        )
    placement_height = device_type.u_height if device_type is not None else proposed_u_height
    placement_full_depth = device_type.is_full_depth if device_type is not None else proposed_full_depth
    if target_rack is None or position is None or placement_height == 0:
        return action, detail, None, effective_face
    rack_face = None if placement_full_depth else effective_face
    excluded_ids = [matched_device.pk] if matched_device is not None else []
    if target_rack.pk is None:
        # A rack this import has not created yet holds no devices, so only its own unit
        # skeleton limits placement. get_available_units() cannot run on an unsaved rack.
        from decimal import Decimal

        from utilities.data import drange

        start = Decimal(str(position))
        occupied = set(drange(start, start + Decimal(str(placement_height)), Decimal("0.5")))
        fits = occupied <= set(target_rack.units)
    else:
        fits = position in target_rack.get_available_units(
            u_height=placement_height,
            rack_face=rack_face,
            exclude=excluded_ids,
        )
    if fits:
        return action, detail, None, effective_face
    detail = (
        f"Rack position {_rack_identity_label(rack_name, ctx.location)} U{position} "
        f"does not have {placement_height}U of available space."
    )
    return "error", detail, "rack_position_occupied", effective_face


def _preview_device_row(  # noqa: C901
    row,
    ctx,
    make,
    model,
    mfg_slug,
    dt_slug,
    source_id,
    device_name,
    serial,
    asset_tag,
    DeviceType,
    Device,
    Rack,
    ip_fields: dict | None = None,
    device_face=None,
    device_airflow=None,
    device_status="active",
    u_position=None,
    is_explicit_mapping: bool = False,
    ambiguous_names: frozenset = frozenset(),
    role_slug=_NOT_PROVIDED,
):
    """Return a RowResult for *dry_run* mode (no DB writes)."""
    # Parse u_height early so it's available in all return paths
    u_height_raw = row.get("u_height", 1)
    review_u_height = _coerce_int(u_height_raw, 1)
    u_height = max(1, review_u_height)

    device_type = _cached_device_type(ctx, DeviceType, mfg_slug, dt_slug)
    dt_exists = device_type is not None
    is_zero_u = _is_zero_u_device_type(mfg_slug, dt_slug, dt_exists, DeviceType, ctx.zero_u_cache)
    from dcim.models import DeviceRole

    source_device_role = _cached_device_role(ctx, DeviceRole, role_slug) if role_slug is not _NOT_PROVIDED else None
    source_role_missing = role_slug is not _NOT_PROVIDED and source_device_role is None
    source_type_missing = device_type is None

    rack_name = _str_val(row.get("rack_name"))
    target_rack = None
    rack_label = rack_name or "(no rack)"
    rack_error = None
    # Re-derive position for display label; u_position param is the pre-resolved value for field_diff
    position = _coerce_position(row.get("u_position"))

    # Keep the source placement available for review. Apply zero-U semantics only
    # after the effective proposal restores any ignored placement fields.
    # _find_existing_device checks DeviceExistingMatch → serial → asset_tag → name in that order,
    # ensuring explicit operator mappings always take precedence over coincidental name matches.
    matched_device, match_method = _find_existing_device(
        ctx.profile,
        source_id,
        ctx.site,
        device_name,
        serial,
        asset_tag,
        Device,
        ambiguous_names,
        tenant=ctx.tenant,
        device_queryset=_device_queryset_for_user(Device, ctx.user, "view") if ctx.user is not None else None,
        review_device_ids=(
            ctx.field_reviewer.review_device_ids(source_id) if ctx.field_reviewer is not None else frozenset()
        ),
    )
    if match_method == "inaccessible device":
        return _perm_denied_row("dcim.view_device", row, device_name, "device")
    if match_method == "ambiguous serial":
        return _ambiguous_serial_row(row, source_id, device_name, rack_name, serial)
    if match_method == "ambiguous asset tag":
        return _ambiguous_asset_tag_row(row, source_id, device_name, rack_name, asset_tag)
    if match_method == "ambiguous stored source ID":
        return _ambiguous_source_id_row(row, source_id, device_name, rack_name)
    if match_method == "ambiguous field review":
        return _ambiguous_field_review_row(row, source_id, device_name, rack_name)
    conflict_row_number = None
    review = None
    review_ctx = ctx
    relation_error = None
    relation_identity_conflict = None
    placement_identity_conflict = None
    placement_error_extra = {}
    if matched_device is not None:
        if matched_device.site_id != ctx.site.pk:
            return RowResult(
                row_number=row["_row_number"],
                source_id=source_id,
                name=device_name,
                action="error",
                object_type="device",
                detail=(
                    f"The {match_method} matches device '{matched_device.name}' at site "
                    f"'{matched_device.site}'. The active import site is '{ctx.site}'. Link a device in the active site."
                ),
                rack_name=rack_name,
                extra_data={
                    "identity_conflict": "cross_site_match",
                    "netbox_device_id": matched_device.pk,
                    "source_serial": serial or "",
                    "asset_tag": asset_tag or "",
                },
            )
        review, review_ctx, effective = _review_device_proposal(
            ctx,
            source_id,
            matched_device,
            rack_name=rack_name,
            device_name=device_name,
            serial=serial,
            asset_tag=asset_tag,
            device_face=device_face,
            device_airflow=device_airflow,
            device_status=device_status,
            u_height=review_u_height,
            u_position=position,
            mfg_slug=mfg_slug,
            dt_slug=dt_slug,
            make=make,
            model=model,
            role_slug=role_slug,
            ip_fields=ip_fields,
        )
        rack_name = effective["rack_name"]
        serial = effective["serial"]
        asset_tag = effective["asset_tag"]
        device_face = effective["face"]
        device_airflow = effective["airflow"]
        device_status = effective["status"]
        position = effective["u_position"]
        if role_slug is not _NOT_PROVIDED:
            role_slug = effective["role"]
        effective_type = effective["device_type"]
        mfg_slug, dt_slug = effective_type[:2]
        if review is not None and "device_type" in review.ignored:
            device_type = _cached_device_type(ctx, DeviceType, mfg_slug, dt_slug)
            dt_exists = device_type is not None
            source_type_missing = False
        can_create_type = ctx.profile.adapter_settings.create_missing_device_types and (
            ctx.user is None or ctx.user.has_perm("dcim.add_devicetype")
        )
        if source_type_missing and not can_create_type and not (review is not None and "device_type" in review.ignored):
            relation_error = f"Device type not found: {make} / {model} (slug: {mfg_slug}/{dt_slug})"
            relation_identity_conflict = "device_type_not_found"
        role_is_ignored = review is not None and "role" in review.ignored
        if source_role_missing and not role_is_ignored:
            can_create_role = ctx.user is None or ctx.user.has_perm("dcim.add_devicerole")
            if not can_create_role:
                relation_error = f"Device role not found: {role_slug}"
                relation_identity_conflict = "device_role_not_found"
        is_zero_u = _is_zero_u_device_type(mfg_slug, dt_slug, dt_exists, DeviceType, ctx.zero_u_cache)
        position, device_face = _zero_u_overrides(
            device_type,
            position,
            device_face,
            review.ignored if review is not None else (),
        )
        review = _zero_u_review(review, matched_device, is_zero_u)
        zero_u_conflict = _zero_u_review_conflict(
            device_type,
            position,
            device_face,
            review.ignored if review is not None else (),
        )
        conflict_source_id = _device_binding_conflict(ctx.profile, source_id, matched_device)
        previous_claim = ctx.claimed_device_ids.get(matched_device.pk)
        if conflict_source_id or (previous_claim and previous_claim[0] != row.get("_row_number")):
            claimed_source = conflict_source_id or previous_claim[1] or f"row {previous_claim[0]}"
            return RowResult(
                row_number=row["_row_number"],
                source_id=source_id,
                name=device_name,
                action="error",
                object_type="device",
                detail=(
                    f"Device '{matched_device.name}' is already bound to source ID "
                    f"'{claimed_source}'. One NetBox device cannot represent two source rows."
                ),
                rack_name=rack_name,
                extra_data={
                    "identity_conflict": "device_already_bound",
                    "netbox_device_id": matched_device.pk,
                },
            )
        ctx.claimed_device_ids[matched_device.pk] = (row.get("_row_number"), source_id)
        action = "update" if ctx.profile.adapter_settings.update_existing else "skip"
        if relation_error is not None:
            action = "error"
            detail = relation_error
        if zero_u_conflict:
            action = "error"
            detail = (
                f"Cannot apply ignored {', '.join(zero_u_conflict)} with zero-U device type "
                f"'{device_type}'. Unignore those fields or ignore the device type."
            )
        if match_method == "name" and relation_error is None:
            netbox_rack = _device_rack_identity_label(matched_device) or "(none)"
            source_rack = _rack_identity_label(rack_name, ctx.location) or "(none)"
            netbox_position = f" U{matched_device.position}" if matched_device.position is not None else ""
            source_position = f" U{position}" if position is not None else ""
            placement_error = _name_placement_conflict_row(
                row,
                review_ctx,
                matched_device,
                source_id,
                device_name,
                rack_name,
                position,
                device_face,
                serial,
                asset_tag,
            )
            if placement_error is not None:
                action = "error"
                detail = placement_error.detail
                placement_identity_conflict = placement_error.extra_data.get("identity_conflict")
                placement_error_extra = placement_error.extra_data
            else:
                detail = (
                    f"Device '{device_name}' already exists in NetBox at rack '{netbox_rack}'{netbox_position}. "
                    f"Source placement is rack '{source_rack}'{source_position}."
                )
        else:
            # Clarify what happens to name: it is NOT updated on matched devices
            name_note = ""
            if matched_device.name != device_name:
                name_note = f"; name stays '{matched_device.name}' (source: '{device_name}')"
            else:
                name_note = "; name unchanged"
            if ctx.profile.adapter_settings.update_existing and not zero_u_conflict and relation_error is None:
                detail = f"Will update '{matched_device.name}' (matched by {match_method}{name_note})"
            elif not zero_u_conflict and relation_error is None:
                detail = (
                    f"Matched to '{matched_device.name}' (by {match_method}{name_note}, skip — update_existing off)"
                )
    else:
        if source_type_missing and not ctx.profile.adapter_settings.create_missing_device_types:
            return RowResult(
                row_number=row["_row_number"],
                source_id=source_id,
                name=device_name,
                action="error",
                object_type="device",
                detail=f"Device type not found: {make} / {model} (slug: {mfg_slug}/{dt_slug})",
                extra_data={
                    "source_make": make,
                    "source_model": model,
                    "mfg_slug": mfg_slug,
                    "dt_slug": dt_slug,
                    "u_height": u_height,
                    "face": device_face or "",
                    "airflow": device_airflow or "",
                    "status": device_status,
                    "asset_tag": asset_tag or "",
                    "source_serial": serial or "",
                    "is_explicit_mapping": is_explicit_mapping,
                    "dt_exists": dt_exists,
                    "extra_columns": row.get("_extra_columns", {}),
                    "conflicts": row.get("_conflicts", {}),
                    **({"_ip": ip_fields} if ip_fields else {}),
                },
            )
        if source_role_missing and ctx.user is not None and not ctx.user.has_perm("dcim.add_devicerole"):
            return RowResult(
                row_number=row["_row_number"],
                source_id=source_id,
                name=device_name,
                action="error",
                object_type="device",
                detail=f"Device role not found: {role_slug}",
                extra_data={"identity_conflict": "device_role_not_found"},
            )
        position, device_face = _zero_u_overrides(device_type, position, device_face)
        action = "create"
        if is_zero_u:
            detail = f"Would create device '{device_name}' in {rack_label} (zero-U)"
        else:
            _pos_label = f" U{position}" if position is not None else ""
            detail = f"Would create device '{device_name}' in {rack_label}{_pos_label}"

    field_diff: dict | None = None
    field_ignored: dict | None = None
    field_informational: dict | None = None
    field_non_writable: dict | None = None
    field_snapshots: dict | None = None
    if matched_device is not None and action != "skip":
        if review is None:
            field_diff = _compute_field_diff(
                matched_device,
                device_name,
                serial,
                asset_tag,
                device_face,
                device_airflow,
                device_status,
                review_u_height,
                position,
                rack_name=_rack_identity_label(rack_name, review_ctx.location),
                device_type_key=(mfg_slug, dt_slug, make, model),
                role_slug=role_slug,
                tenant=review_ctx.tenant,
                location=review_ctx.location,
                ip_fields=ip_fields,
            )
        else:
            field_diff = {**review.differing, **review.informational}
            field_ignored = review.ignored
            field_informational = review.informational
            field_non_writable = {
                field_name: True
                for field_name in review.snapshots
                if field_name in DeviceFieldReviewer.non_writable_fields()
            }
            field_snapshots = {
                field_name: {"file": file_snapshot, "netbox": netbox_snapshot}
                for field_name, (file_snapshot, netbox_snapshot) in review.snapshots.items()
            }

    _annotate_ip_sync_targets(matched_device, field_diff)
    reviewed_rack = _reviewed_rack(review, matched_device)
    rack_error_extra = {}
    if reviewed_rack is not _NOT_PROVIDED:
        target_rack = reviewed_rack
        rack_label = _device_rack_identity_label(matched_device) or "(no rack)"
    else:
        effective_row = dict(row)
        effective_row["rack_name"] = rack_name
        rack_lookup_ctx = ctx
        target_rack, rack_label, rack_error = _resolve_preview_rack(
            effective_row,
            rack_lookup_ctx,
            Rack,
            source_id,
            device_name,
            make,
            model,
            serial,
            asset_tag,
        )
    if rack_error is not None:
        if matched_device is None:
            return rack_error
        action = "error"
        detail = rack_error.detail
        rack_error_extra = rack_error.extra_data
        relation_identity_conflict = relation_identity_conflict or rack_error_extra.get("identity_conflict")

    rack_location_conflict = _rack_location_conflict(target_rack, review_ctx.location)
    if rack_location_conflict:
        action = "error"
        detail = rack_location_conflict
        placement_identity_conflict = "rack_location_conflict"

    pending_full_depth = DeviceType._meta.get_field("is_full_depth").get_default()
    action, detail, identity_conflict, effective_face = _validate_preview_placement(
        action,
        detail,
        rack_name,
        target_rack,
        position,
        device_face,
        device_type,
        u_height,
        pending_full_depth,
        matched_device,
        review_ctx,
    )
    if placement_identity_conflict:
        identity_conflict = placement_identity_conflict
    if relation_identity_conflict:
        identity_conflict = relation_identity_conflict
    placement_height = device_type.u_height if device_type is not None else u_height
    placement_full_depth = device_type.is_full_depth if device_type is not None else pending_full_depth
    placement_face = None if placement_full_depth else effective_face
    placement_rack_label = (
        _device_rack_identity_label(matched_device)
        if reviewed_rack is not _NOT_PROVIDED
        else _rack_identity_label(rack_name, ctx.location)
    )
    action, detail, conflict_row_number, claim_conflict = _claim_rack_slots_for_preview(
        action,
        detail,
        placement_rack_label,
        position,
        placement_face,
        review_ctx,
        row,
        device_name,
        placement_height,
    )
    if claim_conflict is not None:
        identity_conflict = claim_conflict

    primary_contact_plan = None
    contact_suggestion = None
    contact_review = None
    if action in ("create", "update"):
        try:
            contact_review = PrimaryContactResolver.review(
                matched_device,
                row,
                review_ctx.profile,
                ctx.user,
                candidate_source_columns=review_ctx.candidate_source_columns,
            )
            primary_contact_plan = contact_review.plan
            contact_suggestion = contact_review.suggestion
        except _ObjectPermissionDenied as exc:
            return _perm_denied_row(str(exc), row, device_name, "device")
        except ValidationError as exc:
            error_row = _rack_position_error_row(
                row,
                source_id,
                device_name,
                make,
                model,
                asset_tag,
                rack_name,
                position,
                exc,
                action,
            )
            if matched_device is not None:
                error_row.netbox_url = matched_device.get_absolute_url()
                error_row.extra_data["netbox_device_id"] = matched_device.pk
                error_row.extra_data["_identity_state"] = _device_identity_state(matched_device)
            return error_row

    writes_nothing = action == "update" and _matched_device_writes_nothing(
        matched_device,
        review,
        contact_review,
        ip_fields,
        review_ctx.profile,
        source_id,
        asset_tag,
        zero_u=is_zero_u,
    )
    if writes_nothing:
        action = "skip"
        detail = f"Device '{matched_device.name}' matches this row, which writes nothing"

    return RowResult(
        row_number=row["_row_number"],
        source_id=source_id,
        name=device_name,
        action=action,
        object_type="device",
        detail=detail,
        netbox_url=matched_device.get_absolute_url() if matched_device is not None else "",
        rack_name=rack_name,
        extra_data={
            "source_make": make,
            "source_model": model,
            "mfg_slug": mfg_slug,
            "dt_slug": dt_slug,
            "u_height": u_height,
            "u_position": _result_position(position),
            "face": device_face or "",
            "airflow": device_airflow or "",
            "status": device_status,
            "asset_tag": asset_tag or "",
            "source_serial": serial or "",
            "is_explicit_mapping": is_explicit_mapping,
            "dt_exists": dt_exists,
            "extra_columns": row.get("_extra_columns", {}),
            "conflicts": row.get("_conflicts", {}),
            **({"zero_u": True} if is_zero_u else {}),
            **({"writes_nothing": True} if writes_nothing else {}),
            **rack_error_extra,
            **placement_error_extra,
            **({"candidate_values": row["_candidate_values"]} if row.get("_candidate_values") else {}),
            **({"_ip": ip_fields} if ip_fields else {}),
            **({"field_diff": field_diff} if field_diff is not None else {}),
            **({"field_ignored": field_ignored} if field_ignored else {}),
            **({"field_informational": field_informational} if field_informational else {}),
            **({"field_non_writable": field_non_writable} if field_non_writable else {}),
            **({"field_review_snapshots": field_snapshots} if field_snapshots else {}),
            **({"primary_contact_plan": primary_contact_plan} if primary_contact_plan is not None else {}),
            **({"contact_suggestion": contact_suggestion} if contact_suggestion is not None else {}),
            **({"identity_conflict": identity_conflict} if identity_conflict else {}),
            **(
                {"netbox_device_id": matched_device.pk}
                if matched_device is not None and (action != "skip" or writes_nothing)
                else {}
            ),
            **({"_identity_state": _device_identity_state(matched_device)} if matched_device is not None else {}),
            **(
                {
                    "netbox_rack_name": matched_device.rack.name if matched_device.rack_id else "",
                    "netbox_position": _normalize_for_compare(matched_device.position),
                    "netbox_face": matched_device.face or "",
                    # Only set when a placement sync has nothing to write, so an older preview
                    # that predates this key keeps offering the action.
                    **(
                        {"placement_sync_writes_nothing": True}
                        if placement_sync_is_noop(matched_device, rack_name, position, device_face)
                        else {}
                    ),
                }
                if matched_device is not None
                else {}
            ),
            **({"conflict_row_number": conflict_row_number} if action == "error" and conflict_row_number else {}),
        },
    )


def _rack_position_error_row(row, source_id, device_name, make, model, asset_tag, rack_name, position, exc, action):
    """Convert a database or model validation failure to an error row."""
    _rack_label = rack_name if rack_name else "(no rack)"
    _pos_label = f" U{position}" if position is not None else ""
    if isinstance(exc, _DeviceBindingConflict):
        msg = str(exc)
    elif isinstance(exc, ValidationError):
        if hasattr(exc, "message_dict"):
            msg = "; ".join(f"{field}: {', '.join(errors)}" for field, errors in exc.message_dict.items())
        else:
            msg = "; ".join(exc.messages)
    else:
        msg = str(exc).split("\n")[0]
    if isinstance(exc, _DeviceBindingConflict):
        detail = f"Cannot {action} '{device_name}': {msg}"
        identity_conflict = "device_already_bound"
    elif isinstance(exc, _CandidateResolutionRequired):
        detail = msg
        identity_conflict = "candidate_resolution_required"
    elif "unique_rack_position" in msg or ("rack_id" in msg and "position" in msg) or "occupied" in msg:
        detail = f"Cannot {action} '{device_name}': rack position {_rack_label}{_pos_label} is already occupied"
        identity_conflict = "rack_position_occupied"
    else:
        detail = f"Cannot {action} '{device_name}': validation failed: {msg}"
        identity_conflict = "device_validation_failed"
    return RowResult(
        row_number=row["_row_number"],
        source_id=source_id,
        name=device_name,
        action="error",
        object_type="device",
        detail=detail,
        rack_name=rack_name,
        extra_data={
            "source_make": make,
            "source_model": model,
            "asset_tag": asset_tag or "",
            "identity_conflict": identity_conflict,
            **(
                {"candidate_values": {exc.candidate_target: exc.candidate_values}}
                if isinstance(exc, _CandidateResolutionRequired)
                else {}
            ),
            **(
                {"contact_suggestion": exc.suggestion}
                if isinstance(exc, _CandidateResolutionRequired) and exc.suggestion is not None
                else {}
            ),
        },
    )


def _write_device_row(  # noqa: C901
    row,
    ctx,
    make,
    model,
    crm,
    mfg_slug,
    dt_slug,
    source_id,
    device_name,
    serial,
    asset_tag,
    position,
    face,
    airflow,
    status,
    DeviceType,
    DeviceRole,
    Rack,
    Device,
    ip_fields: dict | None = None,
    ambiguous_names: frozenset = frozenset(),
):
    rack_name = _str_val(row.get("rack_name"))
    device_type = _cached_device_type(ctx, DeviceType, mfg_slug, dt_slug)
    device_role = _cached_device_role(ctx, DeviceRole, crm.role_slug)
    source_type_missing = device_type is None
    source_role_missing = device_role is None

    rack = None

    # _find_existing_device checks DeviceExistingMatch → serial → asset_tag → name in that order,
    # ensuring explicit operator mappings always take precedence over coincidental name matches.
    locked_source_match = None
    if source_id:
        locked_source_match = ctx.profile.device_matches.select_for_update().filter(source_id=source_id).first()
    resolve_identity = partial(
        _find_existing_device,
        ctx.profile,
        source_id,
        ctx.site,
        device_name,
        serial,
        asset_tag,
        Device,
        ambiguous_names,
        tenant=ctx.tenant,
        device_queryset=_device_queryset_for_user(Device, ctx.user, "view") if ctx.user is not None else None,
        source_match=locked_source_match,
        source_match_locked=bool(source_id),
        review_device_ids=(
            ctx.field_reviewer.review_device_ids(source_id) if ctx.field_reviewer is not None else frozenset()
        ),
    )
    device, match_method = resolve_identity()
    if match_method == "inaccessible device":
        return _perm_denied_row("dcim.view_device", row, device_name, "device")
    if match_method == "ambiguous serial":
        return _ambiguous_serial_row(row, source_id, device_name, rack_name, serial)
    if match_method == "ambiguous asset tag":
        return _ambiguous_asset_tag_row(row, source_id, device_name, rack_name, asset_tag)
    if match_method == "ambiguous stored source ID":
        return _ambiguous_source_id_row(row, source_id, device_name, rack_name)
    if match_method == "ambiguous field review":
        return _ambiguous_field_review_row(row, source_id, device_name, rack_name)

    if device is None and source_type_missing:
        return RowResult(
            row_number=row["_row_number"],
            source_id=source_id,
            name=device_name,
            action="error",
            object_type="device",
            detail=f"Device type not found: {mfg_slug}/{dt_slug}",
        )
    if device is None and source_role_missing:
        return RowResult(
            row_number=row["_row_number"],
            source_id=source_id,
            name=device_name,
            action="error",
            object_type="device",
            detail=f"Device role not found: {crm.role_slug}",
        )

    write_ctx = ctx
    if device is not None:
        try:
            device = Device.objects.select_for_update(of=("self",)).select_related("rack__location").get(pk=device.pk)
        except Device.DoesNotExist:
            return _identity_state_error(
                row,
                source_id,
                device_name,
                "device",
                f"Device identity changed after preview for '{device_name}'. Refresh the preview before importing.",
                rack_name,
            )
        # The lock covers one row, so re-run the predicate: a device written between
        # resolution and lock can make serial or asset tag ambiguous.
        recheck_device, recheck_method = resolve_identity()
        if recheck_device is None or recheck_device.pk != device.pk or recheck_method != match_method:
            return _identity_state_error(
                row,
                source_id,
                device_name,
                "device",
                f"Device identity changed after preview for '{device_name}'. Refresh the preview before importing.",
                rack_name,
            )
        review, write_ctx, effective = _review_device_proposal(
            ctx,
            source_id,
            device,
            rack_name=rack_name,
            device_name=device_name,
            serial=serial,
            asset_tag=asset_tag,
            device_face=face,
            device_airflow=airflow,
            device_status=status,
            u_height=_coerce_int(row.get("u_height"), 1),
            u_position=position,
            mfg_slug=mfg_slug,
            dt_slug=dt_slug,
            make=make,
            model=model,
            role_slug=crm.role_slug,
            ip_fields=ip_fields,
        )
        rack_name = effective["rack_name"]
        serial = effective["serial"]
        asset_tag = effective["asset_tag"]
        face = effective["face"]
        airflow = effective["airflow"]
        status = effective["status"]
        position = effective["u_position"]
        effective_type = effective["device_type"]
        mfg_slug, dt_slug = effective_type[:2]
        role_slug = effective["role"]
        # An ignored difference means leave the field alone; the writer reads these separately.
        if review is not None:
            ip_fields = {name: value for name, value in (ip_fields or {}).items() if name not in review.ignored}
        if review is not None and "device_type" in review.ignored:
            device_type = _cached_device_type(ctx, DeviceType, mfg_slug, dt_slug)
            if device_type is None:
                return _identity_state_error(
                    row,
                    source_id,
                    device_name,
                    "device",
                    f"Device type changed after preview for '{device_name}'. Refresh the preview before importing.",
                    rack_name,
                )
        elif source_type_missing:
            return RowResult(
                row_number=row["_row_number"],
                source_id=source_id,
                name=device_name,
                action="error",
                object_type="device",
                detail=f"Device type not found: {mfg_slug}/{dt_slug}",
                rack_name=rack_name,
                extra_data={"identity_conflict": "device_type_not_found", "netbox_device_id": device.pk},
            )
        if review is not None and "role" in review.ignored:
            device_role = _cached_device_role(ctx, DeviceRole, role_slug)
            if device_role is None:
                return _identity_state_error(
                    row,
                    source_id,
                    device_name,
                    "device",
                    f"Device role changed after preview for '{device_name}'. Refresh the preview before importing.",
                    rack_name,
                )
        elif source_role_missing:
            return RowResult(
                row_number=row["_row_number"],
                source_id=source_id,
                name=device_name,
                action="error",
                object_type="device",
                detail=f"Device role not found: {role_slug}",
                rack_name=rack_name,
                extra_data={"identity_conflict": "device_role_not_found", "netbox_device_id": device.pk},
            )
        position, face = _zero_u_overrides(
            device_type,
            position,
            face,
            review.ignored if review is not None else (),
        )
        # The intent guard compares this action to the previewed one, so both drop the same fields.
        review = _zero_u_review(review, device, device_type is not None and device_type.u_height == 0)
        zero_u_conflict = _zero_u_review_conflict(
            device_type,
            position,
            face,
            review.ignored if review is not None else (),
        )
        if zero_u_conflict:
            return RowResult(
                row_number=row["_row_number"],
                source_id=source_id,
                name=device_name,
                action="error",
                object_type="device",
                detail=(
                    f"Cannot apply ignored {', '.join(zero_u_conflict)} with zero-U device type "
                    f"'{device_type}'. Unignore those fields or ignore the device type."
                ),
                rack_name=rack_name,
                extra_data={
                    "identity_conflict": "zero_u_review_conflict",
                    "netbox_device_id": device.pk,
                },
            )
        reviewed_rack = _reviewed_rack(review, device)
        rack_lookup_ctx = ctx if reviewed_rack is _NOT_PROVIDED else write_ctx
        if reviewed_rack is not _NOT_PROVIDED:
            rack = reviewed_rack
        else:
            rack = rack_lookup_ctx.rack_map.get(_identity_text(rack_name)) if rack_name else None
            if rack_name and rack is None:
                rack, ambiguous = _get_unique_rack(
                    Rack, rack_lookup_ctx, rack_name, lock=True, permission_action="view"
                )
                if ambiguous:
                    return _ambiguous_rack_row(row, source_id, device_name, rack_name, rack_lookup_ctx, "device")
            if rack_name and not isinstance(rack, Rack):
                return RowResult(
                    row_number=row["_row_number"],
                    source_id=source_id,
                    name=device_name,
                    action="error",
                    object_type="device",
                    detail=(
                        f"Rack '{_rack_identity_label(rack_name, rack_lookup_ctx.location)}' was not found. "
                        "Select the correct import location or fix the rack name."
                    ),
                    rack_name=rack_name,
                    extra_data={"identity_conflict": "rack_not_found"},
                )
        rack_location_conflict = _rack_location_conflict(rack, write_ctx.location)
        if rack_location_conflict:
            return RowResult(
                row_number=row["_row_number"],
                source_id=source_id,
                name=device_name,
                action="error",
                object_type="device",
                detail=rack_location_conflict,
                rack_name=rack_name,
                extra_data={
                    "identity_conflict": "rack_location_conflict",
                    "netbox_device_id": device.pk,
                },
            )
        if device.site_id != ctx.site.pk:
            return RowResult(
                row_number=row["_row_number"],
                source_id=source_id,
                name=device_name,
                action="error",
                object_type="device",
                detail=(
                    f"The {match_method} matches device '{device.name}' at site '{device.site}'. "
                    f"The active import site is '{ctx.site}'. Link a device in the active site."
                ),
                rack_name=rack_name,
                extra_data={"identity_conflict": "cross_site_match", "netbox_device_id": device.pk},
            )
        if match_method == "name":
            placement_error = _name_placement_conflict_row(
                row,
                write_ctx,
                device,
                source_id,
                device_name,
                rack_name,
                position,
                face,
                serial,
                asset_tag,
            )
            if placement_error is not None:
                return placement_error
        contact_review = None
        actual_action = "update" if ctx.profile.adapter_settings.update_existing else "skip"
        if actual_action == "update":
            # Read-only, and needed here because the guard below compares this action to the preview.
            try:
                contact_review = PrimaryContactResolver.review(
                    device,
                    row,
                    ctx.profile,
                    ctx.user,
                    candidate_source_columns=ctx.candidate_source_columns,
                )
            except _ObjectPermissionDenied as exc:
                return _perm_denied_row(str(exc) or "dcim.change_device", row, device_name, "device")
            except (DatabaseError, ValidationError) as exc:
                return _rack_position_error_row(
                    row, source_id, device_name, make, model, asset_tag, rack_name, position, exc, "update"
                )
            if _matched_device_writes_nothing(
                device,
                review,
                contact_review,
                ip_fields,
                ctx.profile,
                source_id,
                asset_tag,
                zero_u=device_type is not None and device_type.u_height == 0,
            ):
                actual_action = "skip"
        if not _intent_matches(ctx, row, "device", actual_action, device.pk, _device_identity_state(device)):
            return _identity_state_error(
                row,
                source_id,
                device_name,
                "device",
                f"Device identity changed after preview for '{device_name}'. Refresh the preview before importing.",
                rack_name,
            )
        conflict_source_id = _device_binding_conflict(ctx.profile, source_id, device)
        if conflict_source_id:
            return RowResult(
                row_number=row["_row_number"],
                source_id=source_id,
                name=device_name,
                action="error",
                object_type="device",
                detail=(
                    f"Device '{device.name}' is already bound to source ID '{conflict_source_id}'. "
                    "One NetBox device cannot represent two source rows."
                ),
                rack_name=rack_name,
                extra_data={"identity_conflict": "device_already_bound", "netbox_device_id": device.pk},
            )
        if actual_action == "update":
            if (
                ctx.user is not None
                and not _device_queryset_for_user(Device, ctx.user, "change").filter(pk=device.pk).exists()
            ):
                return _perm_denied_row("dcim.change_device", row, device_name, "device")
            device.device_type = device_type
            device.role = device_role
            device.rack = rack if isinstance(rack, Rack) else None
            device.position = position
            if device_type.u_height == 0:
                device.face = None
            elif face is not None:
                device.face = face
            if airflow is not None:
                device.airflow = airflow
            device.status = status
            device.serial = serial or device.serial
            if asset_tag:
                device.asset_tag = asset_tag
            device.location = write_ctx.location
            device.tenant = write_ctx.tenant
            ip_json = {}
            try:
                with transaction.atomic():
                    device.full_clean()
                    device.save()
                    _bind_device_source(ctx.profile, source_id, device, asset_tag)
                    for ip_field, ip_str in (ip_fields or {}).items():
                        assigned = _assign_ip_to_device(device, ip_field, ip_str, ctx.user)
                        if not assigned:
                            ip_json[ip_field] = ip_str
                    PrimaryContactResolver.apply(device, ctx.profile, contact_review, ctx.user)
                    _store_source_id(
                        device,
                        ctx.profile,
                        source_id,
                        contact_review.extra_columns,
                        ip_json or None,
                    )
                    _enforce_saved_object_permission(device, ctx.user, "change")
            except _ObjectPermissionDenied as exc:
                return _perm_denied_row(str(exc) or "dcim.change_device", row, device_name, "device")
            except (DatabaseError, ValidationError) as exc:
                return _rack_position_error_row(
                    row, source_id, device_name, make, model, asset_tag, rack_name, position, exc, "update"
                )
            return RowResult(
                row_number=row["_row_number"],
                source_id=source_id,
                name=device_name,
                action="update",
                object_type="device",
                detail=f"Updated device '{device.name}' (matched by {match_method})",
                netbox_url=device.get_absolute_url(),
                rack_name=rack_name,
                extra_data={"source_make": make, "source_model": model, "asset_tag": asset_tag or ""},
            )
        return RowResult(
            row_number=row["_row_number"],
            source_id=source_id,
            name=device_name,
            action="skip",
            object_type="device",
            detail=(
                f"Device '{device.name}' matches this row, which writes nothing"
                if ctx.profile.adapter_settings.update_existing
                else f"Device '{device.name}' already exists (update_existing=False)"
            ),
            rack_name=rack_name,
            extra_data={"source_make": make, "source_model": model, "asset_tag": asset_tag or ""},
        )

    position, face = _zero_u_overrides(device_type, position, face)
    rack = write_ctx.rack_map.get(_identity_text(rack_name)) if rack_name else None
    if rack_name and rack is None:
        rack, ambiguous = _get_unique_rack(Rack, write_ctx, rack_name, lock=True, permission_action="view")
        if ambiguous:
            return _ambiguous_rack_row(row, source_id, device_name, rack_name, write_ctx, "device")
    if rack_name and not isinstance(rack, Rack):
        return RowResult(
            row_number=row["_row_number"],
            source_id=source_id,
            name=device_name,
            action="error",
            object_type="device",
            detail=(
                f"Rack '{_rack_identity_label(rack_name, write_ctx.location)}' was not found. "
                "Select the correct import location or fix the rack name."
            ),
            rack_name=rack_name,
            extra_data={"identity_conflict": "rack_not_found"},
        )
    if not _intent_matches(ctx, row, "device", "create"):
        return _identity_state_error(
            row,
            source_id,
            device_name,
            "device",
            f"Device identity changed after preview for '{device_name}'. Refresh the preview before importing.",
            rack_name,
        )
    if ctx.user is not None and not ctx.user.has_perm("dcim.add_device"):
        return _perm_denied_row("dcim.add_device", row, device_name, "device")
    ip_json = {}
    try:
        contact_review = PrimaryContactResolver.review(
            None,
            row,
            ctx.profile,
            ctx.user,
            candidate_source_columns=ctx.candidate_source_columns,
        )
        with transaction.atomic():
            device = Device(
                site=ctx.site,
                location=ctx.location,
                name=device_name,
                device_type=device_type,
                role=device_role,
                rack=rack if isinstance(rack, Rack) else None,
                position=position,
                face=face,
                airflow=airflow,
                status=status,
                serial=serial,
                asset_tag=asset_tag,
                tenant=ctx.tenant,
            )
            device.full_clean()
            device.save()
            _bind_device_source(ctx.profile, source_id, device, asset_tag)
            # Device.save() instantiates the type's interface templates, so an address can land now.
            for ip_field, ip_str in (ip_fields or {}).items():
                if not _assign_ip_to_device(device, ip_field, ip_str, ctx.user):
                    ip_json[ip_field] = ip_str
            PrimaryContactResolver.apply(device, ctx.profile, contact_review, ctx.user)
            _store_source_id(
                device,
                ctx.profile,
                source_id,
                contact_review.extra_columns,
                ip_json or None,
            )
            _enforce_saved_object_permission(device, ctx.user, "add")
    except _ObjectPermissionDenied as exc:
        return _perm_denied_row(str(exc) or "dcim.add_device", row, device_name, "device")
    except (DatabaseError, ValidationError) as exc:
        return _rack_position_error_row(
            row, source_id, device_name, make, model, asset_tag, rack_name, position, exc, "create"
        )
    _rack_label = rack_name if rack_name else "(no rack)"
    _pos_label = f" U{position}" if position is not None else ""
    return RowResult(
        row_number=row["_row_number"],
        source_id=source_id,
        name=device_name,
        action="create",
        object_type="device",
        detail=f"Created device '{device_name}' in {_rack_label}{_pos_label}",
        netbox_url=device.get_absolute_url(),
        rack_name=rack_name,
        extra_data={"source_make": make, "source_model": model, "asset_tag": asset_tag or ""},
    )


def _assign_ip_to_device(device, ip_field: str, ip_str: str, user=None):
    """Put one address on an interface of *device*; return whether it landed.

    False means there is nowhere to put it, and the caller records the value as unassigned so the
    row does not lose it. The preview names the same interface through the same resolver.
    """
    from . import ip_assignment

    try:
        target = ip_assignment.resolve(device, ip_field, ip_str)
    except ip_assignment.IPAssignmentError:
        return False
    if target.already_held:
        if getattr(device, f"{ip_field}_id", None) != target.existing.pk:
            setattr(device, ip_field, target.existing)
            device.save(update_fields=[ip_field])
        return True
    setattr(device, ip_field, ip_assignment.apply(target, user))
    device.save(update_fields=[ip_field])
    return True


def _pass3_process_devices(rows, ctx, class_role_map):  # noqa: C901
    """Pass 3: create or update Device objects.

    Cyclomatic complexity sits at 16 due to a stack of flat early-return guard
    clauses (ignored-id, position < 1, missing name, no class mapping, ignored
    class, no role slug, etc.). Each clause is simple and self-contained; the
    threshold-exceeding count is accumulation, not nesting, so a per-function
    suppression is preferred over a project-wide bump.
    """
    from dcim.models import Device, DeviceRole, DeviceType, Rack

    side_map, airflow_map, status_map = _get_translation_maps()

    # Identify device names that appear in multiple rows. When the same name occurs
    # in 2+ rows the name-based device lookup would incorrectly match all of them
    # to the same NetBox object (e.g. after one of them is synced), so we skip the
    # name fallback for those ambiguous rows and only match via source-ID link,
    # serial, or asset tag.
    _name_counts: dict[str, int] = {}
    for _row in rows:
        _dc = _str_val(_row.get("device_class"))
        _crm = class_role_map.get(_dc)
        if not _is_writing_device_row(_row, _crm, ctx.ignored_source_ids):
            continue
        _dn = _effective_device_name(_row)
        if _dn:
            name_key = _identity_text(_dn)
            _name_counts[name_key] = _name_counts.get(name_key, 0) + 1
    ambiguous_names: frozenset = frozenset(n for n, c in _name_counts.items() if c > 1)
    _reserve_device_names(rows, ctx, class_role_map, Device)
    effective_identity_values = _effective_duplicate_identity_values(rows, ctx, class_role_map, ambiguous_names, Device)
    serial_rows: dict = {}
    asset_tag_rows = {}
    for identity_row_number, identity in effective_identity_values.items():
        serial = identity.get("serial")
        asset_tag = identity.get("asset_tag")
        if serial:
            serial_rows.setdefault(serial, []).append(identity_row_number)
        if asset_tag:
            asset_tag_key = _identity_text(asset_tag)
            asset_tag_rows.setdefault(asset_tag_key, []).append(identity_row_number)
    ctx.duplicate_serial_rows = {serial: sorted(numbers) for serial, numbers in serial_rows.items() if len(numbers) > 1}
    ctx.duplicate_serials = frozenset(ctx.duplicate_serial_rows)
    ctx.duplicate_asset_tag_rows = {
        asset_tag: sorted(numbers) for asset_tag, numbers in asset_tag_rows.items() if len(numbers) > 1
    }
    ctx.duplicate_asset_tags = frozenset(ctx.duplicate_asset_tag_rows)
    ctx.effective_duplicate_identity = effective_identity_values

    total_rows = len(rows)
    for processed_rows, row in enumerate(rows):
        if ctx.progress_callback is not None:
            ctx.progress_callback(processed_rows, total_rows)
        device_class = _str_val(row.get("device_class"))
        crm = class_role_map.get(device_class)
        if crm and crm.creates_rack:
            continue

        source_id = _str_val(row.get("source_id"))
        device_name = _effective_device_name(row)
        rack_name = _str_val(row.get("rack_name"))
        make = " ".join((_str_val(row.get("make")) or "Unknown").split())
        model = " ".join((_str_val(row.get("model")) or "Unknown").split())
        serial = _str_val(row.get("serial"))
        asset_tag_raw = _str_val(row.get("asset_tag")) or None
        asset_tag = asset_tag_raw[:50] if asset_tag_raw else None

        slug_conflict = ctx.slug_conflicts_by_row.get(row.get("_row_number"))
        if slug_conflict and not ctx.dry_run:
            ctx.result.rows.append(
                RowResult(
                    row_number=row["_row_number"],
                    source_id=source_id,
                    name=device_name,
                    action="error",
                    object_type="device",
                    detail=slug_conflict,
                    rack_name=rack_name,
                    extra_data={"identity_conflict": "derived_slug_collision"},
                )
            )
            continue

        if source_id and source_id in ctx.duplicate_source_ids:
            other_rows = [
                number for number in ctx.duplicate_source_id_rows.get(source_id, []) if number != row["_row_number"]
            ]
            ctx.result.rows.append(
                RowResult(
                    row_number=row["_row_number"],
                    source_id=source_id,
                    name=device_name,
                    action="error",
                    object_type="device",
                    detail=_duplicate_value_detail("source ID", source_id, other_rows),
                    rack_name=rack_name,
                    extra_data={
                        "identity_conflict": "duplicate_source_id",
                        "duplicate_source_id_rows": other_rows,
                    },
                )
            )
            continue

        identity = ctx.effective_duplicate_identity.get(row.get("_row_number"), {})
        if identity.get("serial") and identity["serial"] in ctx.duplicate_serials:
            duplicate_serial = identity["serial"]
            others = [
                number for number in ctx.duplicate_serial_rows.get(duplicate_serial, []) if number != row["_row_number"]
            ]
            ctx.result.rows.append(
                RowResult(
                    row_number=row["_row_number"],
                    source_id=source_id,
                    name=device_name,
                    action="error",
                    object_type="device",
                    detail=_duplicate_value_detail("serial", duplicate_serial, others),
                    rack_name=rack_name,
                    extra_data={
                        "identity_conflict": "duplicate_serial",
                        "duplicate_serial": duplicate_serial,
                        "duplicate_serial_rows": others,
                    },
                )
            )
            continue

        effective_asset_tag = identity.get("asset_tag")
        asset_tag_key = _identity_text(effective_asset_tag) if effective_asset_tag else ""
        if asset_tag_key and asset_tag_key in ctx.duplicate_asset_tags:
            other_rows = [
                number for number in ctx.duplicate_asset_tag_rows.get(asset_tag_key, []) if number != row["_row_number"]
            ]
            ctx.result.rows.append(
                RowResult(
                    row_number=row["_row_number"],
                    source_id=source_id,
                    name=device_name,
                    action="error",
                    object_type="device",
                    detail=_duplicate_value_detail("asset tag", effective_asset_tag, other_rows),
                    rack_name=rack_name,
                    extra_data={
                        "identity_conflict": "duplicate_asset_tag",
                        "duplicate_asset_tag_rows": other_rows,
                    },
                )
            )
            continue

        ip_fields = {}
        for ip_field in ("primary_ip4", "primary_ip6", "oob_ip"):
            raw = str(row.get(ip_field, "")).strip()
            if raw:
                parsed = _parse_ip_with_prefix(raw)
                if parsed:
                    ip_fields[ip_field] = parsed
                else:
                    logger.warning("Row %s: unparseable IP value for %s: %r", row.get("_row_number"), ip_field, raw)

        if source_id and source_id in ctx.ignored_source_ids:
            ctx.result.rows.append(
                RowResult(
                    row_number=row["_row_number"],
                    source_id=source_id,
                    name=device_name,
                    action="ignore",
                    object_type="device",
                    detail="Ignored device",
                    rack_name=rack_name,
                    extra_data={"ignore_kind": "individual"},
                )
            )
            continue

        if _identity_text(device_name) in ambiguous_names:
            strong_device, strong_method = _find_existing_device(
                ctx.profile,
                source_id,
                ctx.site,
                device_name,
                serial,
                asset_tag,
                Device,
                ambiguous_names,
                tenant=ctx.tenant,
                device_queryset=(_device_queryset_for_user(Device, ctx.user, "view") if ctx.user is not None else None),
                review_device_ids=(
                    ctx.field_reviewer.review_device_ids(source_id) if ctx.field_reviewer is not None else frozenset()
                ),
            )
            if strong_method == "inaccessible device":
                ctx.result.rows.append(_perm_denied_row("dcim.view_device", row, device_name, "device"))
                continue
            if strong_method == "ambiguous serial":
                ctx.result.rows.append(_ambiguous_serial_row(row, source_id, device_name, rack_name, serial))
                continue
            if strong_method == "ambiguous asset tag":
                ctx.result.rows.append(_ambiguous_asset_tag_row(row, source_id, device_name, rack_name, asset_tag))
                continue
            if strong_method == "ambiguous stored source ID":
                ctx.result.rows.append(_ambiguous_source_id_row(row, source_id, device_name, rack_name))
                continue
            if strong_method == "ambiguous field review":
                ctx.result.rows.append(_ambiguous_field_review_row(row, source_id, device_name, rack_name))
                continue
            if strong_device is None:
                existing = (
                    _device_queryset_for_user(Device, ctx.user, "view")
                    .filter(
                        site=ctx.site,
                        name__iexact=device_name,
                        **_tenant_filter(ctx.tenant),
                    )
                    .select_related("rack__location")
                    .first()
                )
                placement = f"source rack '{rack_name or '(none)'}'"
                if row.get("u_position") not in (None, ""):
                    placement += f", U{row.get('u_position')}"
                detail = f"Duplicate device name '{device_name}' in this import ({placement})."
                extra_data = {
                    "identity_conflict": "duplicate_name",
                    "suggested_name": _suggest_unique_device_name(row, ctx),
                    "source_rack_name": rack_name,
                    "source_position": row.get("u_position") or "",
                }
                if existing is not None:
                    netbox_rack = _device_rack_identity_label(existing) or "(none)"
                    detail += (
                        f" NetBox already has '{existing.name}' in rack '{netbox_rack}'"
                        f"{f', U{existing.position}' if existing.position is not None else ''}."
                    )
                    extra_data.update(
                        {
                            "netbox_device_id": existing.pk,
                            "netbox_device_name": existing.name,
                            "netbox_rack_name": netbox_rack,
                            "netbox_position": _normalize_for_compare(existing.position),
                            "netbox_face": existing.face or "",
                            "netbox_url": existing.get_absolute_url(),
                        }
                    )
                ctx.result.rows.append(
                    RowResult(
                        row_number=row["_row_number"],
                        source_id=source_id,
                        name=device_name,
                        action="error",
                        object_type="device",
                        detail=detail,
                        rack_name=rack_name,
                        extra_data=extra_data,
                    )
                )
                continue
            conflict_source_id = _device_binding_conflict(ctx.profile, source_id, strong_device)
            if conflict_source_id:
                ctx.result.rows.append(
                    RowResult(
                        row_number=row["_row_number"],
                        source_id=source_id,
                        name=device_name,
                        action="error",
                        object_type="device",
                        detail=(
                            f"Device '{strong_device.name}' matched by {strong_method} is already bound "
                            f"to source ID '{conflict_source_id}'."
                        ),
                        rack_name=rack_name,
                        extra_data={"identity_conflict": "device_already_bound"},
                    )
                )
                continue

        position = _coerce_position(row.get("u_position"))
        if position is not None and position < 1:
            ctx.result.rows.append(
                RowResult(
                    row_number=row["_row_number"],
                    source_id=source_id,
                    name=device_name,
                    action="skip",
                    object_type="device",
                    detail=f"Skipped: position {position} < 1 (under-rack/blanking panel)",
                    rack_name=rack_name,
                )
            )
            continue

        if not device_name:
            ctx.result.rows.append(
                RowResult(
                    row_number=row["_row_number"],
                    source_id=source_id,
                    name="",
                    action="error",
                    object_type="device",
                    detail="Missing device name",
                )
            )
            continue

        mfg_slug, dt_slug, is_explicit_mapping = _resolve_device_type_slugs(
            make,
            model,
            ctx.profile,
            ctx.device_type_identity,
        )
        u_height_raw = row.get("u_height", 1)
        u_height = max(1, _coerce_int(u_height_raw, 1))

        if not crm:
            ctx.result.rows.append(
                RowResult(
                    row_number=row["_row_number"],
                    source_id=source_id,
                    name=device_name,
                    action="error",
                    object_type="device",
                    detail=f"No class→role mapping for class '{device_class}'",
                    extra_data={
                        "source_class": device_class,
                        "profile_id": ctx.profile.pk,
                        "source_make": make,
                        "source_model": model,
                        "asset_tag": asset_tag or "",
                        "mfg_slug": mfg_slug,
                        "dt_slug": dt_slug,
                        "u_height": u_height,
                        "is_explicit_mapping": is_explicit_mapping,
                    },
                )
            )
            continue

        if crm.ignore:
            ctx.result.rows.append(
                RowResult(
                    row_number=row["_row_number"],
                    source_id=source_id,
                    name=device_name,
                    action="ignore",
                    object_type="device",
                    detail=f"Ignored: class '{device_class}'",
                    rack_name=rack_name,
                )
            )
            continue

        if not crm.role_slug:
            ctx.result.rows.append(
                RowResult(
                    row_number=row["_row_number"],
                    source_id=source_id,
                    name=device_name,
                    action="error",
                    object_type="device",
                    detail=(
                        f"Class '{device_class}' has no device role configured "
                        "— edit the class→role mapping to set a role slug"
                    ),
                    rack_name=rack_name,
                    extra_data={
                        "source_class": device_class,
                        "profile_id": ctx.profile.pk,
                        "source_make": make,
                        "source_model": model,
                        "asset_tag": asset_tag or "",
                        "mfg_slug": mfg_slug,
                        "dt_slug": dt_slug,
                        "u_height": u_height,
                        "is_explicit_mapping": is_explicit_mapping,
                    },
                )
            )
            continue

        device_status = status_map.get(_str_val(row.get("status")).lower(), "active")
        device_face = side_map.get(_str_val(row.get("face")).lower())
        device_airflow = airflow_map.get(_str_val(row.get("airflow")).lower())

        if ctx.dry_run:
            row_result = _preview_device_row(
                row,
                ctx,
                make,
                model,
                mfg_slug,
                dt_slug,
                source_id,
                device_name,
                serial,
                asset_tag,
                DeviceType,
                Device,
                Rack,
                ip_fields=ip_fields,
                device_face=device_face,
                device_airflow=device_airflow,
                device_status=device_status,
                u_position=position,
                is_explicit_mapping=is_explicit_mapping,
                ambiguous_names=ambiguous_names,
                role_slug=crm.role_slug,
            )
        else:
            row_result = _write_device_row(
                row,
                ctx,
                make,
                model,
                crm,
                mfg_slug,
                dt_slug,
                source_id,
                device_name,
                serial,
                asset_tag,
                position,
                device_face,
                device_airflow,
                device_status,
                DeviceType,
                DeviceRole,
                Rack,
                Device,
                ip_fields=ip_fields,
                ambiguous_names=ambiguous_names,
            )
        if slug_conflict:
            row_result.action = "error"
            row_result.detail = slug_conflict
            row_result.extra_data.update(
                {
                    "identity_conflict": "derived_slug_collision",
                    "source_make": make,
                    "source_model": model,
                    "mfg_slug": mfg_slug,
                    "dt_slug": dt_slug,
                    "u_height": u_height,
                    "is_explicit_mapping": is_explicit_mapping,
                }
            )
        ctx.result.rows.append(row_result)

    if ctx.progress_callback is not None:
        ctx.progress_callback(total_rows, total_rows)


# ---------------------------------------------------------------------------
# Main import runner
# ---------------------------------------------------------------------------


def _identity_text(value):
    """Normalize a label for source identity comparisons."""
    return " ".join(str(value).split()).casefold()


def _add_within_file_slug_conflicts(
    manufacturer_groups,
    device_type_groups,
    conflicts,
    ignored_device_type_rows=frozenset(),
):
    """Add collisions between different identities in the same source file."""
    for slug, records in manufacturer_groups.items():
        records = [record for record in records if record["row_number"] not in ignored_device_type_rows]
        source_makes = {_identity_text(record["make"]) for record in records}
        if len(source_makes) <= 1 or all(record["explicit_manufacturer"] for record in records):
            continue
        source_labels = {record["make"] for record in records}
        detail = (
            f"Different manufacturer names derive the same slug '{slug}': "
            f"{', '.join(sorted(source_labels))}. Add explicit manufacturer mappings."
        )
        for record in records:
            conflicts[record["row_number"]] = detail

    for (mfg_slug, dt_slug), records in device_type_groups.items():
        records = [record for record in records if record["row_number"] not in ignored_device_type_rows]
        source_types = {(_identity_text(record["make"]), _identity_text(record["model"])) for record in records}
        if len(source_types) <= 1 or all(record["explicit_device_type"] for record in records):
            continue
        source_labels = {(record["make"], record["model"]) for record in records}
        labels = ", ".join(sorted(f"{make} / {model}" for make, model in source_labels))
        detail = (
            f"Different device types derive the same slug '{mfg_slug}/{dt_slug}': {labels}. "
            "Add explicit device type mappings."
        )
        for record in records:
            conflicts[record["row_number"]] = detail


def _add_existing_slug_conflicts(
    manufacturer_groups, device_type_groups, conflicts, ignored_device_type_rows=frozenset()
):
    """Add collisions with differently named objects that already exist in NetBox."""
    from dcim.models import DeviceType, Manufacturer

    existing_manufacturers = {
        manufacturer.slug: manufacturer
        for manufacturer in Manufacturer.objects.filter(slug__in=manufacturer_groups).only("name", "slug")
    }
    for slug, records in manufacturer_groups.items():
        records = [record for record in records if record["row_number"] not in ignored_device_type_rows]
        existing = existing_manufacturers.get(slug)
        if existing is None:
            continue
        for record in records:
            if record["explicit_manufacturer"] or _identity_text(existing.name) == _identity_text(record["make"]):
                continue
            conflicts[record["row_number"]] = (
                f"Manufacturer '{record['make']}' derives slug '{slug}', which already belongs to "
                f"'{existing.name}' in NetBox. Add an explicit manufacturer mapping."
            )

    existing_device_types = {
        (device_type.manufacturer.slug, device_type.slug): device_type
        for device_type in DeviceType.objects.filter(
            manufacturer__slug__in=manufacturer_groups,
            slug__in={key[1] for key in device_type_groups},
        ).select_related("manufacturer")
    }
    for key, records in device_type_groups.items():
        records = [record for record in records if record["row_number"] not in ignored_device_type_rows]
        existing = existing_device_types.get(key)
        if existing is None:
            continue
        for record in records:
            if record["explicit_device_type"] or _identity_text(existing.model) == _identity_text(record["model"]):
                continue
            conflicts[record["row_number"]] = (
                f"Device type '{record['make']} / {record['model']}' derives slug '{key[0]}/{key[1]}', "
                f"which already belongs to '{existing.manufacturer.name} / {existing.model}' in NetBox. "
                "Add an explicit device type mapping."
            )


def _active_ignored_device_type_rows(
    rows,
    profile,
    class_role_map,
    field_reviewer,
    site,
    tenant,
    user,
    device_type_identity,
):
    """Return source rows whose exact active type review suppresses type planning."""
    if field_reviewer is None or site is None:
        return frozenset()
    from dcim.models import Device

    device_queryset = _device_queryset_for_user(Device, user, "view") if user is not None else None
    ignored_rows = set()
    for row in rows:
        source_id = _str_val(row.get("source_id"))
        review_device_ids = field_reviewer.review_device_ids(source_id)
        if not source_id or len(review_device_ids) != 1:
            continue
        crm = class_role_map.get(_str_val(row.get("device_class")))
        if not _is_writing_device_row(row, crm, frozenset()):
            continue
        make = " ".join((_str_val(row.get("make")) or "Unknown").split())
        model = " ".join((_str_val(row.get("model")) or "Unknown").split())
        mfg_slug, dt_slug, _explicit_identity = _resolve_device_type_slugs(
            make,
            model,
            profile,
            device_type_identity,
        )
        matched_device, _match_method = _find_existing_device(
            profile,
            source_id,
            site,
            _effective_device_name(row),
            _str_val(row.get("serial")),
            (_str_val(row.get("asset_tag")) or "")[:50],
            Device,
            tenant=tenant,
            device_queryset=device_queryset,
            review_device_ids=review_device_ids,
        )
        if matched_device is None:
            continue
        review = field_reviewer.review(
            source_id,
            matched_device,
            {"device_type": (mfg_slug, dt_slug, make, model)},
        )
        if "device_type" in review.ignored:
            ignored_rows.add(row.get("_row_number"))
    return frozenset(ignored_rows)


def _derived_slug_conflicts(
    rows,
    profile,
    class_role_map,
    ignored_source_ids=frozenset(),
    *,
    field_reviewer=None,
    site=None,
    tenant=None,
    user=None,
    device_type_identity=None,
):
    """Return per-row errors for different source identities that derive one slug."""
    manufacturer_groups = {}
    device_type_groups = {}
    device_type_identity = device_type_identity or _DeviceTypeIdentityResolver.for_profile(profile)
    mapped_source_makes = device_type_identity.mapped_source_makes
    for row in rows:
        crm = class_role_map.get(_str_val(row.get("device_class")))
        if not _is_writing_device_row(row, crm, ignored_source_ids):
            continue
        make = " ".join((_str_val(row.get("make")) or "Unknown").split())
        model = " ".join((_str_val(row.get("model")) or "Unknown").split())
        mfg_slug, dt_slug, explicit_device_type = _resolve_device_type_slugs(
            make,
            model,
            profile,
            device_type_identity,
        )
        explicit_manufacturer = explicit_device_type or make in mapped_source_makes
        record = {
            "row_number": row.get("_row_number"),
            "make": make,
            "model": model,
            "explicit_device_type": explicit_device_type,
            "explicit_manufacturer": explicit_manufacturer,
        }
        manufacturer_groups.setdefault(mfg_slug, []).append(record)
        device_type_groups.setdefault((mfg_slug, dt_slug), []).append(record)

    ignored_device_type_rows = _active_ignored_device_type_rows(
        rows,
        profile,
        class_role_map,
        field_reviewer,
        site,
        tenant,
        user,
        device_type_identity,
    )
    conflicts = {}
    _add_within_file_slug_conflicts(
        manufacturer_groups,
        device_type_groups,
        conflicts,
        ignored_device_type_rows,
    )
    _add_existing_slug_conflicts(
        manufacturer_groups,
        device_type_groups,
        conflicts,
        ignored_device_type_rows,
    )
    return conflicts


def run_import(
    rows: list[dict],
    profile: ImportProfile,
    context: dict,
    dry_run: bool = True,
    user: object | None = None,
    expected_intents: dict | None = None,
    progress_callback: Callable[[int, int], None] | None = None,
) -> ImportResult:
    """Run (or preview) the import.

    context keys: site, location (optional), tenant (optional)
    dry_run=True  → no DB writes, returns what *would* happen
    dry_run=False → writes to DB; pass user to enforce DCIM object permissions

    Raises ValidationError when the profile names an adapter this release does not register.
    """
    from .models import IgnoredDevice, validate_registered_adapter

    # Every adapter setting is read from deep inside the passes, so reject the profile up front.
    validate_registered_adapter(profile)

    class_role_map: dict[str, object] = {
        crm.source_class: crm for crm in profile.class_role_mappings.select_related("rack_type").all()
    }
    ignored_source_ids = frozenset(
        _str_val(source_id)
        for source_id in IgnoredDevice.objects.filter(profile=profile).values_list("source_id", flat=True)
    )
    source_id_rows = {}
    rack_name_counts = {}
    for row in rows:
        source_id = _str_val(row.get("source_id"))
        crm = class_role_map.get(_str_val(row.get("device_class")))
        if not crm or crm.ignore or source_id in ignored_source_ids:
            continue
        if crm.creates_rack:
            if source_id:
                source_id_rows.setdefault(source_id, []).append(row["_row_number"])
            rack_name = _str_val(row.get("rack_name")) or _str_val(row.get("device_name"))
            if rack_name:
                rack_key = _identity_text(rack_name)
                rack_name_counts[rack_key] = rack_name_counts.get(rack_key, 0) + 1
            continue
        if not _is_writing_device_row(row, crm, ignored_source_ids):
            continue
        if source_id:
            source_id_rows.setdefault(source_id, []).append(row["_row_number"])
    duplicate_source_id_rows = {
        source_id: sorted(numbers) for source_id, numbers in source_id_rows.items() if len(numbers) > 1
    }
    ctx = ImportContext(
        profile=profile,
        site=context["site"],
        location=context.get("location"),
        tenant=context.get("tenant"),
        dry_run=dry_run,
        result=ImportResult(),
        user=user,
        expected_intents=expected_intents or {},
        duplicate_source_ids=frozenset(duplicate_source_id_rows),
        duplicate_source_id_rows=duplicate_source_id_rows,
        duplicate_rack_names=frozenset(name for name, count in rack_name_counts.items() if count > 1),
        ignored_source_ids=ignored_source_ids,
        candidate_source_columns=_build_candidate_source_columns(profile),
        progress_callback=progress_callback,
        field_reviewer=DeviceFieldReviewer.for_profile(profile),
        device_type_identity=_DeviceTypeIdentityResolver.for_profile(profile),
    )
    ctx.slug_conflicts_by_row = _derived_slug_conflicts(
        rows,
        profile,
        class_role_map,
        ignored_source_ids,
        field_reviewer=ctx.field_reviewer,
        site=ctx.site,
        tenant=ctx.tenant,
        user=ctx.user,
        device_type_identity=ctx.device_type_identity,
    )

    if dry_run:
        _pass1_ensure_types(rows, ctx, class_role_map)
        _pass2_process_racks(rows, ctx, class_role_map)
        _pass3_process_devices(rows, ctx, class_role_map)
    else:
        with transaction.atomic():
            if profile.adapter_settings.primary_contact_role:
                PrimaryContactResolver.lock_imports()
            _pass1_ensure_types(rows, ctx, class_role_map)
            _pass2_process_racks(rows, ctx, class_role_map)
            _pass3_process_devices(rows, ctx, class_role_map)

    ctx.result._recompute_counts()
    if ctx.pending_device_roles:
        ctx.result.counts["device_roles_pending"] = len(ctx.pending_device_roles)
    return ctx.result


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _build_candidate_source_columns(profile: ImportProfile) -> dict[str, frozenset[str]]:
    """Return configured source columns grouped by candidate target."""
    grouped = {}
    for mapping in profile.column_mappings.all():
        if mapping.target_field.startswith(CANDIDATE_TARGET_PREFIX):
            target = mapping.target_field.removeprefix(CANDIDATE_TARGET_PREFIX)
            grouped.setdefault(target, set()).add(mapping.source_column)
    return {target: frozenset(source_columns) for target, source_columns in grouped.items()}


def _store_source_id(
    obj, profile: ImportProfile, source_id: str, extra_columns: dict | None = None, ip_data: dict | None = None
):
    """Store the source ID in the profile's custom field and the plugin's import record.

    The import record covers Devices. A Rack keeps only the operator-configured custom field.
    """
    # Per-profile custom field (e.g. cans_id → plain string)
    if profile.adapter_settings.custom_field_name and source_id:
        try:
            obj.custom_field_data[profile.adapter_settings.custom_field_name] = source_id
            obj.save(update_fields=["custom_field_data"])
        except (AttributeError, KeyError):  # pragma: no cover
            logger.warning("Failed to set custom field '%s' on %s", profile.adapter_settings.custom_field_name, obj)

    from dcim.models import Device

    if not isinstance(obj, Device):
        return
    DeviceImportSource.objects.update_or_create(
        device=obj,
        defaults={
            "profile": profile,
            "source_id": source_id or "",
            "extra_columns": extra_columns or {},
            "unassigned_ips": ip_data or {},
        },
    )
