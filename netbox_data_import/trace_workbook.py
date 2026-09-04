# SPDX-License-Identifier: Apache-2.0
# SPDX-FileCopyrightText: 2026 Marcin Zieba <marcinpsk@gmail.com>
"""Interpret trace workbooks without accessing the ORM."""

from __future__ import annotations

from collections import defaultdict
from dataclasses import asdict, dataclass, replace
from hashlib import sha256
from io import BytesIO
import json
from typing import Iterable, Mapping, Sequence

import openpyxl

from .adapters import SourceDiagnostic, SourceUnreadable
from .field_keys import (
    FRONT_PORT_CLASSES,
    INTERFACE_PORT_CLASSES,
    PORT_CLASS_CLAIMED_KINDS,
    PORT_CLASSES,
    REAR_PORT_CLASSES,
    same_device_and_cards,
)
from .trace_schema import TRACE_EXPORT_TIMESTAMP_MAX_LENGTH
from .values import identity_text, source_text

TRACE_PATH_SHEET = "Trace From To"
TRACE_LIST_SHEET = "Trace List"

IdentityKey = tuple[str, str, str, str]

_PATH_HEADER = (
    "Port",
    "PortClass",
    "Cards",
    "Device",
    "UPos",
    "Rack",
    "Location",
    "CableClass",
    "Port",
    "PortClass",
    "Cards",
    "Device",
    "UPos",
    "Rack",
    "Location",
)
_LIST_HEADER = ("Location", "Rack", "UPos", "Device", "Cards", "Port", "PortClass", "Cable")
_BLOCK_LINE_MARKERS = ("From", "To")


@dataclass(frozen=True)
class TerminationReference:
    """Name one source termination and retain its corroboration values."""

    device: str
    cards: str
    port: str
    port_class: str
    u_position: str = ""
    rack: str = ""
    location: str = ""

    @property
    def identity_key(self) -> IdentityKey:
        """Return the normalized device, cards, port, and claimed-kind identity."""
        return (
            identity_text(self.device),
            identity_text(self.cards),
            identity_text(self.port),
            # An unrecognized PortClass keeps its own text, so two unknown values stay apart.
            PORT_CLASS_CLAIMED_KINDS.get(self.port_class, identity_text(self.port_class)),
        )


@dataclass(frozen=True)
class EndpointSummary:
    """Retain the source From and To statements in their original direction."""

    from_termination: TerminationReference
    to_termination: TerminationReference
    from_text: str
    to_text: str


@dataclass(frozen=True)
class SegmentEvidence:
    """Describe one source-claimed cable between two Termination References."""

    left: TerminationReference
    cable_class: str
    right: TerminationReference


@dataclass(frozen=True)
class PassThroughClaim:
    """Describe the source-claimed continuation through one device."""

    device: str
    cards: str
    entry_port: str
    exit_port: str


@dataclass(frozen=True)
class TraceProvenance:
    """Locate one Source Trace occurrence in its workbook."""

    workbook_fingerprint: str
    sheet: str
    block_ordinal: int
    row_start: int
    row_end: int
    export_timestamp: str
    from_text: str
    to_text: str
    direction: str


@dataclass(frozen=True)
class SourceTrace:
    """Carry canonical Segment Evidence for one path or an Endpoint Summary fallback."""

    endpoint_summary: EndpointSummary
    segments: tuple[SegmentEvidence, ...]
    pass_through_claims: tuple[PassThroughClaim, ...]
    corroboration: tuple[TerminationReference, ...]
    identity: str
    content_fingerprint: str
    provenance: tuple[TraceProvenance, ...]
    errors: tuple[SourceDiagnostic, ...] = ()

    @property
    def valid(self) -> bool:
        """Return whether source validation found no errors on this trace."""
        return not self.errors

    @property
    def ends_at_rear_port(self) -> bool:
        """Return whether the original To termination is a rear port."""
        return self.endpoint_summary.to_termination.port_class in REAR_PORT_CLASSES


@dataclass(frozen=True)
class _RawRow:
    """One spreadsheet row, holding the cells the sheet width defines."""

    row_number: int
    values: tuple[object, ...]


@dataclass(frozen=True)
class _Block:
    """One From and To block with its header row, data rows, and workbook provenance."""

    sheet: str
    ordinal: int
    row_start: int
    row_end: int
    from_text: str
    to_text: str
    has_to_line: bool
    header: tuple[str, ...]
    rows: tuple[_RawRow, ...]
    export_timestamp: str
    workbook_fingerprint: str

    @property
    def pair_key(self) -> tuple[str, str]:
        """Return the From and To text that pairs this block with the other sheet."""
        return source_text(self.from_text), source_text(self.to_text)


@dataclass(frozen=True)
class _ParsedSegment:
    """One Segment Evidence entry with the row that stated it."""

    evidence: SegmentEvidence
    row_number: int


@dataclass(frozen=True)
class _ParsedVisit:
    """One Trace List visit with the row that stated it."""

    termination: TerminationReference
    row_number: int


_SegmentClaim = tuple[tuple[IdentityKey, IdentityKey], str]


def parse_endpoint_line(line: str) -> TerminationReference:
    """Parse ``Device > [Cards > ]Port (PortClass)`` into a Termination Reference."""
    parts = source_text(line).split(" > ")
    if len(parts) not in (2, 3):
        raise ValueError("Endpoint line must contain a device, an optional cards label, and a port.")
    device = source_text(parts[0])
    cards = source_text(parts[1]) if len(parts) == 3 else ""
    port_and_class = source_text(parts[-1])
    marker = port_and_class.rfind(" (")
    if marker <= 0 or not port_and_class.endswith(")"):
        raise ValueError("Endpoint line must end with a PortClass in parentheses.")
    port = source_text(port_and_class[:marker])
    port_class = source_text(port_and_class[marker + 2 : -1])
    if not device or not port or not port_class:
        raise ValueError("Endpoint line contains an empty device, port, or PortClass.")
    return TerminationReference(device=device, cards=cards, port=port, port_class=port_class)


def canonical_trace_identity(
    first: TerminationReference,
    second: TerminationReference,
) -> str:
    """Return the sorted endpoint pair as compact canonical JSON."""
    endpoints = sorted((first.identity_key, second.identity_key))
    return json.dumps(endpoints, ensure_ascii=False, separators=(",", ":"))


def canonical_orientation(
    from_termination: TerminationReference,
    to_termination: TerminationReference,
    segments: Sequence[SegmentEvidence],
) -> tuple[SegmentEvidence, ...]:
    """Orient Segment Evidence from the endpoint whose identity key sorts first."""
    if from_termination.identity_key <= to_termination.identity_key:
        return tuple(segments)
    return tuple(
        SegmentEvidence(left=segment.right, cable_class=segment.cable_class, right=segment.left)
        for segment in reversed(segments)
    )


def content_fingerprint(
    from_termination: TerminationReference,
    to_termination: TerminationReference,
    segments: Sequence[SegmentEvidence],
) -> str:
    """Hash the canonical endpoints, Segment Evidence, and Pass-Through Claims."""
    oriented = canonical_orientation(from_termination, to_termination, segments)
    claims = _pass_through_claims(oriented)
    endpoints = sorted((from_termination, to_termination), key=lambda termination: termination.identity_key)
    payload = {
        # An Endpoint Summary fallback states no segment, so the endpoints carry its whole content.
        "endpoints": [[termination.identity_key, termination.port_class] for termination in endpoints],
        "segments": [
            [
                segment.left.identity_key,
                segment.left.port_class,
                segment.cable_class,
                segment.right.identity_key,
                segment.right.port_class,
            ]
            for segment in oriented
        ],
        "pass_through_claims": [
            [
                identity_text(claim.device),
                identity_text(claim.cards),
                identity_text(claim.entry_port),
                identity_text(claim.exit_port),
            ]
            for claim in claims
        ],
    }
    serialized = json.dumps(payload, ensure_ascii=False, separators=(",", ":"), sort_keys=True)
    return sha256(serialized.encode()).hexdigest()


def _cell_text(value: object) -> str:
    """Return the raw text of a cell value, empty for None."""
    return "" if value is None else str(value)


def _normalized_header(values: Iterable[object]) -> tuple[str, ...]:
    """Return one header row as trimmed text."""
    return tuple(source_text(value) for value in values)


def _row_has_data(values: Iterable[object]) -> bool:
    """Return whether any cell in the row carries text."""
    return any(source_text(value) for value in values)


def _row_values(sheet, row_number: int, width: int) -> tuple[object, ...]:
    """Return the cells of one row across the sheet's fixed width."""
    return tuple(sheet.cell(row=row_number, column=column).value for column in range(1, width + 1))


def _last_data_row(values_by_row: Mapping[int, tuple[object, ...]], start: int, stop: int) -> int:
    """Return the last row in the range that carries data."""
    for row_number in range(stop, start - 1, -1):
        if _row_has_data(values_by_row[row_number]):
            return row_number
    return start


def _extract_blocks(sheet, workbook_fingerprint: str) -> tuple[_Block, ...]:
    """Return every block on one sheet, including visits the next block's lines overwrite."""
    width = len(_PATH_HEADER) if sheet.title == TRACE_PATH_SHEET else len(_LIST_HEADER)
    # openpyxl builds a cell object on first access, so the sheet is read once and every step reuses it.
    values_by_row = {row: _row_values(sheet, row, width) for row in range(1, sheet.max_row + 1)}
    starts = [row for row, values in values_by_row.items() if source_text(values[0]) == "From"]
    first_row = values_by_row.get(1, ())
    export_timestamp = source_text(first_row[1]) if first_row and source_text(first_row[0]) == "Executed" else ""
    blocks = []
    for index, start in enumerate(starts):
        boundary = starts[index + 1] if index + 1 < len(starts) else sheet.max_row + 1
        to_row = start + 1
        has_to_line = to_row < boundary and source_text(values_by_row[to_row][0]) == "To"
        header_row = start + 2
        header = _normalized_header(values_by_row[header_row]) if header_row < boundary else ()
        rows = [
            _RawRow(row_number, values_by_row[row_number])
            for row_number in range(start + 3, boundary)
            if _row_has_data(values_by_row[row_number])
        ]
        row_end = _last_data_row(values_by_row, start, boundary - 1)
        # A block that fills its separator row keeps its last rows under the next block's lines.
        for carry_row in (boundary, boundary + 1):
            if carry_row > sheet.max_row:
                continue
            values = values_by_row[carry_row]
            if source_text(values[0]) not in _BLOCK_LINE_MARKERS:
                continue
            if _row_has_data(values[2:]):
                rows.append(_RawRow(carry_row, ("", "", *values[2:])))
                row_end = carry_row
        blocks.append(
            _Block(
                sheet=sheet.title,
                ordinal=index + 1,
                row_start=start,
                row_end=row_end,
                from_text=_cell_text(values_by_row[start][1]),
                to_text=_cell_text(values_by_row[to_row][1]) if has_to_line else "",
                has_to_line=has_to_line,
                header=header,
                rows=tuple(rows),
                export_timestamp=export_timestamp,
                workbook_fingerprint=workbook_fingerprint,
            )
        )
    return tuple(blocks)


def _location(block: _Block) -> str:
    """Return the readable source position of one block."""
    return f"{block.sheet} block {block.ordinal} (rows {block.row_start}-{block.row_end})"


def _error(block: _Block, code: str, detail: str, row_number: int | None = None) -> SourceDiagnostic:
    """Return one source diagnostic that names the block it came from."""
    return SourceDiagnostic(
        code=code,
        message=f"{_location(block)}: {detail}",
        row_number=block.row_start if row_number is None else row_number,
    )


def _metadata_errors(blocks: Iterable[_Block | None]) -> list[SourceDiagnostic]:
    """Return diagnostics for raw workbook metadata that cannot fit its stored representation."""
    errors = []
    for block in blocks:
        if block is not None and len(block.export_timestamp) > TRACE_EXPORT_TIMESTAMP_MAX_LENGTH:
            errors.append(
                _error(
                    block,
                    "trace.metadata_too_long",
                    f"The export timestamp exceeds {TRACE_EXPORT_TIMESTAMP_MAX_LENGTH} characters.",
                    row_number=1,
                )
            )
    return errors


def _endpoint_summary(block: _Block) -> tuple[EndpointSummary | None, list[SourceDiagnostic]]:
    """Return the Endpoint Summary of one block, or the reason it has none."""
    errors = []
    if not block.has_to_line:
        errors.append(_error(block, "trace.incomplete_block", "The block has no To line."))
        return None, errors
    try:
        from_termination = parse_endpoint_line(block.from_text)
        to_termination = parse_endpoint_line(block.to_text)
    except ValueError as exc:
        errors.append(_error(block, "trace.incomplete_block", str(exc)))
        return None, errors
    return (
        EndpointSummary(
            from_termination=from_termination,
            to_termination=to_termination,
            from_text=block.from_text,
            to_text=block.to_text,
        ),
        errors,
    )


def _termination_from_path(values: tuple[object, ...], offset: int) -> TerminationReference:
    """Return the Termination Reference one side of a Segment Evidence row states."""
    termination = TerminationReference(
        port=source_text(values[offset]),
        port_class=source_text(values[offset + 1]),
        cards=source_text(values[offset + 2]),
        device=source_text(values[offset + 3]),
        u_position=source_text(values[offset + 4]),
        rack=source_text(values[offset + 5]),
        location=source_text(values[offset + 6]),
    )
    if not termination.device or not termination.port or not termination.port_class:
        raise ValueError("A Segment Evidence row has an empty device, port, or PortClass.")
    return termination


def _parse_segments(block: _Block) -> tuple[tuple[_ParsedSegment, ...], list[SourceDiagnostic]]:
    """Return the Segment Evidence a path block states, and the rows it could not read."""
    if block.header != _PATH_HEADER:
        return (), [_error(block, "trace.incomplete_block", "The block has no recognized Segment Evidence header.")]
    parsed = []
    errors = []
    for row in block.rows:
        try:
            left = _termination_from_path(row.values, 0)
            cable_class = source_text(row.values[7])
            right = _termination_from_path(row.values, 8)
            if not cable_class:
                raise ValueError("A Segment Evidence row has an empty CableClass.")
        except (IndexError, ValueError) as exc:
            errors.append(_error(block, "trace.incomplete_block", str(exc), row.row_number))
            continue
        parsed.append(_ParsedSegment(SegmentEvidence(left, cable_class, right), row.row_number))
    return tuple(parsed), errors


def _parse_visits(block: _Block) -> tuple[tuple[_ParsedVisit, ...], list[SourceDiagnostic]]:
    """Return the visits a Trace List block states, and the rows it could not read."""
    if block.header != _LIST_HEADER:
        return (), [_error(block, "trace.incomplete_block", "The block has no recognized Trace List header.")]
    parsed = []
    errors = []
    for row in block.rows:
        try:
            termination = TerminationReference(
                location=source_text(row.values[0]),
                rack=source_text(row.values[1]),
                u_position=source_text(row.values[2]),
                device=source_text(row.values[3]),
                cards=source_text(row.values[4]),
                port=source_text(row.values[5]),
                port_class=source_text(row.values[6]),
            )
        except IndexError:
            termination = TerminationReference("", "", "", "")
        if not termination.device or not termination.port or not termination.port_class:
            errors.append(
                _error(
                    block,
                    "trace.corroboration_mismatch",
                    "A Trace List visit has an empty device, port, or PortClass.",
                    row.row_number,
                )
            )
            continue
        parsed.append(_ParsedVisit(termination, row.row_number))
    return tuple(parsed), errors


def _unknown_port_class_error(
    block: _Block,
    terminations: Iterable[tuple[TerminationReference, int]],
) -> SourceDiagnostic | None:
    """Return the first PortClass outside the fixed vocabulary as a diagnostic."""
    for termination, row_number in terminations:
        if termination.port_class not in PORT_CLASSES:
            return _error(
                block,
                "trace.unknown_port_class",
                f"PortClass '{termination.port_class}' is outside the fixed vocabulary.",
                row_number,
            )
    return None


def _linearity_error(
    block: _Block, summary: EndpointSummary, segments: Sequence[_ParsedSegment]
) -> SourceDiagnostic | None:
    """Return the first structural break in a path, or None when the path is linear."""
    # A path states two terminations whether or not it states the segments between them.
    if summary.from_termination.identity_key == summary.to_termination.identity_key:
        return _error(
            block,
            "trace.non_linear_path",
            "The From and To lines name one termination, so the path does not join two terminations.",
            segments[-1].row_number if segments else block.row_start,
        )
    if not segments:
        return None
    if segments[0].evidence.left.identity_key != summary.from_termination.identity_key:
        return _error(
            block,
            "trace.non_linear_path",
            "The first Segment Evidence termination contradicts the From line.",
            segments[0].row_number,
        )
    if segments[-1].evidence.right.identity_key != summary.to_termination.identity_key:
        return _error(
            block,
            "trace.non_linear_path",
            "The last Segment Evidence termination contradicts the To line.",
            segments[-1].row_number,
        )
    seen_segments = set()
    left_destinations: dict[IdentityKey, set[IdentityKey]] = defaultdict(set)
    right_sources: dict[IdentityKey, set[IdentityKey]] = defaultdict(set)
    for parsed in segments:
        left_key = parsed.evidence.left.identity_key
        right_key = parsed.evidence.right.identity_key
        # One cable has no direction, so a reversed row restates the segment rather than adding one.
        pair = tuple(sorted((left_key, right_key)))
        if pair in seen_segments:
            return _error(
                block,
                "trace.non_linear_path",
                "The path repeats a Segment Evidence row.",
                parsed.row_number,
            )
        seen_segments.add(pair)
        left_destinations[left_key].add(right_key)
        right_sources[right_key].add(left_key)
        if len(left_destinations[left_key]) > 1 or len(right_sources[right_key]) > 1:
            return _error(
                block,
                "trace.non_linear_path",
                "The Segment Evidence rows branch.",
                parsed.row_number,
            )
    for previous, following in zip(segments, segments[1:], strict=False):
        if not same_device_and_cards(previous.evidence.right, following.evidence.left):
            return _error(
                block,
                "trace.non_linear_path",
                "Consecutive Segment Evidence rows do not share a device and cards label.",
                following.row_number,
            )
    return None


def _pass_through_claims(segments: Sequence[SegmentEvidence]) -> tuple[PassThroughClaim, ...]:
    """Return the continuation each pair of consecutive segments claims."""
    claims = []
    for previous, following in zip(segments, segments[1:], strict=False):
        if not same_device_and_cards(previous.right, following.left):
            continue
        claims.append(
            PassThroughClaim(
                device=previous.right.device,
                cards=previous.right.cards,
                entry_port=previous.right.port,
                exit_port=following.left.port,
            )
        )
    return tuple(claims)


def _pass_through_error(block: _Block, segments: Sequence[_ParsedSegment]) -> SourceDiagnostic | None:
    """Return a diagnostic for the first Pass-Through Claim at an interface PortClass."""
    for previous, following in zip(segments, segments[1:], strict=False):
        if not same_device_and_cards(previous.evidence.right, following.evidence.left):
            continue
        if (
            previous.evidence.right.port_class in INTERFACE_PORT_CLASSES
            or following.evidence.left.port_class in INTERFACE_PORT_CLASSES
        ):
            return _error(
                block,
                "trace.pass_through_at_interface",
                "A Pass-Through Claim enters or exits through an interface PortClass.",
                following.row_number,
            )
    return None


def _group_consecutive_visits(visits: Sequence[_ParsedVisit]) -> tuple[tuple[_ParsedVisit, ...], ...]:
    """Group Trace List visits into one entry per visited device."""
    groups: list[list[_ParsedVisit]] = []
    for visit in visits:
        device = identity_text(visit.termination.device)
        if not groups or identity_text(groups[-1][0].termination.device) != device:
            groups.append([])
        groups[-1].append(visit)
    return tuple(tuple(group) for group in groups)


def _expected_visits(segments: Sequence[SegmentEvidence]) -> tuple[tuple[TerminationReference, ...], ...]:
    """Return the terminations each device visit of the path presents."""
    if not segments:
        return ()
    visits: list[tuple[TerminationReference, ...]] = [(segments[0].left,)]
    visits.extend((previous.right, following.left) for previous, following in zip(segments, segments[1:], strict=False))
    visits.append((segments[-1].right,))
    return tuple(visits)


def _corroboration_error(
    block: _Block,
    segments: Sequence[SegmentEvidence],
    visits: Sequence[_ParsedVisit],
) -> SourceDiagnostic | None:
    """Return the first contradiction between the Trace List and the path rows."""
    if not visits or not segments:
        return None
    actual = _group_consecutive_visits(visits)
    expected = _expected_visits(segments)
    if segments[-1].right.port_class in REAR_PORT_CLASSES and len(actual) == len(expected) - 1:
        expected = expected[:-1]
    if len(actual) != len(expected):
        return _error(
            block,
            "trace.corroboration_mismatch",
            "The Trace List device sequence does not match the Segment Evidence visits.",
            actual[0][0].row_number,
        )
    for actual_group, expected_group in zip(actual, expected, strict=True):
        allowed = {termination.identity_key for termination in expected_group}
        if any(visit.termination.identity_key not in allowed for visit in actual_group):
            return _error(
                block,
                "trace.corroboration_mismatch",
                "A Trace List termination contradicts its Segment Evidence visit.",
                actual_group[0].row_number,
            )
    return None


def _enrich_termination(
    termination: TerminationReference,
    corroboration: Iterable[TerminationReference],
) -> TerminationReference:
    """Return the termination with the corroboration values the source states for it."""
    enriched = termination
    # A path row is read before a Trace List visit, so an empty cell there must not hide a later one.
    for candidate in corroboration:
        if candidate.identity_key == termination.identity_key:
            enriched = replace(
                enriched,
                u_position=enriched.u_position or candidate.u_position,
                rack=enriched.rack or candidate.rack,
                location=enriched.location or candidate.location,
            )
    return enriched


def _enrich_summary(
    summary: EndpointSummary,
    corroboration: Iterable[TerminationReference],
) -> EndpointSummary:
    """Return the Endpoint Summary with corroboration on both endpoints."""
    corroboration = tuple(corroboration)
    return replace(
        summary,
        from_termination=_enrich_termination(summary.from_termination, corroboration),
        to_termination=_enrich_termination(summary.to_termination, corroboration),
    )


def _provenance(block: _Block, summary: EndpointSummary) -> TraceProvenance:
    """Return where one block stated the trace, and in which direction."""
    direction = (
        "canonical" if summary.from_termination.identity_key <= summary.to_termination.identity_key else "reversed"
    )
    return TraceProvenance(
        workbook_fingerprint=block.workbook_fingerprint,
        sheet=block.sheet,
        block_ordinal=block.ordinal,
        row_start=block.row_start,
        row_end=block.row_end,
        export_timestamp=block.export_timestamp,
        from_text=block.from_text,
        to_text=block.to_text,
        direction=direction,
    )


def _deduplicate_errors(errors: Iterable[SourceDiagnostic]) -> tuple[SourceDiagnostic, ...]:
    """Drop a repeated diagnostic, so every distinct finding keeps its own row."""
    return tuple(dict.fromkeys(errors))


def _path_trace(
    path_block: _Block, list_block: _Block | None
) -> tuple[SourceTrace | None, tuple[SourceDiagnostic, ...]]:
    """Return the Source Trace one path block states, with any block-level diagnostic."""
    summary, errors = _endpoint_summary(path_block)
    errors.extend(_metadata_errors((path_block, list_block)))
    if summary is None:
        return None, tuple(errors)
    parsed_segments, segment_errors = _parse_segments(path_block)
    errors.extend(segment_errors)
    parsed_visits: tuple[_ParsedVisit, ...] = ()
    if list_block is not None:
        parsed_visits, visit_errors = _parse_visits(list_block)
        errors.extend(visit_errors)
    endpoint_terms = (
        (summary.from_termination, path_block.row_start),
        (summary.to_termination, path_block.row_start + 1),
    )
    path_terms = tuple(
        (termination, parsed.row_number)
        for parsed in parsed_segments
        for termination in (parsed.evidence.left, parsed.evidence.right)
    )
    visit_terms = tuple((visit.termination, visit.row_number) for visit in parsed_visits)
    unknown = _unknown_port_class_error(path_block, (*endpoint_terms, *path_terms))
    if unknown is None and list_block is not None:
        unknown = _unknown_port_class_error(list_block, visit_terms)
    if unknown is not None:
        errors.append(unknown)
    incomplete = any(error.code == "trace.incomplete_block" for error in errors)
    linearity = None if incomplete else _linearity_error(path_block, summary, parsed_segments)
    if linearity is not None:
        errors.append(linearity)
    if not incomplete and linearity is None:
        pass_through = _pass_through_error(path_block, parsed_segments)
        if pass_through is not None:
            errors.append(pass_through)
        if list_block is not None:
            mismatch = _corroboration_error(
                list_block,
                tuple(parsed.evidence for parsed in parsed_segments),
                parsed_visits,
            )
            if mismatch is not None:
                errors.append(mismatch)
    stated_segments = tuple(parsed.evidence for parsed in parsed_segments)
    visits = tuple(visit.termination for visit in parsed_visits)
    path_corroboration = tuple(
        termination for segment in stated_segments for termination in (segment.left, segment.right)
    )
    summary = _enrich_summary(summary, (*path_corroboration, *visits))
    segments = canonical_orientation(summary.from_termination, summary.to_termination, stated_segments)
    provenance = [_provenance(path_block, summary)]
    if list_block is not None:
        provenance.append(_provenance(list_block, summary))
    return (
        SourceTrace(
            endpoint_summary=summary,
            segments=segments,
            pass_through_claims=_pass_through_claims(segments),
            corroboration=visits,
            identity=canonical_trace_identity(summary.from_termination, summary.to_termination),
            content_fingerprint=content_fingerprint(
                summary.from_termination,
                summary.to_termination,
                stated_segments,
            ),
            provenance=tuple(provenance),
            errors=_deduplicate_errors(errors),
        ),
        (),
    )


def _fallback_trace(block: _Block) -> tuple[SourceTrace | None, tuple[SourceDiagnostic, ...]]:
    """Return the Endpoint Summary fallback an unpaired Trace List block states."""
    summary, errors = _endpoint_summary(block)
    errors.extend(_metadata_errors((block,)))
    visits, visit_errors = _parse_visits(block)
    errors.extend(visit_errors)
    if summary is None or not visits:
        return None, tuple(errors)
    terms = (
        (summary.from_termination, block.row_start),
        (summary.to_termination, block.row_start + 1),
        *((visit.termination, visit.row_number) for visit in visits),
    )
    unknown = _unknown_port_class_error(block, terms)
    if unknown is not None:
        errors.append(unknown)
    loop = _linearity_error(block, summary, ())
    if loop is not None:
        errors.append(loop)
    corroboration = tuple(visit.termination for visit in visits)
    summary = _enrich_summary(summary, corroboration)
    return (
        SourceTrace(
            endpoint_summary=summary,
            segments=(),
            pass_through_claims=(),
            corroboration=corroboration,
            identity=canonical_trace_identity(summary.from_termination, summary.to_termination),
            content_fingerprint=content_fingerprint(summary.from_termination, summary.to_termination, ()),
            provenance=(_provenance(block, summary),),
            errors=_deduplicate_errors(errors),
        ),
        (),
    )


def _trace_selection_key(trace: SourceTrace) -> tuple[bool, bool, str]:
    """Rank Segment Evidence first, then canonical direction and serialized occurrence content."""
    summary = trace.endpoint_summary
    return (
        not bool(trace.segments),
        summary.from_termination.identity_key > summary.to_termination.identity_key,
        json.dumps(asdict(trace), ensure_ascii=False, separators=(",", ":"), sort_keys=True),
    )


def _collapse_duplicates(traces: Sequence[SourceTrace]) -> tuple[SourceTrace, ...]:
    """Collapse occurrences that share an identity, and flag differing evidence."""
    by_identity: dict[str, list[SourceTrace]] = {}
    for trace in traces:
        by_identity.setdefault(trace.identity, []).append(trace)
    collapsed = []
    for occurrences in by_identity.values():
        ordered = sorted(occurrences, key=_trace_selection_key)
        selected = ordered[0]
        provenance = tuple(dict.fromkeys(item for trace in ordered for item in trace.provenance))
        # The fingerprint excludes Trace List data, so a later occurrence can state its own finding.
        errors = _first_of_each_code(ordered)
        # An endpoint-only fallback states no segments, so it contradicts no segment evidence.
        compared = [trace for trace in ordered if trace.segments] or ordered
        if len({trace.content_fingerprint for trace in compared}) > 1:
            locations = "; ".join(_location_from_provenance(trace.provenance[0]) for trace in compared)
            errors.append(
                SourceDiagnostic(
                    code="trace.duplicate_conflict",
                    message=f"Source Trace occurrences have differing evidence: {locations}.",
                    row_number=min(item.row_start for item in provenance),
                )
            )
        collapsed.append(replace(selected, provenance=provenance, errors=_deduplicate_errors(errors)))
    return tuple(collapsed)


def _first_of_each_code(occurrences: Sequence[SourceTrace]) -> list[SourceDiagnostic]:
    """Return every occurrence's findings, reporting one repeated condition once."""
    errors: list[SourceDiagnostic] = []
    seen_codes: set[str] = set()
    for occurrence in occurrences:
        for error in occurrence.errors:
            if error.code in seen_codes and occurrence is not occurrences[0]:
                continue
            seen_codes.add(error.code)
            errors.append(error)
    return errors


def _location_from_provenance(provenance: TraceProvenance) -> str:
    """Return the readable source position one provenance record holds."""
    return f"{provenance.sheet} block {provenance.block_ordinal} (rows {provenance.row_start}-{provenance.row_end})"


def _cross_trace_conflicts(traces: Sequence[SourceTrace]) -> tuple[SourceTrace, ...]:
    """Flag every trace that shares a termination or disagrees about a CableClass."""
    termination_claims: dict[IdentityKey, dict[int, set[_SegmentClaim]]] = defaultdict(lambda: defaultdict(set))
    segment_classes: dict[tuple[IdentityKey, IdentityKey], dict[str, set[int]]] = defaultdict(lambda: defaultdict(set))
    for index, trace in enumerate(traces):
        claims: dict[IdentityKey, set[_SegmentClaim]] = defaultdict(set)
        for segment in trace.segments:
            ordered_pair = sorted((segment.left.identity_key, segment.right.identity_key))
            segment_pair = ordered_pair[0], ordered_pair[1]
            # The CableClass label keys its own mapping row, so it compares as the source states it.
            cable_class = segment.cable_class
            claim = segment_pair, cable_class
            claims[segment.left.identity_key].add(claim)
            claims[segment.right.identity_key].add(claim)
            segment_classes[segment_pair][cable_class].add(index)
        claims.setdefault(trace.endpoint_summary.from_termination.identity_key, set())
        claims.setdefault(trace.endpoint_summary.to_termination.identity_key, set())
        for termination, trace_claims in claims.items():
            termination_claims[termination][index].update(trace_claims)
    conflicts: dict[int, set[str]] = defaultdict(set)
    for claims_by_owner in termination_claims.values():
        owner_claims = list(claims_by_owner.values())
        identical_shared_segment = bool(owner_claims[0]) and all(claims == owner_claims[0] for claims in owner_claims)
        if len(claims_by_owner) > 1 and not identical_shared_segment:
            for owner in claims_by_owner:
                conflicts[owner].add("a termination is claimed by another Source Trace")
    for classes in segment_classes.values():
        owners = set().union(*classes.values())
        if len(classes) > 1 and len(owners) > 1:
            for owner in owners:
                conflicts[owner].add("a shared segment has conflicting CableClass labels")
    checked = []
    for index, trace in enumerate(traces):
        if index not in conflicts:
            checked.append(trace)
            continue
        provenance = trace.provenance[0]
        detail = "; ".join(sorted(conflicts[index]))
        error = SourceDiagnostic(
            code="trace.cross_trace_conflict",
            message=f"{_location_from_provenance(provenance)}: {detail}.",
            row_number=provenance.row_start,
        )
        checked.append(replace(trace, errors=_deduplicate_errors((*trace.errors, error))))
    return tuple(checked)


def interpret(content: bytes) -> tuple[tuple[SourceTrace, ...], tuple[SourceDiagnostic, ...]]:
    """Return typed Source Traces and source diagnostics from workbook bytes."""
    try:
        book = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise SourceUnreadable(f"Cannot open Excel file: {exc}") from exc
    try:
        recognized = [name for name in (TRACE_PATH_SHEET, TRACE_LIST_SHEET) if name in book.sheetnames]
        if not recognized:
            diagnostic = SourceDiagnostic(
                code="trace.no_recognized_sheet",
                message=f"Workbook has neither '{TRACE_PATH_SHEET}' nor '{TRACE_LIST_SHEET}'.",
            )
            return (), (diagnostic,)
        workbook_fingerprint = sha256(content).hexdigest()
        path_blocks = (
            _extract_blocks(book[TRACE_PATH_SHEET], workbook_fingerprint) if TRACE_PATH_SHEET in recognized else ()
        )
        list_blocks = (
            _extract_blocks(book[TRACE_LIST_SHEET], workbook_fingerprint) if TRACE_LIST_SHEET in recognized else ()
        )
    finally:
        # close() releases an archive only in read-only mode, but it stays paired with the load.
        book.close()
    lists_by_pair: dict[tuple[str, str], list[_Block]] = defaultdict(list)
    for block in list_blocks:
        lists_by_pair[block.pair_key].append(block)
    traces: list[SourceTrace] = []
    diagnostics: list[SourceDiagnostic] = []
    paired_lists: set[tuple[str, int]] = set()
    for path_block in path_blocks:
        candidates = lists_by_pair[path_block.pair_key]
        list_block = candidates.pop(0) if candidates else None
        if list_block is not None:
            paired_lists.add((list_block.sheet, list_block.ordinal))
        trace, block_diagnostics = _path_trace(path_block, list_block)
        diagnostics.extend(block_diagnostics)
        if trace is not None:
            traces.append(trace)
    for block in list_blocks:
        if (block.sheet, block.ordinal) in paired_lists:
            continue
        trace, block_diagnostics = _fallback_trace(block)
        diagnostics.extend(block_diagnostics)
        if trace is not None:
            traces.append(trace)
    checked = _cross_trace_conflicts(_collapse_duplicates(traces))
    diagnostics.extend(error for trace in checked for error in trace.errors)
    return checked, tuple(diagnostics)


__all__ = (
    "EndpointSummary",
    "FRONT_PORT_CLASSES",
    "INTERFACE_PORT_CLASSES",
    "PORT_CLASSES",
    "PassThroughClaim",
    "REAR_PORT_CLASSES",
    "SegmentEvidence",
    "SourceTrace",
    "TerminationReference",
    "TraceProvenance",
    "canonical_orientation",
    "canonical_trace_identity",
    "content_fingerprint",
    "interpret",
    "parse_endpoint_line",
)
