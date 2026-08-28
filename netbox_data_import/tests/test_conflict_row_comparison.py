# SPDX-License-Identifier: Apache-2.0
# SPDX-FileCopyrightText: 2026 Marcin Zieba <marcinpsk@gmail.com>
"""Conflicting import rows show their source facts together in the preview."""

import io
import re
from html import unescape

import openpyxl
from django.urls import reverse
from django.utils.html import strip_tags

from netbox_data_import.tests.test_views import BaseViewTestCase, _make_profile


class ConflictRowComparisonTest(BaseViewTestCase):
    """The preview compares each error row with every source row that it names."""

    def _workbook(self):
        """Return a workbook with every supported within-import identity conflict."""
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = "Data"
        sheet.append(
            [
                "Id",
                "Rack",
                "Name",
                "Class",
                "Make",
                "Model",
                "UHeight",
                "UPosition",
                "Side",
                "Airflow",
                "Serial Number",
                "Asset Tag",
                "Status",
            ]
        )
        sheet.append(
            [
                "serial-source-a",
                "Rack-Compare",
                "serial-device-a",
                "Server",
                "Example Vendor",
                "Example Model",
                1,
                1,
                "front",
                "",
                "SERIAL-DUPLICATE",
                "ASSET-SERIAL-A",
                "active",
            ]
        )
        sheet.append(
            [
                "serial-source-b",
                "Rack-Compare",
                "serial-device-b",
                "Server",
                "Example Vendor",
                "Example Model",
                1,
                2,
                "front",
                "",
                "SERIAL-DUPLICATE",
                "ASSET-SERIAL-B",
                "active",
            ]
        )
        sheet.append(
            [
                "source-duplicate",
                "Rack-Compare",
                "source-device-a",
                "Server",
                "Example Vendor",
                "Example Model",
                1,
                3,
                "front",
                "",
                "SERIAL-SOURCE-A",
                "ASSET-SOURCE-A",
                "active",
            ]
        )
        sheet.append(
            [
                "source-duplicate",
                "Source-Rack-B",
                "Source-Rack-B",
                "Cabinet",
                "Example Vendor",
                "Example Model",
                1,
                4,
                "front",
                "",
                "SERIAL-SOURCE-B",
                "ASSET-SOURCE-B",
                "active",
            ]
        )
        sheet.append(
            [
                "asset-source-a",
                "Rack-Compare",
                "asset-device-a",
                "Server",
                "Example Vendor",
                "Example Model",
                1,
                5,
                "front",
                "",
                "SERIAL-ASSET-A",
                "ASSET-DUPLICATE",
                "active",
            ]
        )
        sheet.append(
            [
                "asset-source-b",
                "Rack-Compare",
                "asset-device-b",
                "Server",
                "Example Vendor",
                "Example Model",
                1,
                6,
                "front",
                "",
                "SERIAL-ASSET-B",
                "ASSET-DUPLICATE",
                "active",
            ]
        )
        sheet.append(
            [
                "rack-source-a",
                "Rack-Compare",
                "rack-device-a",
                "Server",
                "Example Vendor",
                "Example Model",
                1,
                7,
                "front",
                "",
                "SERIAL-RACK-A",
                "ASSET-RACK-A",
                "active",
            ]
        )
        sheet.append(
            [
                "rack-source-b",
                "Rack-Compare",
                "rack-device-b",
                "Server",
                "Example Vendor",
                "Example Model",
                1,
                7,
                "front",
                "",
                "SERIAL-RACK-B",
                "ASSET-RACK-B",
                "active",
            ]
        )
        workbook = io.BytesIO()
        book.save(workbook)
        workbook.seek(0)
        return workbook

    def _numeric_class_workbook(self):
        """Return duplicate source rows whose class cells contain a number."""
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = "Data"
        sheet.append(
            [
                "Id",
                "Rack",
                "Name",
                "Class",
                "Make",
                "Model",
                "UHeight",
                "UPosition",
                "Side",
                "Serial Number",
                "Asset Tag",
            ]
        )
        sheet.append(
            [
                "numeric-source-duplicate",
                "Rack-Compare",
                "numeric-device-a",
                1,
                "Example Vendor",
                "Example Model",
                1,
                10,
                "front",
                "SERIAL-NUMERIC-A",
                "ASSET-NUMERIC-A",
            ]
        )
        sheet.append(
            [
                "numeric-source-duplicate",
                "Rack-Compare",
                "numeric-device-b",
                1,
                "Example Vendor",
                "Example Model",
                1,
                11,
                "front",
                "SERIAL-NUMERIC-B",
                "ASSET-NUMERIC-B",
            ]
        )
        workbook = io.BytesIO()
        book.save(workbook)
        workbook.seek(0)
        return workbook

    def _preview(self, workbook=None, additional_source_class=None):
        """Run a real dry-run import and return the rendered preview response."""
        from dcim.models import DeviceRole, DeviceType, Manufacturer, Rack, Site

        from netbox_data_import.engine import parse_file, run_import
        from netbox_data_import.models import ClassRoleMapping
        from netbox_data_import.views import _serialize_rows

        site = Site.objects.create(name="Comparison Site", slug="comparison-site")
        Rack.objects.create(name="Rack-Compare", site=site, u_height=42)
        DeviceRole.objects.create(name="Server", slug="server")
        manufacturer = Manufacturer.objects.create(name="Example Vendor", slug="example-vendor")
        DeviceType.objects.create(
            manufacturer=manufacturer,
            model="Example Model",
            slug="example-vendor-example-model",
            u_height=1,
            is_full_depth=False,
        )
        profile = _make_profile("Conflict comparison")
        if additional_source_class is not None:
            ClassRoleMapping.objects.create(
                profile=profile,
                source_class=additional_source_class,
                creates_rack=False,
                role_slug="server",
            )
        rows = parse_file(workbook or self._workbook(), profile)
        result = run_import(rows, profile, {"site": site}, dry_run=True)

        session = self.client.session
        session["import_result"] = result.to_session_dict()
        session["import_rows"] = _serialize_rows(rows)
        session["import_context"] = {
            "profile_id": profile.pk,
            "site_id": site.pk,
            "location_id": None,
            "tenant_id": None,
            "filename": "conflicts.xlsx",
        }
        session.save()
        response = self.client.get(reverse("plugins:netbox_data_import:import_preview"))
        self.assertEqual(response.status_code, 200)
        return response

    def _resolution_workbook(self):
        """Return two rows whose serials differ until a saved resolution makes them collide."""
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = "Data"
        sheet.append(["Id", "Rack", "Name", "Class", "Make", "Model", "UHeight", "UPosition", "Serial Number"])
        sheet.append(
            ["res-a", "Rack-Compare", "res-device-a", "Server", "Example Vendor", "Example Model", 1, 1, "SERIAL-A"]
        )
        sheet.append(
            ["res-b", "Rack-Compare", "res-device-b", "Server", "Example Vendor", "Example Model", 1, 2, "SERIAL-B"]
        )
        workbook = io.BytesIO()
        book.save(workbook)
        workbook.seek(0)
        return workbook

    def _materialized_preview(self):
        """Render the preview from the stored result, the way the setup step reaches it once."""
        from dcim.models import DeviceRole, DeviceType, Manufacturer, Rack, Site

        from netbox_data_import.engine import derive_effective_rows, parse_file, run_import
        from netbox_data_import.models import SourceResolution
        from netbox_data_import.preview_row_actions import PREVIEW_USE_MATERIALIZED_ONCE_SESSION_KEY
        from netbox_data_import.views import _serialize_rows

        site = Site.objects.create(name="Resolution Site", slug="resolution-site")
        Rack.objects.create(name="Rack-Compare", site=site, u_height=42)
        DeviceRole.objects.create(name="Server", slug="server")
        manufacturer = Manufacturer.objects.create(name="Example Vendor", slug="example-vendor")
        DeviceType.objects.create(
            manufacturer=manufacturer, model="Example Model", slug="example-vendor-example-model", u_height=1
        )
        profile = _make_profile("Resolution comparison")
        SourceResolution.objects.create(
            profile=profile,
            source_id="res-a",
            source_column="serial",
            original_value="SERIAL-A",
            resolved_fields={"serial": "SERIAL-B"},
        )
        rows = parse_file(self._resolution_workbook(), profile)
        result = run_import(derive_effective_rows(rows, profile), profile, {"site": site}, dry_run=True)

        session = self.client.session
        session["import_result"] = result.to_session_dict()
        session["import_rows"] = _serialize_rows(rows)
        session["import_context"] = {
            "profile_id": profile.pk,
            "site_id": site.pk,
            "location_id": None,
            "tenant_id": None,
            "filename": "resolutions.xlsx",
        }
        session[PREVIEW_USE_MATERIALIZED_ONCE_SESSION_KEY] = True
        session.save()
        response = self.client.get(reverse("plugins:netbox_data_import:import_preview"))
        self.assertEqual(response.status_code, 200)
        return response

    def test_the_comparison_shows_the_value_that_caused_the_conflict(self):
        """The stored result is built from resolved rows, so the table cannot read the raw ones."""
        response = self._materialized_preview()
        line = self._comparison_line(self._detail_for(response, "res-device-a"), "res-device-a")
        self.assertIn("SERIAL-B", line)
        self.assertNotIn("SERIAL-A", line)

    def _detail_for(self, response, name):
        """Return one named device row's detail markup."""
        rows = list(response.context["result"].rows)
        index = next(
            index for index, row in enumerate(rows, start=1) if row.object_type == "device" and row.name == name
        )
        html = response.content.decode()
        start = html.index(f'<tr id="diff-{index}" class="ndi-diff-row"')
        end = html.find('<tr id="prow-', start)
        if end == -1:
            end = html.index('id="previewNoFilterResults"', start)
        return html[start:end]

    def _comparison_line(self, detail, name):
        """Return normalized text from the comparison line for one source row."""
        lines = re.findall(r'<tr class="ndi-conflict-comparison-row[^"]*">(.*?)</tr>', detail, re.DOTALL)
        name_cell = re.compile(rf"<td>\s*{re.escape(name)}\s*</td>")
        matching = [line for line in lines if name_cell.search(line)]
        self.assertEqual(len(matching), 1, f"the comparison must contain one line for {name}")
        return " ".join(unescape(strip_tags(matching[0])).split())

    def test_each_conflict_detail_compares_every_involved_source_row(self):
        """Each supported conflict gives the operator all facts needed to choose a row."""
        response = self._preview()
        cases = [
            {
                "current": (
                    "serial-device-b",
                    "Row 3",
                    "serial-source-b",
                    "SERIAL-DUPLICATE",
                    "ASSET-SERIAL-B",
                    "Rack-Compare, U2, front",
                    "Duplicate serial 'SERIAL-DUPLICATE' appears more than once in this import, also on row 2.",
                ),
                "other": (
                    "serial-device-a",
                    "Row 2",
                    "serial-source-a",
                    "SERIAL-DUPLICATE",
                    "ASSET-SERIAL-A",
                    "Rack-Compare, U1, front",
                    "Duplicate serial 'SERIAL-DUPLICATE' appears more than once in this import, also on row 3.",
                ),
                "actions": ("Error", "Error"),
            },
            {
                "current": (
                    "source-device-a",
                    "Row 4",
                    "source-duplicate",
                    "SERIAL-SOURCE-A",
                    "ASSET-SOURCE-A",
                    "Rack-Compare, U3, front",
                    "Duplicate source ID 'source-duplicate' appears more than once in this import, also on row 5.",
                ),
                "other": (
                    "Source-Rack-B",
                    "Row 5",
                    "source-duplicate",
                    "SERIAL-SOURCE-B",
                    "ASSET-SOURCE-B",
                    "Source-Rack-B, U4, front",
                    "Duplicate source ID 'source-duplicate' appears more than once in this import, also on row 4.",
                ),
                "actions": ("Error", "Error"),
            },
            {
                "current": (
                    "asset-device-b",
                    "Row 7",
                    "asset-source-b",
                    "SERIAL-ASSET-B",
                    "ASSET-DUPLICATE",
                    "Rack-Compare, U6, front",
                    "Duplicate asset tag 'ASSET-DUPLICATE' appears more than once in this import, also on row 6.",
                ),
                "other": (
                    "asset-device-a",
                    "Row 6",
                    "asset-source-a",
                    "SERIAL-ASSET-A",
                    "ASSET-DUPLICATE",
                    "Rack-Compare, U5, front",
                    "Duplicate asset tag 'ASSET-DUPLICATE' appears more than once in this import, also on row 7.",
                ),
                "actions": ("Error", "Error"),
            },
            {
                "current": (
                    "rack-device-b",
                    "Row 9",
                    "rack-source-b",
                    "SERIAL-RACK-B",
                    "ASSET-RACK-B",
                    "Rack-Compare, U7, front",
                    "Rack position conflict: Rack-Compare U7 (front) also claimed by row 8 ('rack-device-a')",
                ),
                "other": (
                    "rack-device-a",
                    "Row 8",
                    "rack-source-a",
                    "SERIAL-RACK-A",
                    "ASSET-RACK-A",
                    "Rack-Compare, U7, front",
                    "Would create device 'rack-device-a' in Rack-Compare U7",
                ),
                "actions": ("Error", "Create"),
            },
        ]

        for case in cases:
            with self.subTest(current=case["current"][0]):
                detail = self._detail_for(response, case["current"][0])
                current_line = self._comparison_line(detail, case["current"][0])
                other_line = self._comparison_line(detail, case["other"][0])
                for fact in case["current"]:
                    self.assertIn(fact, current_line)
                for fact in case["other"]:
                    self.assertIn(fact, other_line)
                self.assertIn(case["actions"][0], current_line)
                self.assertIn(case["actions"][1], other_line)
                self.assertIn("Current row", current_line)
                self.assertNotIn("Current row", other_line)

    def test_numeric_class_codes_keep_the_conflict_comparison_renderable(self):
        """A numeric class cell that matches a text policy must not take down the preview."""
        response = self._preview(self._numeric_class_workbook(), additional_source_class="1")
        detail = self._detail_for(response, "numeric-device-b")

        current_line = self._comparison_line(detail, "numeric-device-b")
        other_line = self._comparison_line(detail, "numeric-device-a")

        self.assertIn("Row 3", current_line)
        self.assertIn("Row 2", other_line)
        self.assertIn("Current row", current_line)


class ConflictComparisonCarriesItsActionTest(ConflictRowComparisonTest):
    """A comparison that only states facts leaves the operator to find the control elsewhere."""

    def _comparison_row_html(self, detail, name):
        """Return the raw markup of the comparison line for one source row."""
        lines = re.findall(r'<tr class="ndi-conflict-comparison-row[^"]*">(.*?)</tr>', detail, re.DOTALL)
        name_cell = re.compile(rf"<td>\s*{re.escape(name)}\s*</td>")
        matching = [line for line in lines if name_cell.search(line)]
        self.assertEqual(len(matching), 1, f"the comparison must contain one line for {name}")
        return matching[0]

    def test_every_row_in_a_serial_collision_can_give_the_serial_up(self):
        """The engine marks both rows, so the operator picks which one loses the serial."""
        response = self._preview()
        detail = self._detail_for(response, "serial-device-b")
        for name, row_number in (("serial-device-b", 3), ("serial-device-a", 2)):
            line = self._comparison_row_html(detail, name)
            self.assertIn("ndi-conflict-comparison-action", line, name)
            self.assertIn(f'name="row_number" value="{row_number}"', line, name)
            self.assertIn("Ignore serial", line, name)

    def test_a_conflict_with_no_safe_endpoint_offers_no_action(self):
        """Offering a control that cannot resolve the conflict is worse than offering none."""
        response = self._preview()
        detail = self._detail_for(response, "source-device-a")
        for name in ("source-device-a", "Source-Rack-B"):
            line = self._comparison_row_html(detail, name)
            self.assertNotIn("ndi-conflict-comparison-action", line, name)
