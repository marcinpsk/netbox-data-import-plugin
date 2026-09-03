# SPDX-License-Identifier: Apache-2.0
# SPDX-FileCopyrightText: 2026 Marcin Zieba <marcinpsk@gmail.com>
"""Exercise the trace Source Adapter with no database access."""

from io import BytesIO
import json
from pathlib import Path
from unittest import mock

from django.test import SimpleTestCase
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from netbox_data_import.adapters import SourceBatch, SourceUnreadable, TraceWorkbookAdapter
from netbox_data_import.catalog import OutputKind
from netbox_data_import import trace_workbook
from netbox_data_import.trace_workbook import parse_endpoint_line

FIXTURES = Path(__file__).parent / "fixtures"

PATH_HEADER = (
    "Port",
    "PortClass",
    "Cards",
    "Device",
    "UPos",
    "Rack",
    "Location",
    "CableClass",
    "Port",
    "PortClass",
    "Cards",
    "Device",
    "UPos",
    "Rack",
    "Location",
)
LIST_HEADER = ("Location", "Rack", "UPos", "Device", "Cards", "Port", "PortClass", "Cable")


def _termination(device, cards, port, port_class):
    """Return one compact termination tuple for an in-memory workbook."""
    return device, cards, port, port_class


def _endpoint_line(termination):
    """Render one endpoint line in the source format."""
    device, cards, port, port_class = termination
    parts = [device]
    if cards:
        parts.append(cards)
    parts.append(f"{port} ({port_class})")
    return " > ".join(parts)


def _segment(left, cable_class, right, corroboration=("", "", "")):
    """Render one Segment Evidence row in the source column order."""
    left_device, left_cards, left_port, left_class = left
    right_device, right_cards, right_port, right_class = right
    return (
        left_port,
        left_class,
        left_cards,
        left_device,
        *corroboration,
        cable_class,
        right_port,
        right_class,
        right_cards,
        right_device,
        *corroboration,
    )


def _visit(termination):
    """Render one Trace List visit row."""
    device, cards, port, port_class = termination
    return "", "", "", device, cards, port, port_class, "Ignored"


def _add_sheet(book, name, header, blocks):
    """Add one trace sheet with the supplied source blocks."""
    sheet = book.create_sheet(name)
    sheet.append(("Executed", "2026-08-31 12:00:00"))
    sheet.append(())
    for from_line, to_line, rows in blocks:
        sheet.append(("From", from_line))
        sheet.append(("To", to_line))
        sheet.append(header)
        for row in rows:
            sheet.append(row)
    return sheet


def _workbook(*, path_blocks=(), list_blocks=(), include_path=True, include_list=False):
    """Build workbook bytes with fixed trace sheet names."""
    book = openpyxl.Workbook()
    active = book.active
    if isinstance(active, Worksheet):
        book.remove(active)
    if include_path:
        _add_sheet(book, "Trace From To", PATH_HEADER, path_blocks)
    if include_list:
        _add_sheet(book, "Trace List", LIST_HEADER, list_blocks)
    buffer = BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def _two_traces_over_one_pair(first_class, second_class, *, shared_entry=False):
    """Return two path blocks that cross one rear-port segment, each from its own endpoints."""
    shared_left = _termination("PANEL-P", "CARD-P", "REAR", "Punch-Down")
    shared_right = _termination("PANEL-Q", "CARD-Q", "REAR", "Punch-Down")
    blocks = []
    for index, cable_class in enumerate((first_class, second_class), start=1):
        front = "FRONT" if shared_entry else f"FRONT-{index}"
        endpoint = _termination(f"DEVICE-{index}", "", f"PORT-{index}", "Port")
        far_endpoint = _termination(f"DEVICE-{index}0", "", f"PORT-{index}0", "NIC")
        blocks.append(
            (
                _endpoint_line(endpoint),
                _endpoint_line(far_endpoint),
                (
                    _segment(endpoint, "Lead", _termination("PANEL-P", "CARD-P", front, "Position Front")),
                    _segment(shared_left, cable_class, shared_right),
                    _segment(_termination("PANEL-Q", "CARD-Q", front, "Position Front"), "Tail", far_endpoint),
                ),
            )
        )
    return tuple(blocks)


def _interpret(content):
    """Run the public Source Adapter seam with its empty configuration."""
    return TraceWorkbookAdapter.interpret(content, {})


def _codes(batch):
    """Return all batch diagnostic codes in source order."""
    return [diagnostic.code for diagnostic in batch.diagnostics]


class TraceWorkbookFixtureTest(SimpleTestCase):
    """The committed workbooks define the real trace format."""

    def test_copper_fixture_collapses_duplicate_blocks(self):
        """The copper corpus produces ten valid three-segment Source Traces."""
        content = (FIXTURES / "trace_copper.xlsx").read_bytes()

        batch = _interpret(content)

        self.assertIsInstance(batch, SourceBatch)
        self.assertEqual(batch.output_kinds, frozenset({OutputKind.SOURCE_TRACE}))
        self.assertEqual(len(batch.rows), 10)
        self.assertTrue(all(trace.valid for trace in batch.rows))
        self.assertEqual({len(trace.segments) for trace in batch.rows}, {3})
        self.assertEqual({len(trace.provenance) for trace in batch.rows}, {4})
        self.assertEqual(batch.diagnostics, ())

    def test_fiber_fixture_preserves_valid_and_invalid_traces(self):
        """The fiber corpus keeps its two structural failures and every other trace."""
        content = (FIXTURES / "trace_fiber.xlsx").read_bytes()

        batch = _interpret(content)

        valid = [trace for trace in batch.rows if trace.valid]
        self.assertEqual(len(batch.rows), 10)
        self.assertEqual(len(valid), 8)
        self.assertTrue(all(4 <= len(trace.segments) <= 9 for trace in valid))
        self.assertEqual(sum(trace.ends_at_rear_port for trace in valid), 4)
        self.assertEqual(_codes(batch).count("trace.non_linear_path"), 1)
        self.assertEqual(_codes(batch).count("trace.pass_through_at_interface"), 1)
        self.assertNotIn("trace.corroboration_mismatch", _codes(batch))
        self.assertNotIn("trace.duplicate_conflict", _codes(batch))
        self.assertNotIn("trace.cross_trace_conflict", _codes(batch))
        self.assertEqual(batch.diagnostics, tuple(error for trace in batch.rows for error in trace.errors))
        interface_trace = next(
            trace
            for trace in batch.rows
            if any(error.code == "trace.pass_through_at_interface" for error in trace.errors)
        )
        self.assertTrue(
            any(claim.entry_port == claim.exit_port == "R01" for claim in interface_trace.pass_through_claims)
        )
        self.assertEqual(interface_trace.corroboration[-1].device, "DEV-07")

    def test_fixture_diagnostics_never_cross_the_target_boundary(self):
        """The Source Adapter never emits a diagnostic that needs NetBox state."""
        forbidden = {
            "cable.unsupported_termination_kind",
            "trace.device_unresolved",
            "trace.endpoint_evidence_only",
        }

        for fixture in ("trace_copper.xlsx", "trace_fiber.xlsx"):
            batch = _interpret((FIXTURES / fixture).read_bytes())
            self.assertTrue(forbidden.isdisjoint(_codes(batch)))

    def test_fixture_provenance_captures_the_workbook_and_both_sheets(self):
        """Collapsed occurrences retain their source positions and export timestamp."""
        batch = _interpret((FIXTURES / "trace_copper.xlsx").read_bytes())
        trace = batch.rows[0]

        self.assertEqual({item.sheet for item in trace.provenance}, {"Trace From To", "Trace List"})
        self.assertEqual({item.export_timestamp for item in trace.provenance}, {"2026-08-16 21:28:51"})
        self.assertEqual({len(item.workbook_fingerprint) for item in trace.provenance}, {64})
        self.assertTrue(all(item.from_text and item.to_text for item in trace.provenance))


class TraceWorkbookIdentityTest(SimpleTestCase):
    """Trace identity and content use direction-independent canonical forms."""

    def test_endpoint_lines_keep_cards_and_port_classes_distinct(self):
        """The endpoint grammar accepts every source shape used by the fixtures."""
        examples = {
            "DEV-02 > 08 (Switch Port)": ("DEV-02", "", "08", "Switch Port"),
            "DEV-01 > Slot 0 > 2/1 (Port)": ("DEV-01", "Slot 0", "2/1", "Port"),
            "DEV-10 > Switch A - MOD-01 > Eth 01 (NIC)": ("DEV-10", "Switch A - MOD-01", "Eth 01", "NIC"),
            "H1-Q10-52-F > H1-Q10-52-1-F - H1-Q7-2-1-F > R03 (Fiber Pair Back)": (
                "H1-Q10-52-F",
                "H1-Q10-52-1-F - H1-Q7-2-1-F",
                "R03",
                "Fiber Pair Back",
            ),
            "DEV-12 > SFP - 4 (Switch Port)": ("DEV-12", "", "SFP - 4", "Switch Port"),
        }

        for line, expected in examples.items():
            with self.subTest(line=line):
                termination = parse_endpoint_line(line)
                self.assertEqual(
                    (termination.device, termination.cards, termination.port, termination.port_class),
                    expected,
                )

    def test_a_reversed_restatement_collapses_without_fingerprint_churn(self):
        """Direction changes provenance but not trace identity or content."""
        endpoint_a = _termination("DEVICE-A", "CARD-A", "PORT-A", "Port")
        panel_entry = _termination("PANEL-A", "CARD-P", "FRONT-A", "Position Front")
        panel_exit = _termination("PANEL-A", "CARD-P", "REAR-A", "Punch-Down")
        endpoint_b = _termination("DEVICE-B", "", "PORT-B", "NIC")
        forward = (
            _endpoint_line(endpoint_a),
            _endpoint_line(endpoint_b),
            (_segment(endpoint_a, "Cable A", panel_entry), _segment(panel_exit, "Cable B", endpoint_b)),
        )
        reverse = (
            _endpoint_line(endpoint_b),
            _endpoint_line(endpoint_a),
            (_segment(endpoint_b, "Cable B", panel_exit), _segment(panel_entry, "Cable A", endpoint_a)),
        )

        forward_trace = _interpret(_workbook(path_blocks=(forward,))).rows[0]
        reverse_trace = _interpret(_workbook(path_blocks=(reverse,))).rows[0]
        combined = _interpret(_workbook(path_blocks=(forward, reverse)))

        self.assertEqual(forward_trace.identity, reverse_trace.identity)
        self.assertEqual(forward_trace.content_fingerprint, reverse_trace.content_fingerprint)
        self.assertEqual(len(combined.rows), 1)
        self.assertEqual(len(combined.rows[0].provenance), 2)
        self.assertEqual({item.direction for item in combined.rows[0].provenance}, {"canonical", "reversed"})
        self.assertEqual(combined.diagnostics, ())

    def test_a_device_move_leaves_both_canonical_forms_alone(self):
        """Corroboration is display evidence, so a rack change must not churn the fingerprint."""
        endpoint_a = _termination("DEVICE-A", "", "PORT-A", "Port")
        panel_in = _termination("PANEL-A", "CARD-A", "FRONT", "Position Front")
        panel_out = _termination("PANEL-A", "CARD-A", "REAR", "Punch-Down")
        endpoint_b = _termination("DEVICE-B", "", "PORT-B", "NIC")
        lines = _endpoint_line(endpoint_a), _endpoint_line(endpoint_b)
        stated = (*lines, (_segment(endpoint_a, "Cable A", panel_in), _segment(panel_out, "Cable B", endpoint_b)))
        moved = (
            *lines,
            (
                _segment(endpoint_a, "Cable A", panel_in, ("9", "RACK-Z", "Site Z")),
                _segment(panel_out, "Cable B", endpoint_b, ("9", "RACK-Z", "Site Z")),
            ),
        )

        original = _interpret(_workbook(path_blocks=(stated,))).rows[0]
        relocated = _interpret(_workbook(path_blocks=(moved,))).rows[0]

        self.assertEqual(relocated.endpoint_summary.from_termination.rack, "RACK-Z")
        self.assertEqual(original.identity, relocated.identity)
        self.assertEqual(original.content_fingerprint, relocated.content_fingerprint)

    def test_a_repatched_path_keeps_its_identity_and_moves_its_fingerprint(self):
        """Endpoint identity is stable across patching, which is why the fingerprint exists."""
        endpoint_a = _termination("DEVICE-A", "", "PORT-A", "Port")
        endpoint_b = _termination("DEVICE-B", "", "PORT-B", "NIC")
        first_in = _termination("PANEL-A", "CARD-A", "FRONT", "Position Front")
        first_out = _termination("PANEL-A", "CARD-A", "REAR", "Punch-Down")
        second_in = _termination("PANEL-B", "CARD-B", "FRONT", "Position Front")
        second_out = _termination("PANEL-B", "CARD-B", "REAR", "Punch-Down")
        lines = _endpoint_line(endpoint_a), _endpoint_line(endpoint_b)
        before = (*lines, (_segment(endpoint_a, "Cable A", first_in), _segment(first_out, "Cable B", endpoint_b)))
        after = (*lines, (_segment(endpoint_a, "Cable A", second_in), _segment(second_out, "Cable B", endpoint_b)))

        original = _interpret(_workbook(path_blocks=(before,))).rows[0]
        repatched = _interpret(_workbook(path_blocks=(after,))).rows[0]

        self.assertEqual(original.identity, repatched.identity)
        self.assertNotEqual(original.content_fingerprint, repatched.content_fingerprint)

    def test_json_identity_does_not_collide_on_separator_characters(self):
        """JSON preserves field boundaries that a separator-joined identity would lose."""
        first_a = _termination("A|B", "C", "D", "Port")
        first_b = _termination("Z|Y", "X", "W", "NIC")
        second_a = _termination("A", "B|C", "D", "Port")
        second_b = _termination("Z", "Y|X", "W", "NIC")
        blocks = (
            (_endpoint_line(first_a), _endpoint_line(first_b), (_segment(first_a, "Cable", first_b),)),
            (_endpoint_line(second_a), _endpoint_line(second_b), (_segment(second_a, "Cable", second_b),)),
        )

        batch = _interpret(_workbook(path_blocks=blocks))

        identities = {trace.identity for trace in batch.rows}
        self.assertEqual(len(identities), 2)
        self.assertTrue(all(isinstance(json.loads(identity), list) for identity in identities))


class TraceWorkbookCorroborationTest(SimpleTestCase):
    """Corroboration reaches a termination from a path row and from a Trace List visit."""

    def _pair(self):
        """Return the two endpoints and the From/To lines both sheets share."""
        endpoint_a = _termination("DEVICE-A", "", "PORT-A", "Port")
        endpoint_b = _termination("DEVICE-B", "", "PORT-B", "NIC")
        return endpoint_a, endpoint_b, (_endpoint_line(endpoint_a), _endpoint_line(endpoint_b))

    def test_a_trace_list_visit_fills_what_the_path_row_left_empty(self):
        """The path row is read first, so its empty cells must not hide what the list states."""
        endpoint_a, endpoint_b, lines = self._pair()
        path_block = (*lines, (_segment(endpoint_a, "Cable A", endpoint_b),))
        stated = ("Site Q", "RACK-Q", "7", "DEVICE-A", "", "PORT-A", "Port", "Ignored")
        list_block = (*lines, (stated, _visit(endpoint_b)))

        batch = _interpret(_workbook(path_blocks=(path_block,), list_blocks=(list_block,), include_list=True))
        stated_end = batch.rows[0].endpoint_summary.from_termination

        self.assertEqual(stated_end.location, "Site Q")
        self.assertEqual(stated_end.rack, "RACK-Q")
        self.assertEqual(stated_end.u_position, "7")

    def test_a_path_row_value_survives_a_later_empty_visit(self):
        """The first non-empty value wins, so a blank list row cannot clear a stated one."""
        endpoint_a, endpoint_b, lines = self._pair()
        path_block = (*lines, (_segment(endpoint_a, "Cable A", endpoint_b, ("9", "RACK-Z", "Site Z")),))
        list_block = (*lines, (_visit(endpoint_a), _visit(endpoint_b)))

        batch = _interpret(_workbook(path_blocks=(path_block,), list_blocks=(list_block,), include_list=True))
        stated_end = batch.rows[0].endpoint_summary.from_termination

        self.assertEqual(stated_end.rack, "RACK-Z")
        self.assertEqual(stated_end.location, "Site Z")
        self.assertEqual(stated_end.u_position, "9")


class TraceWorkbookHandleReleaseTest(SimpleTestCase):
    """openpyxl holds the archive of the BytesIO until close()."""

    def test_the_workbook_is_closed_when_block_extraction_raises(self):
        """Without a finally the archive stays open for every unreadable sheet."""
        endpoint_a = _termination("DEVICE-A", "", "PORT-A", "Port")
        endpoint_b = _termination("DEVICE-B", "", "PORT-B", "NIC")
        block = (_endpoint_line(endpoint_a), _endpoint_line(endpoint_b), (_segment(endpoint_a, "Cable A", endpoint_b),))
        content = _workbook(path_blocks=(block,))
        closed = []
        real_load = openpyxl.load_workbook

        def recording_load(*args, **kwargs):
            """Return the real workbook, with its close recorded."""
            book = real_load(*args, **kwargs)
            original_close = book.close
            book.close = lambda: (closed.append(True), original_close())[1]
            return book

        with (
            mock.patch.object(trace_workbook.openpyxl, "load_workbook", recording_load),
            mock.patch.object(trace_workbook, "_extract_blocks", side_effect=RuntimeError("unreadable sheet")),
            self.assertRaises(RuntimeError),
        ):
            _interpret(content)

        self.assertTrue(closed, "interpret must close the workbook when extraction raises")


class TraceWorkbookTaxonomyTest(SimpleTestCase):
    """Small workbooks cover source-only validation conditions absent from the corpus."""

    def test_no_recognized_sheet_is_a_batch_error(self):
        """An unrelated workbook returns its diagnostic code without creating a unit."""
        book = openpyxl.Workbook()
        sheet = book.active
        if isinstance(sheet, Worksheet):
            sheet.title = "Other"
            sheet.append(("not", "a trace"))
        buffer = BytesIO()
        book.save(buffer)

        batch = _interpret(buffer.getvalue())

        self.assertEqual(batch.rows, ())
        self.assertEqual(_codes(batch), ["trace.no_recognized_sheet"])

    def test_unreadable_bytes_raise_the_adapter_error(self):
        """A broken archive keeps the same unreadable-source contract as the flat adapter."""
        with self.assertRaises(SourceUnreadable):
            _interpret(b"not a workbook")

    def test_an_incomplete_trace_does_not_stop_a_valid_trace(self):
        """A missing header invalidates one trace while the next block remains usable."""
        invalid_a = _termination("DEVICE-A", "", "PORT-A", "Port")
        invalid_b = _termination("DEVICE-B", "", "PORT-B", "NIC")
        valid_a = _termination("DEVICE-C", "", "PORT-C", "Port")
        valid_b = _termination("DEVICE-D", "", "PORT-D", "NIC")
        book = openpyxl.Workbook()
        sheet = book.active
        if not isinstance(sheet, Worksheet):
            sheet = book.create_sheet()
        sheet.title = "Trace From To"
        sheet.append(("Executed", "2026-08-31 12:00:00"))
        sheet.append(())
        sheet.append(("From", _endpoint_line(invalid_a)))
        sheet.append(("To", _endpoint_line(invalid_b)))
        sheet.append(("Wrong", "Header"))
        sheet.append(("From", _endpoint_line(valid_a)))
        sheet.append(("To", _endpoint_line(valid_b)))
        sheet.append(PATH_HEADER)
        sheet.append(_segment(valid_a, "Cable", valid_b))
        buffer = BytesIO()
        book.save(buffer)

        batch = _interpret(buffer.getvalue())

        self.assertEqual(len(batch.rows), 2)
        self.assertEqual(sum(trace.valid for trace in batch.rows), 1)
        self.assertEqual(_codes(batch), ["trace.incomplete_block"])

    def test_trace_list_mismatch_invalidates_only_its_trace(self):
        """A wrong ordered device visit emits the corroboration diagnostic."""
        endpoint_a = _termination("DEVICE-A", "", "PORT-A", "Port")
        panel_in = _termination("PANEL-A", "CARD-A", "FRONT", "Position Front")
        panel_out = _termination("PANEL-A", "CARD-A", "REAR", "Punch-Down")
        endpoint_b = _termination("DEVICE-B", "", "PORT-B", "NIC")
        lines = _endpoint_line(endpoint_a), _endpoint_line(endpoint_b)
        path_block = (*lines, (_segment(endpoint_a, "Cable A", panel_in), _segment(panel_out, "Cable B", endpoint_b)))
        list_block = (*lines, (_visit(endpoint_a), _visit(_termination("WRONG", "", "P", "Port")), _visit(endpoint_b)))

        batch = _interpret(_workbook(path_blocks=(path_block,), list_blocks=(list_block,), include_list=True))

        self.assertFalse(batch.rows[0].valid)
        self.assertEqual(_codes(batch), ["trace.corroboration_mismatch"])

    def test_differing_duplicate_evidence_becomes_one_invalid_trace(self):
        """One identity with two CableClass claims emits a duplicate conflict."""
        endpoint_a = _termination("DEVICE-A", "", "PORT-A", "Port")
        endpoint_b = _termination("DEVICE-B", "", "PORT-B", "NIC")
        lines = _endpoint_line(endpoint_a), _endpoint_line(endpoint_b)
        blocks = (
            (*lines, (_segment(endpoint_a, "Cable A", endpoint_b),)),
            (*lines, (_segment(endpoint_a, "Cable B", endpoint_b),)),
        )

        batch = _interpret(_workbook(path_blocks=blocks))

        self.assertEqual(len(batch.rows), 1)
        self.assertFalse(batch.rows[0].valid)
        self.assertEqual(_codes(batch), ["trace.duplicate_conflict"])
        self.assertIn("block 1", batch.diagnostics[0].message)
        self.assertIn("block 2", batch.diagnostics[0].message)

    def test_unknown_port_class_is_detected_without_target_state(self):
        """A class outside the fixed PortClass vocabulary is a source error."""
        endpoint_a = _termination("DEVICE-A", "", "PORT-A", "Unknown Class")
        endpoint_b = _termination("DEVICE-B", "", "PORT-B", "NIC")
        block = (
            _endpoint_line(endpoint_a),
            _endpoint_line(endpoint_b),
            (_segment(endpoint_a, "Cable", endpoint_b),),
        )

        batch = _interpret(_workbook(path_blocks=(block,)))

        self.assertFalse(batch.rows[0].valid)
        self.assertEqual(_codes(batch), ["trace.unknown_port_class"])

    def test_a_shared_termination_invalidates_every_involved_trace(self):
        """Two traces landing different cables on one front port is an occupancy conflict alone."""
        blocks = _two_traces_over_one_pair("Shared", "Shared", shared_entry=True)

        batch = _interpret(_workbook(path_blocks=blocks))

        self.assertEqual(len(batch.rows), 2)
        self.assertTrue(all(not trace.valid for trace in batch.rows))
        self.assertEqual(_codes(batch), ["trace.cross_trace_conflict", "trace.cross_trace_conflict"])
        self.assertIn("claimed by another Source Trace", batch.diagnostics[0].message)
        self.assertNotIn("CableClass", batch.diagnostics[0].message)

    def test_one_segment_claimed_with_two_cable_classes_invalidates_both_traces(self):
        """A shared segment stated with two labels cannot become one Cable."""
        blocks = _two_traces_over_one_pair("Shared A", "Shared B")

        batch = _interpret(_workbook(path_blocks=blocks))

        self.assertEqual(len(batch.rows), 2)
        self.assertTrue(all(not trace.valid for trace in batch.rows))
        self.assertEqual(_codes(batch), ["trace.cross_trace_conflict", "trace.cross_trace_conflict"])
        self.assertIn("conflicting CableClass labels", batch.diagnostics[0].message)

    def test_a_cable_class_that_differs_only_in_case_is_a_disagreement(self):
        """A CableClass label keys its own mapping row, so it is not normalized like an identity."""
        batch = _interpret(_workbook(path_blocks=_two_traces_over_one_pair("Shared", "shared")))

        self.assertTrue(all(not trace.valid for trace in batch.rows))
        self.assertIn("conflicting CableClass labels", batch.diagnostics[0].message)

    def test_two_unreadable_rows_each_report_their_own_location(self):
        """Issue #84 asks the diagnostics to name the malformed rows, not just the first."""
        endpoint_a = _termination("DEVICE-A", "", "PORT-A", "Port")
        endpoint_b = _termination("DEVICE-B", "", "PORT-B", "NIC")
        nameless = _termination("", "", "PORT-X", "Port")
        block = (
            _endpoint_line(endpoint_a),
            _endpoint_line(endpoint_b),
            (
                _segment(nameless, "Cable", endpoint_b),
                _segment(nameless, "Cable", endpoint_b),
                _segment(endpoint_a, "Cable", endpoint_b),
            ),
        )

        batch = _interpret(_workbook(path_blocks=(block,)))

        self.assertEqual(_codes(batch), ["trace.incomplete_block", "trace.incomplete_block"])
        self.assertEqual([diagnostic.row_number for diagnostic in batch.diagnostics], [6, 7])

    def test_an_unpaired_block_without_a_to_line_is_incomplete(self):
        """A From line with nothing under it is the section 5.6 incomplete block, not silence."""
        book = openpyxl.Workbook()
        sheet = book.active
        if not isinstance(sheet, Worksheet):
            sheet = book.create_sheet()
        sheet.title = "Trace List"
        sheet.append(("Executed", "2026-08-31 12:00:00"))
        sheet.append(())
        sheet.append(("From", _endpoint_line(_termination("DEVICE-A", "", "PORT-A", "Port"))))
        buffer = BytesIO()
        book.save(buffer)

        batch = _interpret(buffer.getvalue())

        self.assertEqual(batch.rows, ())
        self.assertIn("trace.incomplete_block", _codes(batch))

    def test_fallback_occurrences_with_different_endpoint_evidence_conflict(self):
        """An Endpoint Summary fallback states no segment, so its endpoints carry all its evidence."""
        endpoint_b = _termination("DEVICE-B", "", "PORT-B", "NIC")
        blocks = tuple(
            (
                _endpoint_line(endpoint_a),
                _endpoint_line(endpoint_b),
                (_visit(endpoint_a), _visit(endpoint_b)),
            )
            for endpoint_a in (
                _termination("DEVICE-A", "", "PORT-A", "Port"),
                _termination("DEVICE-A", "", "PORT-A", "Switch Port"),
            )
        )

        batch = _interpret(_workbook(list_blocks=blocks, include_path=False, include_list=True))

        self.assertEqual(len(batch.rows), 1)
        self.assertEqual(_codes(batch), ["trace.duplicate_conflict"])

    def test_a_same_rear_port_continuation_is_legal_evidence(self):
        """A trunk and a patch recorded on one rear port stay valid, unlike an interface hop."""
        endpoint_a = _termination("DEVICE-A", "", "PORT-A", "Port")
        panel_front = _termination("PANEL-A", "CARD-A", "FRONT", "Position Front")
        panel_rear = _termination("PANEL-A", "CARD-A", "REAR", "Punch-Down")
        shared_rear = _termination("PANEL-B", "CARD-B", "REAR", "Punch-Down")
        endpoint_b = _termination("DEVICE-B", "", "PORT-B", "NIC")
        block = (
            _endpoint_line(endpoint_a),
            _endpoint_line(endpoint_b),
            (
                _segment(endpoint_a, "Cable A", panel_front),
                _segment(panel_rear, "Trunk", shared_rear),
                _segment(shared_rear, "Patch", endpoint_b),
            ),
        )

        batch = _interpret(_workbook(path_blocks=(block,)))

        self.assertTrue(batch.rows[0].valid)
        self.assertEqual(batch.diagnostics, ())
        self.assertIn(
            ("PANEL-B", "REAR", "REAR"),
            [(claim.device, claim.entry_port, claim.exit_port) for claim in batch.rows[0].pass_through_claims],
        )

    def test_an_identical_shared_segment_is_allowed(self):
        """Two traces may contribute the same segment with the same CableClass."""
        endpoint_a = _termination("DEVICE-A", "", "PORT-A", "Port")
        endpoint_b = _termination("DEVICE-B", "", "PORT-B", "NIC")
        endpoint_c = _termination("DEVICE-C", "", "PORT-C", "Port")
        endpoint_d = _termination("DEVICE-D", "", "PORT-D", "NIC")
        panel_p_first = _termination("PANEL-P", "CARD-P", "FRONT-1", "Position Front")
        panel_p_second = _termination("PANEL-P", "CARD-P", "FRONT-2", "Position Front")
        panel_p_shared = _termination("PANEL-P", "CARD-P", "REAR", "Punch-Down")
        panel_q_shared = _termination("PANEL-Q", "CARD-Q", "REAR", "Punch-Down")
        panel_q_first = _termination("PANEL-Q", "CARD-Q", "FRONT-1", "Position Front")
        panel_q_second = _termination("PANEL-Q", "CARD-Q", "FRONT-2", "Position Front")
        first = (
            _endpoint_line(endpoint_a),
            _endpoint_line(endpoint_b),
            (
                _segment(endpoint_a, "Cable A", panel_p_first),
                _segment(panel_p_shared, "Shared", panel_q_shared),
                _segment(panel_q_first, "Cable B", endpoint_b),
            ),
        )
        second = (
            _endpoint_line(endpoint_c),
            _endpoint_line(endpoint_d),
            (
                _segment(endpoint_c, "Cable C", panel_p_second),
                _segment(panel_p_shared, "Shared", panel_q_shared),
                _segment(panel_q_second, "Cable D", endpoint_d),
            ),
        )

        batch = _interpret(_workbook(path_blocks=(first, second)))

        self.assertEqual(len(batch.rows), 2)
        self.assertTrue(all(trace.valid for trace in batch.rows))
        self.assertNotIn("trace.cross_trace_conflict", _codes(batch))

    def test_an_unpaired_trace_list_block_becomes_endpoint_summary_evidence(self):
        """Trace List data can produce a valid Source Trace without Segment Evidence."""
        endpoint_a = _termination("DEVICE-A", "", "PORT-A", "Port")
        endpoint_b = _termination("DEVICE-B", "", "PORT-B", "NIC")
        block = (
            _endpoint_line(endpoint_a),
            _endpoint_line(endpoint_b),
            (_visit(endpoint_a), _visit(endpoint_b)),
        )

        batch = _interpret(_workbook(list_blocks=(block,), include_path=False, include_list=True))

        self.assertEqual(len(batch.rows), 1)
        self.assertTrue(batch.rows[0].valid)
        self.assertEqual(batch.rows[0].segments, ())
        self.assertEqual(batch.rows[0].pass_through_claims, ())
        self.assertEqual(batch.rows[0].provenance[0].sheet, "Trace List")


class TraceWorkbookExportArtifactTest(SimpleTestCase):
    """The exporter overruns a block's separator row when the block fills it."""

    endpoint_a = _termination("DEVICE-A", "", "PORT-A", "Port")
    panel_front = _termination("PANEL-A", "CARD-A", "FRONT", "Position Front")
    panel_rear = _termination("PANEL-A", "CARD-A", "REAR", "Punch-Down")
    endpoint_b = _termination("DEVICE-B", "", "PORT-B", "NIC")
    endpoint_c = _termination("DEVICE-C", "", "PORT-C", "Port")
    endpoint_d = _termination("DEVICE-D", "", "PORT-D", "NIC")

    def _sheet(self, title, header):
        """Return an empty trace sheet with its export timestamp row."""
        book = openpyxl.Workbook()
        sheet = book.active
        if not isinstance(sheet, Worksheet):
            sheet = book.create_sheet()
        sheet.title = title
        sheet.append(("Executed", "2026-08-31 12:00:00"))
        sheet.append(())
        return book, sheet

    @staticmethod
    def _overrun(line_marker, line_text, row):
        """Return the next block's line with the previous block's trailing row still under it."""
        return (line_marker, line_text, *row[2:])

    def test_a_trace_list_visit_under_the_next_block_still_corroborates(self):
        """The fiber corpus depends on this recovery, so an isolated case must prove it."""
        path_block = (
            _endpoint_line(self.endpoint_a),
            _endpoint_line(self.endpoint_b),
            (
                _segment(self.endpoint_a, "Cable A", self.panel_front),
                _segment(self.panel_rear, "Cable B", self.endpoint_b),
            ),
        )
        book, sheet = self._sheet("Trace List", LIST_HEADER)
        sheet.append(("From", _endpoint_line(self.endpoint_a)))
        sheet.append(("To", _endpoint_line(self.endpoint_b)))
        sheet.append(LIST_HEADER)
        sheet.append(_visit(self.endpoint_a))
        sheet.append(_visit(self.panel_front))
        sheet.append(self._overrun("From", _endpoint_line(self.endpoint_c), _visit(self.endpoint_b)))
        sheet.append(("To", _endpoint_line(self.endpoint_d)))
        sheet.append(LIST_HEADER)
        path = book.create_sheet("Trace From To")
        path.append(("Executed", "2026-08-31 12:00:00"))
        path.append(())
        for from_line, to_line, rows in (path_block,):
            path.append(("From", from_line))
            path.append(("To", to_line))
            path.append(PATH_HEADER)
            for row in rows:
                path.append(row)
        buffer = BytesIO()
        book.save(buffer)

        batch = _interpret(buffer.getvalue())

        self.assertEqual(len(batch.rows), 1)
        self.assertTrue(batch.rows[0].valid)
        self.assertEqual([visit.device for visit in batch.rows[0].corroboration], ["DEVICE-A", "PANEL-A", "DEVICE-B"])
        self.assertEqual(batch.diagnostics, ())

    def test_a_segment_row_under_the_next_block_is_reported_not_dropped(self):
        """The overrun destroys the row's Port and PortClass cells, so it must fail loudly."""
        book, sheet = self._sheet("Trace From To", PATH_HEADER)
        sheet.append(("From", _endpoint_line(self.endpoint_a)))
        sheet.append(("To", _endpoint_line(self.endpoint_b)))
        sheet.append(PATH_HEADER)
        sheet.append(_segment(self.endpoint_a, "Cable", self.endpoint_b))
        sheet.append(
            self._overrun("From", _endpoint_line(self.endpoint_c), _segment(self.endpoint_a, "Cable", self.endpoint_b))
        )
        sheet.append(("To", _endpoint_line(self.endpoint_d)))
        sheet.append(PATH_HEADER)
        sheet.append(_segment(self.endpoint_c, "Cable", self.endpoint_d))
        buffer = BytesIO()
        book.save(buffer)

        batch = _interpret(buffer.getvalue())

        overrun_trace, following_trace = batch.rows
        self.assertFalse(overrun_trace.valid)
        self.assertEqual([error.code for error in overrun_trace.errors], ["trace.incomplete_block"])
        self.assertTrue(following_trace.valid, "the next block still reads normally")


class TraceWorkbookCollapseAndPairingTest(SimpleTestCase):
    """Two statements of one identity must pair, corroborate, and report independently."""

    endpoint_a = _termination("DEVICE-A", "", "PORT-A", "Port")
    panel_front = _termination("PANEL-A", "CARD-A", "FRONT", "Position Front")
    panel_rear = _termination("PANEL-A", "CARD-A", "REAR", "Punch-Down")
    endpoint_b = _termination("DEVICE-B", "", "PORT-B", "NIC")

    def _path_block(self, from_line, to_line):
        """Return one two-segment path block under the supplied endpoint lines."""
        return (
            from_line,
            to_line,
            (
                _segment(self.endpoint_a, "Cable A", self.panel_front),
                _segment(self.panel_rear, "Cable B", self.endpoint_b),
            ),
        )

    def test_a_later_occurrence_keeps_the_finding_its_own_trace_list_states(self):
        """The fingerprint excludes Trace List data, so equal fingerprints are not equal findings."""
        lines = _endpoint_line(self.endpoint_a), _endpoint_line(self.endpoint_b)
        corroborating = (*lines, (_visit(self.endpoint_a), _visit(self.panel_front), _visit(self.endpoint_b)))
        contradicting = (*lines, (_visit(self.endpoint_a), _visit(_termination("WRONG", "", "P", "Port"))))

        batch = _interpret(
            _workbook(
                path_blocks=(self._path_block(*lines), self._path_block(*lines)),
                list_blocks=(corroborating, contradicting),
                include_list=True,
            )
        )

        self.assertEqual(len(batch.rows), 1)
        self.assertEqual(_codes(batch), ["trace.corroboration_mismatch"])
        self.assertFalse(batch.rows[0].valid)

    def test_surrounding_space_on_one_sheet_still_pairs_the_block(self):
        """An untrimmed export line must not turn its own Trace List block into a rival trace."""
        from_line, to_line = _endpoint_line(self.endpoint_a), _endpoint_line(self.endpoint_b)
        list_block = (
            f" {from_line}",
            f"{to_line} ",
            (_visit(self.endpoint_a), _visit(self.panel_front), _visit(self.endpoint_b)),
        )

        batch = _interpret(
            _workbook(
                path_blocks=(self._path_block(from_line, to_line),),
                list_blocks=(list_block,),
                include_list=True,
            )
        )

        self.assertEqual(len(batch.rows), 1)
        self.assertTrue(batch.rows[0].valid)
        self.assertEqual(batch.diagnostics, ())
        self.assertEqual(len(batch.rows[0].corroboration), 3)

    def test_the_next_block_header_is_not_absorbed_as_a_carried_row(self):
        """Only a From or To line overwrites a trailing row, so a header row is never carried."""
        book = openpyxl.Workbook()
        sheet = book.active
        if not isinstance(sheet, Worksheet):
            sheet = book.create_sheet()
        sheet.title = "Trace From To"
        sheet.append(("Executed", "2026-08-31 12:00:00"))
        sheet.append(())
        sheet.append(("From", _endpoint_line(self.endpoint_a)))
        sheet.append(("To", _endpoint_line(self.endpoint_b)))
        sheet.append(PATH_HEADER)
        sheet.append(_segment(self.endpoint_a, "Cable A", self.panel_front))
        sheet.append(_segment(self.panel_rear, "Cable B", self.endpoint_b))
        sheet.append(("From", _endpoint_line(_termination("DEVICE-C", "", "PORT-C", "Port"))))
        sheet.append(PATH_HEADER)
        buffer = BytesIO()
        book.save(buffer)

        batch = _interpret(buffer.getvalue())

        complete = next(trace for trace in batch.rows if trace.segments)
        self.assertTrue(complete.valid, [error.code for error in complete.errors])
        self.assertEqual(len(complete.segments), 2)
