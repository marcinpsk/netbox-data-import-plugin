# SPDX-License-Identifier: Apache-2.0
# Copyright (C) 2025 Marcin Zieba <marcinpsk@gmail.com>
"""Unit tests for the import engine: parse_file and run_import (dry_run mode)."""

import os
from io import BytesIO

from django.test import TestCase

from netbox_data_import.engine import (
    ImportContext,
    ImportResult,
    ParseError,
    RowResult,
    _ensure_device_type,
    _ip_already_assigned,
    _normalize_for_compare,
    _preview_device_row,
    parse_file,
    run_import,
)
from netbox_data_import.models import ClassRoleMapping, ColumnMapping, ImportProfile


FIXTURE_PATH = os.path.join(os.path.dirname(__file__), "fixtures", "sample_cans.xlsx")


def _make_profile(name="Test") -> ImportProfile:
    """Create a fully configured ImportProfile matching the sample fixture."""
    profile = ImportProfile.objects.create(
        name=name,
        adapter_config={
            "sheet_name": "Data",
            "source_id_column": "Id",
            "custom_field_name": "",
            "update_existing": True,
            "create_missing_device_types": True,
        },
    )
    # Standard CANS column mappings
    field_map = {
        "Id": "source_id",
        "Rack": "rack_name",
        "Name": "device_name",
        "Class": "device_class",
        "Side": "face",
        "Airflow": "airflow",
        "UPosition": "u_position",
        "Status": "status",
        "Make": "make",
        "Model": "model",
        "UHeight": "u_height",
        "Serial Number": "serial",
        "Asset Tag": "asset_tag",
    }
    for src, tgt in field_map.items():
        ColumnMapping.objects.create(profile=profile, source_column=src, target_field=tgt)

    # Cabinet class → rack
    ClassRoleMapping.objects.create(
        profile=profile,
        source_class="Cabinet",
        creates_rack=True,
    )
    # Server class → device role
    ClassRoleMapping.objects.create(
        profile=profile,
        source_class="Server",
        creates_rack=False,
        role_slug="server",
    )
    # Switch class → device role
    ClassRoleMapping.objects.create(
        profile=profile,
        source_class="Switch",
        creates_rack=False,
        role_slug="network-switch",
    )
    return profile


class ParseFileTest(TestCase):
    """Tests for engine.parse_file."""

    def test_parse_sample_fixture(self):
        """parse_file returns one row-dict per non-empty data row."""
        profile = _make_profile("ParseTest")
        with open(FIXTURE_PATH, "rb") as f:
            rows = parse_file(f, profile)

        # The fixture has 3 data rows (1 rack + 2 devices)
        self.assertEqual(len(rows), 3)

    def test_row_keys_match_target_fields(self):
        """Each row-dict is keyed by target_field names, not source column names."""
        profile = _make_profile("KeyTest")
        with open(FIXTURE_PATH, "rb") as f:
            rows = parse_file(f, profile)

        for row in rows:
            # No raw source column names should appear (only target fields + _row_number)
            self.assertIn("_row_number", row)
            self.assertNotIn("Serial Number", row)  # source name must be replaced
            self.assertNotIn("UPosition", row)

    def test_rack_row_has_rack_class(self):
        """The Cabinet row maps device_class to 'Cabinet'."""
        profile = _make_profile("RackRow")
        with open(FIXTURE_PATH, "rb") as f:
            rows = parse_file(f, profile)

        rack_rows = [r for r in rows if r.get("device_class") == "Cabinet"]
        self.assertEqual(len(rack_rows), 1)
        self.assertEqual(rack_rows[0]["rack_name"], "Rack-01")

    def test_missing_sheet_raises_parse_error(self):
        """ParseError is raised when the sheet name doesn't exist."""
        profile = _make_profile("BadSheet")
        profile.adapter_config["sheet_name"] = "NonExistent"
        with open(FIXTURE_PATH, "rb") as f:
            with self.assertRaises(ParseError):
                parse_file(f, profile)

    def test_invalid_file_raises_parse_error(self):
        """ParseError is raised for non-Excel binary data."""
        profile = _make_profile("BadFile")
        garbage = BytesIO(b"this is not an excel file")
        with self.assertRaises(ParseError):
            parse_file(garbage, profile)


class MultiColumnMergeTest(TestCase):
    """Tests for multi-source column merging in parse_file."""

    def _make_merge_profile(self) -> ImportProfile:
        """Profile with 'Serial Number' and 'Service Tag' both mapping to 'serial'."""
        profile = ImportProfile.objects.create(
            name="MergeTest",
            adapter_config={"sheet_name": "Data", "update_existing": True, "create_missing_device_types": True},
        )
        for src, tgt in [
            ("Id", "source_id"),
            ("Rack", "rack_name"),
            ("Name", "device_name"),
            ("Class", "device_class"),
            ("Make", "make"),
            ("Model", "model"),
            ("Serial Number", "serial"),
            ("Service Tag", "serial"),
        ]:
            ColumnMapping.objects.create(profile=profile, source_column=src, target_field=tgt)
        return profile

    def _make_single_row_workbook(
        self,
        serial_number: str | None,
        service_tag: str | None,
    ) -> BytesIO:
        """Build an in-memory Excel file with one data row."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["Id", "Rack", "Name", "Class", "Make", "Model", "Serial Number", "Service Tag"])
        ws.append(["100", "Rack-01", "Dev-01", "Server", "Cisco", "C9300", serial_number, service_tag])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_single_source_no_conflict(self):
        """When only one source column has a value, it is used with no conflict."""
        profile = self._make_merge_profile()
        rows = parse_file(self._make_single_row_workbook("SN-001", None), profile)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["serial"], "SN-001")
        self.assertNotIn("_conflicts", rows[0])

    def test_both_sources_same_value_no_conflict(self):
        """When both sources have identical values, the value is used with no conflict."""
        profile = self._make_merge_profile()
        rows = parse_file(self._make_single_row_workbook("SAME-42", "SAME-42"), profile)
        self.assertEqual(rows[0]["serial"], "SAME-42")
        self.assertNotIn("_conflicts", rows[0])

    def test_both_sources_different_values_conflict(self):
        """When both sources have different non-empty values, _conflicts is populated."""
        profile = self._make_merge_profile()
        rows = parse_file(self._make_single_row_workbook("ABC-111", "XYZ-999"), profile)
        self.assertIsNone(rows[0].get("serial"))
        self.assertIn("_conflicts", rows[0])
        conflict = rows[0]["_conflicts"]["serial"]
        self.assertIn("Serial Number", conflict)
        self.assertIn("Service Tag", conflict)
        self.assertEqual(conflict["Serial Number"], "ABC-111")
        self.assertEqual(conflict["Service Tag"], "XYZ-999")

    def test_conflict_cleared_by_saved_resolution(self):
        """A saved SourceResolution for the target field clears the conflict when rows are derived."""
        from netbox_data_import.engine import derive_effective_rows
        from netbox_data_import.models import SourceResolution

        profile = self._make_merge_profile()
        SourceResolution.objects.create(
            profile=profile,
            source_id="100",
            source_column="_merge_serial",
            original_value="",
            resolved_fields={"serial": "ABC-111"},
        )

        parsed = parse_file(self._make_single_row_workbook("ABC-111", "XYZ-999"), profile)
        self.assertIsNone(parsed[0].get("serial"), "parsing stays pristine")

        rows = derive_effective_rows(parsed, profile)
        self.assertEqual(rows[0]["serial"], "ABC-111")
        self.assertFalse(rows[0].get("_conflicts", {}).get("serial"))


class DerivationOrderTest(TestCase):
    """Two resolutions for one source row must apply in the same order on every derivation."""

    def _profile_with_two_resolutions_for_one_row(self):
        """Save the later source_column first, so insertion order is not the order to apply."""
        from netbox_data_import.models import SourceResolution

        profile = ImportProfile.objects.create(
            name="Derivation Order",
            adapter_config={"sheet_name": "Data", "source_id_column": "Id"},
        )
        ColumnMapping.objects.create(profile=profile, source_column="Name", target_field="device_name")
        # (profile, source_id, source_column) is unique, so one source row can own several.
        for column, decision in (("zz_column", "from-zz"), ("aa_column", "from-aa")):
            SourceResolution.objects.create(
                profile=profile,
                source_id="ORDER-1",
                source_column=column,
                original_value="pristine",
                resolved_fields={"device_name": decision},
            )
        return profile

    def test_the_database_is_asked_for_a_total_order(self):
        """The job compares two derivations, so an undefined order reads as a changed policy."""
        from django.db import connection

        from netbox_data_import.engine import derive_effective_rows
        from netbox_data_import.models import SourceResolution

        profile = self._profile_with_two_resolutions_for_one_row()
        reads = []

        def record_the_resolution_read(execute, sql, params, many, context):
            if SourceResolution._meta.db_table in sql and "ORDER BY" in sql:
                reads.append(sql)
            return execute(sql, params, many, context)

        with connection.execute_wrapper(record_the_resolution_read):
            derive_effective_rows([{"source_id": "ORDER-1", "device_name": "pristine"}], profile)

        self.assertTrue(reads, "the resolutions were never read with an ORDER BY")
        for sql in reads:
            # Meta.ordering stops at source_id, and SQL leaves ties there in no defined order.
            self.assertIn("source_column", sql.rsplit("ORDER BY", 1)[1], sql)

    def test_the_resolution_that_wins_does_not_depend_on_the_database(self):
        """Applying the same rows in the other order would name the other decision."""
        from netbox_data_import.engine import derive_effective_rows

        profile = self._profile_with_two_resolutions_for_one_row()
        rows = [{"source_id": "ORDER-1", "device_name": "pristine"}]

        derived = derive_effective_rows(rows, profile)

        # source_column completes the order, so the last writer for the field is always this one.
        self.assertEqual(derived[0]["device_name"], "from-zz")


class ApplyColumnMappingsTest(TestCase):
    """Tests for apply_column_mappings — re-applying mappings to already-parsed rows."""

    def _make_profile_with_mapping(self, source: str, target: str) -> ImportProfile:
        profile = ImportProfile.objects.create(
            name="ApplyMapTest",
            adapter_config={"sheet_name": "Data", "update_existing": True, "create_missing_device_types": True},
        )
        ColumnMapping.objects.create(profile=profile, source_column=source, target_field=target)
        return profile

    def test_basic_mapping_applies_value(self):
        """A raw source column is replaced with the target field key."""
        from netbox_data_import.engine import apply_column_mappings

        profile = self._make_profile_with_mapping("Asset Tag", "asset_tag")
        rows = [{"Asset Tag": "TAG-001", "device_name": "Dev-01"}]
        result = apply_column_mappings(rows, profile)
        self.assertEqual(result[0]["asset_tag"], "TAG-001")
        self.assertNotIn("Asset Tag", result[0])

    def test_empty_source_value_skipped(self):
        """A source column with an empty value is skipped, leaving target field absent."""
        from netbox_data_import.engine import apply_column_mappings

        profile = self._make_profile_with_mapping("Asset Tag", "asset_tag")
        rows = [{"Asset Tag": "", "device_name": "Dev-01"}]
        result = apply_column_mappings(rows, profile)
        self.assertNotIn("asset_tag", result[0])

    def test_multi_source_same_value_no_conflict(self):
        """Two source columns with identical values merge without a conflict."""
        from netbox_data_import.engine import apply_column_mappings

        profile = ImportProfile.objects.create(name="MultiSame", adapter_config={"sheet_name": "Data"})
        ColumnMapping.objects.create(profile=profile, source_column="SN1", target_field="serial")
        ColumnMapping.objects.create(profile=profile, source_column="SN2", target_field="serial")
        rows = [{"SN1": "ABC-100", "SN2": "ABC-100"}]
        result = apply_column_mappings(rows, profile)
        self.assertEqual(result[0]["serial"], "ABC-100")
        self.assertNotIn("_conflicts", result[0])

    def test_multi_source_different_values_conflict(self):
        """Two source columns with different values produce a conflict entry."""
        from netbox_data_import.engine import apply_column_mappings

        profile = ImportProfile.objects.create(name="MultiDiff", adapter_config={"sheet_name": "Data"})
        ColumnMapping.objects.create(profile=profile, source_column="SN1", target_field="serial")
        ColumnMapping.objects.create(profile=profile, source_column="SN2", target_field="serial")
        rows = [{"SN1": "ABC-100", "SN2": "XYZ-999"}]
        result = apply_column_mappings(rows, profile)
        self.assertIsNone(result[0].get("serial"))
        self.assertIn("serial", result[0].get("_conflicts", {}))

    def test_existing_target_with_new_conflicting_source_records_conflict(self):
        """When a target field already has a value and a new mapping brings a different value, conflict is recorded."""
        from netbox_data_import.engine import apply_column_mappings

        profile = self._make_profile_with_mapping("Service Tag", "serial")
        rows = [{"Service Tag": "NEW-999", "serial": "OLD-001"}]
        result = apply_column_mappings(rows, profile)
        self.assertIsNone(result[0].get("serial"))
        conflicts = result[0].get("_conflicts", {})
        self.assertIn("serial", conflicts)

    def test_existing_target_with_same_value_no_conflict(self):
        """When a new source column provides the same value as the existing target, no conflict is recorded."""
        from netbox_data_import.engine import apply_column_mappings

        profile = self._make_profile_with_mapping("Service Tag", "serial")
        rows = [{"Service Tag": "SAME-001", "serial": "SAME-001"}]
        result = apply_column_mappings(rows, profile)
        self.assertEqual(result[0].get("serial"), "SAME-001")
        self.assertNotIn("_conflicts", result[0])

    def test_serial_exact_comparison_no_false_merge(self):
        """Serial numbers '0042' and '42' must NOT be merged — they differ as text identifiers."""
        from netbox_data_import.engine import apply_column_mappings

        profile = ImportProfile.objects.create(name="SerialExact", adapter_config={"sheet_name": "Data"})
        ColumnMapping.objects.create(profile=profile, source_column="SN1", target_field="serial")
        ColumnMapping.objects.create(profile=profile, source_column="SN2", target_field="serial")
        rows = [{"SN1": "0042", "SN2": "42"}]
        result = apply_column_mappings(rows, profile)
        self.assertIsNone(result[0].get("serial"))
        self.assertIn("serial", result[0].get("_conflicts", {}))

    def test_u_position_float_int_no_false_conflict(self):
        """u_position '35.0' and '35' should merge without conflict (numeric normalization)."""
        from netbox_data_import.engine import apply_column_mappings

        profile = ImportProfile.objects.create(name="UPosFloat", adapter_config={"sheet_name": "Data"})
        ColumnMapping.objects.create(profile=profile, source_column="Pos1", target_field="u_position")
        ColumnMapping.objects.create(profile=profile, source_column="Pos2", target_field="u_position")
        rows = [{"Pos1": "35.0", "Pos2": "35"}]
        result = apply_column_mappings(rows, profile)
        self.assertNotIn("_conflicts", result[0])

    def test_source_in_extra_columns_is_promoted(self):
        """A source column stored under _extra_columns is promoted to the target field."""
        from netbox_data_import.engine import apply_column_mappings

        profile = self._make_profile_with_mapping("Asset Tag", "asset_tag")
        rows = [{"_extra_columns": {"Asset Tag": "TAG-EXTRA"}, "device_name": "Dev-01"}]
        result = apply_column_mappings(rows, profile)
        self.assertEqual(result[0]["asset_tag"], "TAG-EXTRA")
        self.assertNotIn("Asset Tag", result[0].get("_extra_columns", {}))

    def test_candidate_mapping_collects_values_without_promoting_a_field(self):
        """Candidate mappings retain their source labels for row-level review."""
        from netbox_data_import.engine import apply_column_mappings

        profile = self._make_profile_with_mapping("Contact Email", "candidate:contact")
        rows = [{"_extra_columns": {"Contact Email": "candidate@example.invalid"}}]

        result = apply_column_mappings(rows, profile)

        self.assertEqual(
            result[0]["_candidate_values"]["contact"],
            {"Contact Email": "candidate@example.invalid"},
        )
        self.assertNotIn("Contact Email", result[0].get("_extra_columns", {}))

    def test_returns_rows(self):
        """apply_column_mappings returns the modified rows list."""
        from netbox_data_import.engine import apply_column_mappings

        profile = self._make_profile_with_mapping("Asset Tag", "asset_tag")
        rows = [{"device_name": "X"}]
        result = apply_column_mappings(rows, profile)
        self.assertIsInstance(result, list)


class RunImportDryRunTest(TestCase):
    """Tests for engine.run_import with dry_run=True (no DB writes)."""

    def setUp(self):
        from dcim.models import Site

        self.site = Site.objects.create(name="Test Site", slug="test-site")
        self.profile = _make_profile("DryRun")

    def test_dry_run_returns_import_result(self):
        """run_import returns an ImportResult instance."""
        with open(FIXTURE_PATH, "rb") as f:
            rows = parse_file(f, self.profile)

        result = run_import(rows, self.profile, {"site": self.site}, dry_run=True)
        self.assertIsInstance(result, ImportResult)

    def test_dry_run_has_no_errors(self):
        """The sample fixture produces no error rows in dry-run mode."""
        with open(FIXTURE_PATH, "rb") as f:
            rows = parse_file(f, self.profile)

        result = run_import(rows, self.profile, {"site": self.site}, dry_run=True)
        error_rows = [r for r in result.rows if r.action == "error"]
        self.assertEqual(error_rows, [], msg=f"Unexpected errors: {error_rows}")

    def test_dry_run_identifies_rack_and_devices(self):
        """Dry-run result contains both rack and device rows."""
        with open(FIXTURE_PATH, "rb") as f:
            rows = parse_file(f, self.profile)

        result = run_import(rows, self.profile, {"site": self.site}, dry_run=True)
        types = {r.object_type for r in result.rows}
        self.assertIn("rack", types)
        self.assertIn("device", types)

    def test_dry_run_does_not_write_to_db(self):
        """No Rack or Device rows are created in dry-run mode."""
        from dcim.models import Device, Rack

        with open(FIXTURE_PATH, "rb") as f:
            rows = parse_file(f, self.profile)

        run_import(rows, self.profile, {"site": self.site}, dry_run=True)
        self.assertEqual(Rack.objects.filter(site=self.site).count(), 0)
        self.assertEqual(Device.objects.filter(site=self.site).count(), 0)

    def test_dry_run_counts(self):
        """Result counts reflect what would be created."""
        with open(FIXTURE_PATH, "rb") as f:
            rows = parse_file(f, self.profile)

        result = run_import(rows, self.profile, {"site": self.site}, dry_run=True)
        # There should be at least 1 rack to create
        self.assertGreater(result.counts.get("racks_created", 0), 0)


class RowResultSerializationTest(TestCase):
    """Tests for RowResult and ImportResult serialization helpers."""

    def test_row_result_roundtrip(self):
        """RowResult.to_dict() and RowResult.from_dict() are inverse operations."""
        r = RowResult(
            row_number=5,
            source_id="42",
            name="switch-01",
            action="create",
            object_type="device",
            detail="Would create device",
            netbox_url="",
        )
        d = r.to_dict()
        restored = RowResult.from_dict(d)
        self.assertEqual(restored.name, r.name)
        self.assertEqual(restored.action, r.action)

    def test_import_result_session_roundtrip(self):
        """ImportResult can be serialised to a dict and restored correctly."""
        result = ImportResult()
        result.rows = [
            RowResult(1, "1", "rack-01", "create", "rack", "Would create rack"),
            RowResult(2, "2", "server-01", "create", "device", "Would create device"),
        ]
        result._recompute_counts()

        session_dict = result.to_session_dict()
        restored = ImportResult.from_session_dict(session_dict)

        self.assertEqual(len(restored.rows), 2)
        self.assertEqual(restored.counts.get("racks_created"), 1)
        self.assertEqual(restored.counts.get("devices_created"), 1)


class PreviewDeviceRowTest(TestCase):
    """Unit tests for _preview_device_row internals."""

    def setUp(self):
        from dcim.models import Site

        self.site = Site.objects.create(name="Preview Site", slug="preview-site")
        self.profile = _make_profile("Preview")

    def test_position_without_rack_is_an_error(self):
        """A rack position without a rack is rejected."""
        from dcim.models import Device, DeviceType, Rack

        row = {
            "_row_number": 1,
            "rack_name": "",
            "u_position": 3,
        }
        ctx = ImportContext(
            profile=self.profile, site=self.site, location=None, tenant=None, dry_run=True, result=ImportResult()
        )
        result_row = _preview_device_row(
            row=row,
            ctx=ctx,
            make="TestMake",
            model="TestModel",
            mfg_slug="test-mfg",
            dt_slug="test-dt",
            source_id="1",
            device_name="test-device-01",
            serial="",
            asset_tag="",
            DeviceType=DeviceType,
            Device=Device,
            Rack=Rack,
        )
        self.assertEqual(result_row.action, "error")
        self.assertEqual(result_row.extra_data["identity_conflict"], "rack_required")

    def test_unknown_rack_is_an_error(self):
        """A non-empty rack name must resolve in the active location."""
        from dcim.models import Device, DeviceType, Rack

        row = {
            "_row_number": 2,
            "rack_name": "RACK-99",
            "u_position": 5,
        }
        ctx = ImportContext(
            profile=self.profile, site=self.site, location=None, tenant=None, dry_run=True, result=ImportResult()
        )
        result_row = _preview_device_row(
            row=row,
            ctx=ctx,
            make="TestMake",
            model="TestModel",
            mfg_slug="test-mfg",
            dt_slug="test-dt",
            source_id="2",
            device_name="test-device-02",
            serial="",
            asset_tag="",
            DeviceType=DeviceType,
            Device=Device,
            Rack=Rack,
        )
        self.assertEqual(result_row.action, "error")
        self.assertEqual(result_row.extra_data["identity_conflict"], "rack_not_found")
        self.assertIn("RACK-99", result_row.detail)

    def test_extra_data_includes_slugs(self):
        """extra_data includes mfg_slug and dt_slug for device rows."""
        from dcim.models import Device, DeviceType, Rack

        row = {
            "_row_number": 3,
            "rack_name": "",
            "u_position": None,
        }
        ctx = ImportContext(
            profile=self.profile, site=self.site, location=None, tenant=None, dry_run=True, result=ImportResult()
        )
        result_row = _preview_device_row(
            row=row,
            ctx=ctx,
            make="Dell",
            model="PowerEdge R640",
            mfg_slug="dell",
            dt_slug="poweredge-r640",
            source_id="3",
            device_name="server-03",
            serial="SN12345",
            asset_tag="AT789",
            DeviceType=DeviceType,
            Device=Device,
            Rack=Rack,
        )
        # Verify we're on the success path (not error)
        self.assertEqual(result_row.action, "create")
        # Verify extra_data contains all required fields
        self.assertIn("source_make", result_row.extra_data)
        self.assertIn("source_model", result_row.extra_data)
        self.assertIn("mfg_slug", result_row.extra_data)
        self.assertIn("dt_slug", result_row.extra_data)
        self.assertEqual(result_row.extra_data["source_make"], "Dell")
        self.assertEqual(result_row.extra_data["source_model"], "PowerEdge R640")
        self.assertEqual(result_row.extra_data["mfg_slug"], "dell")
        self.assertEqual(result_row.extra_data["dt_slug"], "poweredge-r640")
        self.assertIn("u_height", result_row.extra_data)
        self.assertIn("asset_tag", result_row.extra_data)
        self.assertEqual(result_row.extra_data["asset_tag"], "AT789")

    def test_extra_data_includes_slugs_on_error(self):
        """extra_data includes slugs even when device type is not found."""
        from dcim.models import Device, DeviceType, Rack

        # Disable create_missing_device_types to force error path
        self.profile.adapter_config["create_missing_device_types"] = False
        self.profile.save()

        row = {
            "_row_number": 4,
            "rack_name": "",
            "u_position": None,
        }
        ctx = ImportContext(
            profile=self.profile, site=self.site, location=None, tenant=None, dry_run=True, result=ImportResult()
        )
        result_row = _preview_device_row(
            row=row,
            ctx=ctx,
            make="UnknownMfg",
            model="UnknownModel",
            mfg_slug="unknownmfg",
            dt_slug="unknownmodel",
            source_id="4",
            device_name="unknown-device",
            serial="",
            asset_tag="",
            DeviceType=DeviceType,
            Device=Device,
            Rack=Rack,
        )
        # Should be an error action
        self.assertEqual(result_row.action, "error")
        # extra_data should still contain slugs
        self.assertIn("source_make", result_row.extra_data)
        self.assertIn("source_model", result_row.extra_data)
        self.assertIn("mfg_slug", result_row.extra_data)
        self.assertIn("dt_slug", result_row.extra_data)
        self.assertEqual(result_row.extra_data["mfg_slug"], "unknownmfg")
        self.assertEqual(result_row.extra_data["dt_slug"], "unknownmodel")
        self.assertIn("u_height", result_row.extra_data)
        self.assertIn("asset_tag", result_row.extra_data)
        self.assertEqual(result_row.extra_data["asset_tag"], "")

    def test_ip_fields_in_preview_extra_data(self):
        """extra_data includes _ip dict when ip_fields is passed."""
        from dcim.models import Device, DeviceType, Rack

        row = {
            "_row_number": 5,
            "rack_name": "",
            "u_position": None,
        }
        ctx = ImportContext(
            profile=self.profile, site=self.site, location=None, tenant=None, dry_run=True, result=ImportResult()
        )
        result_row = _preview_device_row(
            row=row,
            ctx=ctx,
            make="TestMake",
            model="TestModel",
            mfg_slug="test-mfg",
            dt_slug="test-dt",
            source_id="5",
            device_name="test-device-05",
            serial="",
            asset_tag="",
            DeviceType=DeviceType,
            Device=Device,
            Rack=Rack,
            ip_fields={"primary_ip4": "192.168.1.1/32"},
        )
        # Verify we're on the success path
        self.assertEqual(result_row.action, "create")
        # Verify extra_data contains _ip
        self.assertIn("_ip", result_row.extra_data)
        self.assertEqual(result_row.extra_data["_ip"], {"primary_ip4": "192.168.1.1/32"})

    def test_ip_fields_absent_when_not_provided(self):
        """extra_data does NOT contain _ip key when ip_fields is empty/None."""
        from dcim.models import Device, DeviceType, Rack

        row = {
            "_row_number": 6,
            "rack_name": "",
            "u_position": None,
        }
        ctx = ImportContext(
            profile=self.profile, site=self.site, location=None, tenant=None, dry_run=True, result=ImportResult()
        )
        result_row = _preview_device_row(
            row=row,
            ctx=ctx,
            make="TestMake",
            model="TestModel",
            mfg_slug="test-mfg",
            dt_slug="test-dt",
            source_id="6",
            device_name="test-device-06",
            serial="",
            asset_tag="",
            DeviceType=DeviceType,
            Device=Device,
            Rack=Rack,
        )
        # Verify we're on the success path
        self.assertEqual(result_row.action, "create")
        # Verify extra_data does NOT contain _ip
        self.assertNotIn("_ip", result_row.extra_data)


class ParseIPWithPrefixTest(TestCase):
    """Tests for _parse_ip_with_prefix helper function."""

    def test_parse_ip_with_prefix_adds_cidr(self):
        """_parse_ip_with_prefix adds /32 or /128 prefix if absent."""
        from netbox_data_import.engine import _parse_ip_with_prefix

        # IPv4 without prefix
        self.assertEqual(_parse_ip_with_prefix("192.168.1.1"), "192.168.1.1/32")

        # IPv4 with prefix
        self.assertEqual(_parse_ip_with_prefix("192.168.1.1/24"), "192.168.1.1/24")

        # IPv6 without prefix
        self.assertEqual(_parse_ip_with_prefix("::1"), "::1/128")

        # Invalid IP
        self.assertIsNone(_parse_ip_with_prefix("not-an-ip"))

        # Empty string
        self.assertIsNone(_parse_ip_with_prefix(""))

        # IPv4 with prefix preserved
        self.assertEqual(_parse_ip_with_prefix("192.168.1.1/24"), "192.168.1.1/24")


class IPBuriedInSourceTextTest(TestCase):
    """Source systems export an address inside a label, and the row still names one address."""

    def _parse(self, raw):
        from netbox_data_import.engine import _parse_ip_with_prefix

        return _parse_ip_with_prefix(raw)

    def test_a_trailing_suffix_the_octets_do_not_own_is_dropped(self):
        """`192.0.2.99_5` is one address and a separator the exporter left behind."""
        self.assertEqual(self._parse("192.0.2.99_5"), "192.0.2.99/32")

    def test_an_address_after_a_descriptive_prefix_is_read(self):
        """A VLAN label in front of the address must not cost the row its IP."""
        self.assertEqual(self._parse("Site_Mgmt - 512 - 192.0.2.150"), "192.0.2.150/32")

    def test_a_trailing_comment_is_dropped(self):
        self.assertEqual(self._parse("192.0.2.150 (mgmt)"), "192.0.2.150/32")

    def test_a_buried_address_keeps_the_prefix_length_it_carries(self):
        self.assertEqual(self._parse("mgmt: 192.0.2.66/28"), "192.0.2.66/28")

    def test_a_buried_ipv6_address_is_read(self):
        self.assertEqual(self._parse("mgmt 2001:db8::1 primary"), "2001:db8::1/128")

    def test_text_carrying_no_address_is_still_refused(self):
        """Reading an address out of a label must not invent one out of any text."""
        for raw in ("not-an-ip", "", "Site_Mgmt - 512", "12:30", "999.999.999.999"):
            with self.subTest(raw=raw):
                self.assertIsNone(self._parse(raw))

    def test_an_address_a_word_runs_into_is_refused_rather_than_truncated(self):
        """Hex letters continue an address, so a word can leave a shorter valid one behind.

        Refusing costs the row an address it never really carried. Truncating gives the device a
        different address than the source names, which no later review would catch.
        """
        for raw in ("2001:db8::1backup", "10.0.0.1and10.0.0.2", "192.0.2.1a"):
            with self.subTest(raw=raw):
                self.assertIsNone(self._parse(raw))

    def test_a_separator_that_is_not_a_word_still_ends_the_address(self):
        """`_`, `-`, `/` and whitespace end a value; only letters and digits continue one."""
        self.assertEqual(self._parse("192.0.2.7_vlan"), "192.0.2.7/32")
        self.assertEqual(self._parse("[192.0.2.8]"), "192.0.2.8/32")
        self.assertEqual(self._parse("192.0.2.9, 192.0.2.10"), "192.0.2.9/32")


class EnsureDeviceTypeExecuteModeTest(TestCase):
    """Tests that _ensure_device_type never appends RowResult rows in execute mode."""

    def setUp(self):
        self.profile = _make_profile("EnsureDT")

    def test_execute_mode_no_row_results_create_missing_false(self):
        """Execute mode with create_missing_device_types=False appends no RowResult rows."""
        from dcim.models import DeviceType, Manufacturer

        self.profile.adapter_config["create_missing_device_types"] = False
        result = ImportResult()
        row = {"_row_number": 1, "source_id": "1"}
        ctx = ImportContext(profile=self.profile, site=None, location=None, tenant=None, dry_run=False, result=result)
        _ensure_device_type(
            "unknown-mfg",
            "unknown-dt",
            "Unknown Make",
            "Unknown Model",
            1,
            set(),
            ctx,
            row,
            Manufacturer,
            DeviceType,
        )
        device_type_rows = [r for r in result.rows if r.object_type == "device_type"]
        self.assertEqual(device_type_rows, [], "Execute mode must not append device_type RowResult rows")

    def test_execute_mode_no_row_results_create_missing_true(self):
        """Execute mode with create_missing_device_types=True appends no RowResult rows (creates silently)."""
        from dcim.models import DeviceType, Manufacturer

        self.profile.adapter_config["create_missing_device_types"] = True
        result = ImportResult()
        row = {"_row_number": 1, "source_id": "1"}
        ctx = ImportContext(profile=self.profile, site=None, location=None, tenant=None, dry_run=False, result=result)
        _ensure_device_type(
            "silent-mfg",
            "silent-dt",
            "Silent Make",
            "Silent Model",
            1,
            set(),
            ctx,
            row,
            Manufacturer,
            DeviceType,
        )
        device_type_rows = [r for r in result.rows if r.object_type == "device_type"]
        self.assertEqual(device_type_rows, [], "Execute mode must not append device_type RowResult rows")
        # Verify the device type was actually created in DB
        self.assertTrue(DeviceType.objects.filter(manufacturer__slug="silent-mfg", slug="silent-dt").exists())

    def test_dry_run_appends_error_row_when_create_missing_false(self):
        """Dry-run with create_missing_device_types=False does append an error RowResult."""
        from dcim.models import DeviceType, Manufacturer

        self.profile.adapter_config["create_missing_device_types"] = False
        result = ImportResult()
        row = {"_row_number": 1, "source_id": "1"}
        ctx = ImportContext(profile=self.profile, site=None, location=None, tenant=None, dry_run=True, result=result)
        _ensure_device_type(
            "dry-mfg",
            "dry-dt",
            "Dry Make",
            "Dry Model",
            1,
            set(),
            ctx,
            row,
            Manufacturer,
            DeviceType,
        )
        device_type_rows = [r for r in result.rows if r.object_type == "device_type"]
        self.assertEqual(len(device_type_rows), 1)
        self.assertEqual(device_type_rows[0].action, "error")


class ParseFileEdgeCasesTest(TestCase):
    """Tests for parse_file edge cases: empty rows and missing column headers."""

    def test_empty_rows_are_skipped(self):
        """parse_file skips fully-empty data rows (line 192 coverage)."""
        import openpyxl
        from io import BytesIO

        profile = _make_profile("EmptyRowTest")

        # Build an xlsx with one data row and one empty row
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(
            [
                "Id",
                "Rack",
                "Name",
                "Class",
                "Make",
                "Model",
                "UHeight",
                "UPosition",
                "Serial Number",
                "Asset Tag",
                "Status",
            ]
        )
        ws.append(["SRC001", "Rack-01", "dev-01", "Server", "Dell", "R740", "1", "1", "", "", "active"])
        ws.append([None, None, None, None, None, None, None, None, None, None, None])  # empty row
        ws.append(["SRC002", "Rack-01", "dev-02", "Server", "Dell", "R740", "1", "2", "", "", "active"])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        rows = parse_file(buf, profile)
        # Only 2 non-empty rows
        self.assertEqual(len(rows), 2)

    def test_mapping_with_missing_source_column_skips(self):
        """parse_file silently skips column mappings whose header doesn't exist in the file (line 198)."""
        import openpyxl
        from io import BytesIO

        profile = _make_profile("MissingColTest")
        # Add a mapping for a column that doesn't exist in the file
        ColumnMapping.objects.create(profile=profile, source_column="NonExistentCol", target_field="tenant")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["Id", "Name", "Class"])
        ws.append(["SRC001", "dev-01", "Server"])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        rows = parse_file(buf, profile)
        self.assertEqual(len(rows), 1)
        # tenant should be absent (no mapping target applied since column didn't exist)
        self.assertNotIn("tenant", rows[0])


class FieldDiffComputationTest(TestCase):
    """Tests for field_diff computation in _preview_device_row."""

    def setUp(self):
        from dcim.models import DeviceRole, DeviceType, Manufacturer, Site

        self.site = Site.objects.create(name="Diff Site", slug="diff-site")
        self.profile = _make_profile("FieldDiff")

        mfg = Manufacturer.objects.create(name="TestMfg", slug="testmfg")
        self.device_type = DeviceType.objects.create(
            manufacturer=mfg,
            model="TestModel",
            slug="testmodel",
            u_height=1,
        )
        self.role = DeviceRole.objects.create(name="Server", slug="server", color="000000")

    def _make_existing_device(self, name="existing-server", serial="OLD123", asset_tag="OLD-TAG"):
        from dcim.models import Device

        return Device.objects.create(
            name=name,
            site=self.site,
            device_type=self.device_type,
            role=self.role,
            serial=serial,
            asset_tag=asset_tag,
            status="active",
        )

    def _call_preview(
        self,
        device_name,
        serial,
        asset_tag,
        device_status="active",
        ip_fields=None,
        u_position=None,
        device_face=None,
        device_airflow=None,
    ):
        from dcim.models import Device, DeviceType, Rack

        row = {"_row_number": 1, "rack_name": "", "u_position": u_position}
        ctx = ImportContext(
            profile=self.profile,
            site=self.site,
            location=None,
            tenant=None,
            dry_run=True,
            result=ImportResult(),
        )
        return _preview_device_row(
            row=row,
            ctx=ctx,
            make="TestMfg",
            model="TestModel",
            mfg_slug="testmfg",
            dt_slug="testmodel",
            source_id="99",
            device_name=device_name,
            serial=serial,
            asset_tag=asset_tag,
            DeviceType=DeviceType,
            Device=Device,
            Rack=Rack,
            ip_fields=ip_fields,
            device_face=device_face,
            device_airflow=device_airflow,
            device_status=device_status,
            u_position=u_position,
        )

    def _call_preview_with_row(self, device_name, extra_row_fields=None, **kwargs):
        """Like _call_preview but allows injecting extra row dict keys."""
        from dcim.models import DeviceType, Device, Rack
        from netbox_data_import.engine import ImportContext, _preview_device_row

        row = {
            "_row_number": 1,
            "source_id": "test-id",
            "device_name": device_name,
            "rack_name": None,
            "make": "TestMfg",
            "model": "TestModel",
            "u_height": 1,
            "serial": None,
            "asset_tag": None,
        }
        if extra_row_fields:
            row.update(extra_row_fields)
        ctx = ImportContext(
            profile=self.profile,
            site=self.site,
            location=None,
            tenant=None,
            dry_run=True,
            result=ImportResult(),
        )
        return _preview_device_row(
            row,
            ctx,
            make="TestMfg",
            model="TestModel",
            mfg_slug="testmfg",
            dt_slug="testmodel",
            source_id="test-id",
            device_name=device_name,
            serial=row.get("serial"),
            asset_tag=row.get("asset_tag"),
            DeviceType=DeviceType,
            Device=Device,
            Rack=Rack,
            **kwargs,
        )

    def test_field_diff_on_update_row(self):
        """Update rows include field_diff with changed serial and asset_tag."""
        self._make_existing_device(serial="OLD123", asset_tag="OLD-TAG")
        result = self._call_preview("existing-server", serial="NEW456", asset_tag="A-001")
        self.assertEqual(result.action, "update")
        self.assertIn("field_diff", result.extra_data)
        diff = result.extra_data["field_diff"]
        self.assertIn("serial", diff)
        self.assertEqual(diff["serial"]["netbox"], "OLD123")
        self.assertEqual(diff["serial"]["file"], "NEW456")
        self.assertIn("asset_tag", diff)
        self.assertEqual(diff["asset_tag"]["netbox"], "OLD-TAG")
        self.assertEqual(diff["asset_tag"]["file"], "A-001")

    def test_field_diff_absent_on_create_row(self):
        """Create rows must not have field_diff in extra_data."""
        result = self._call_preview("brand-new-device", serial="SN001", asset_tag="AT001")
        self.assertEqual(result.action, "create")
        self.assertNotIn("field_diff", result.extra_data)

    def test_field_diff_absent_on_skip_row(self):
        """Skip rows (update_existing=False) must not have field_diff in extra_data."""
        self.profile.adapter_config["update_existing"] = False
        self.profile.save()
        self._make_existing_device(serial="OLD123", asset_tag="OLD-TAG")
        result = self._call_preview("existing-server", serial="NEW456", asset_tag="A-001")
        self.assertEqual(result.action, "skip")
        self.assertNotIn("field_diff", result.extra_data)

    def _assign_primary_ip4(self, device, address):
        """Give a device the address on a real interface, the way the writer leaves it."""
        from dcim.models import Interface
        from ipam.models import IPAddress

        interface = Interface.objects.create(device=device, name="mgmt", type="1000base-t")
        ip = IPAddress.objects.create(address=address, assigned_object=interface)
        device.primary_ip4 = ip
        device.save()
        return ip

    def test_field_diff_reports_an_ip_the_import_would_assign(self):
        """The import writes this field, so a preview that hides it cannot say what it will do."""
        self._make_existing_device(serial="OLD123", asset_tag="OLD-TAG")
        result = self._call_preview(
            "existing-server",
            serial="NEW456",
            asset_tag="A-001",
            ip_fields={"primary_ip4": "10.0.0.1/32"},
        )
        self.assertEqual(result.action, "update")
        diff = result.extra_data["field_diff"]
        self.assertIn("primary_ip4", diff)
        self.assertEqual(diff["primary_ip4"]["file"], "10.0.0.1/32")
        self.assertEqual(diff["primary_ip4"]["netbox"], "")

    def test_field_diff_reports_an_ip_the_import_would_replace(self):
        """The operator has to see which address the row takes the device away from."""
        device = self._make_existing_device(serial="OLD123", asset_tag="OLD-TAG")
        self._assign_primary_ip4(device, "10.0.0.9/32")
        result = self._call_preview(
            "existing-server",
            serial="NEW456",
            asset_tag="A-001",
            ip_fields={"primary_ip4": "10.0.0.1/32"},
        )
        diff = result.extra_data["field_diff"]["primary_ip4"]
        self.assertEqual(diff["netbox"], "10.0.0.9/32")
        self.assertEqual(diff["file"], "10.0.0.1/32")
        self.assertEqual(diff["ip_target"], "would go to mgmt", "the row must name the interface it would use")

    def test_field_diff_leaves_out_an_ip_the_device_already_carries(self):
        """An address that agrees is not a difference, whichever way each side spells it."""
        device = self._make_existing_device(serial="OLD123", asset_tag="OLD-TAG")
        self._assign_primary_ip4(device, "10.0.0.1/32")
        result = self._call_preview(
            "existing-server",
            serial="NEW456",
            asset_tag="A-001",
            ip_fields={"primary_ip4": "10.0.0.1/32"},
        )
        self.assertNotIn("primary_ip4", result.extra_data["field_diff"])

    def test_field_diff_leaves_out_an_ip_field_the_row_does_not_supply(self):
        """A profile that maps no IPv6 column must not report an empty IPv6 difference."""
        self._make_existing_device(serial="OLD123", asset_tag="OLD-TAG")
        result = self._call_preview(
            "existing-server",
            serial="NEW456",
            asset_tag="A-001",
            ip_fields={"primary_ip4": "10.0.0.1/32"},
        )
        diff = result.extra_data["field_diff"]
        self.assertNotIn("primary_ip6", diff)
        self.assertNotIn("oob_ip", diff)

    def test_field_diff_excludes_matching_fields(self):
        """field_diff must not include a field when xls value matches the existing device value."""
        self._make_existing_device(serial="SAME-SERIAL", asset_tag="DIFF-TAG")
        result = self._call_preview("existing-server", serial="SAME-SERIAL", asset_tag="NEW-TAG")
        self.assertEqual(result.action, "update")
        diff = result.extra_data.get("field_diff", {})
        self.assertNotIn("serial", diff)
        self.assertIn("asset_tag", diff)

    def test_field_diff_no_u_height_when_matches(self):
        """u_height must not appear in diff when XLS value equals device type u_height."""
        self._make_existing_device(serial="S1", asset_tag="A1")
        # row has no u_height key → defaults to 1; device_type was created with u_height=1
        result = self._call_preview("existing-server", serial="S1", asset_tag="A1")
        diff = result.extra_data.get("field_diff", {})
        self.assertNotIn("u_height", diff, "u_height must not appear in diff when values match")

    def test_extra_data_includes_face_airflow_status(self):
        """_preview_device_row must include face, airflow, status in extra_data."""
        result = self._call_preview(
            "new-device-face",
            serial="SN-FIELDS",
            asset_tag="AT-FIELDS",
            device_face="front",
            device_airflow="front-to-rear",
            device_status="staged",
        )
        self.assertEqual(result.action, "create")
        self.assertEqual(result.extra_data.get("face"), "front")
        self.assertEqual(result.extra_data.get("airflow"), "front-to-rear")
        self.assertEqual(result.extra_data.get("status"), "staged")

    def test_extra_data_includes_extra_columns(self):
        """_preview_device_row must pass through _extra_columns from the row."""
        result = self._call_preview_with_row(
            "new-device-extra",
            extra_row_fields={"_extra_columns": {"cf_location": "DC1"}},
        )
        self.assertEqual(result.action, "create")
        self.assertEqual(result.extra_data.get("extra_columns"), {"cf_location": "DC1"})

    def test_extra_data_includes_conflicts(self):
        """_preview_device_row must pass through _conflicts from the row."""
        result = self._call_preview_with_row(
            "new-device-conflict",
            extra_row_fields={"_conflicts": {"serial": {"Serial Number": "AAA", "Service Tag": "BBB"}}},
        )
        self.assertEqual(result.action, "create")
        self.assertIn("conflicts", result.extra_data)
        self.assertIn("serial", result.extra_data["conflicts"])

    def test_field_diff_u_position_float_vs_int(self):
        """u_position '35.0' from source file vs 35 (int) from NetBox must NOT appear in diff."""
        self._make_existing_device(serial="SN-POS", asset_tag="AT-POS")
        from dcim.models import Device

        dev = Device.objects.get(name="existing-server")
        dev.position = 35
        dev.save()

        # Pass "35.0" to simulate the float-like string that comes from an Excel cell
        result = self._call_preview("existing-server", serial="SN-POS", asset_tag="AT-POS", u_position="35.0")
        diff = result.extra_data.get("field_diff", {})
        self.assertNotIn("u_position", diff, "'35.0' vs 35 must not appear as a diff after float normalisation")

    def test_field_diff_text_fields_use_exact_comparison(self):
        """Text fields (serial, asset_tag, device_name) use exact str comparison, not float normalization.

        Serial numbers and asset tags are identifiers where '35.0' and '35' are meaningfully
        different — float normalization must NOT be applied to them.
        """
        from unittest.mock import MagicMock

        from netbox_data_import.engine import _compute_field_differences

        mock_device = MagicMock()
        mock_device.name = "some-device"
        mock_device.status = "active"
        mock_device.serial = "35.0"
        mock_device.asset_tag = "AT-001"
        mock_device.face = ""
        mock_device.airflow = ""
        mock_device.position = None
        mock_device.device_type_id = None

        diff, _informational = _compute_field_differences(
            matched_device=mock_device,
            device_name="some-device",
            serial="35",
            asset_tag="AT-001",
            device_face=None,
            device_airflow=None,
            device_status="active",
            u_height=1,
            u_position=None,
        )

        self.assertIn("serial", diff, "'35' vs '35.0' must appear as a diff for serial (text field)")
        self.assertEqual(diff["serial"]["netbox"], "35.0")
        self.assertEqual(diff["serial"]["file"], "35")

    def test_extra_data_includes_netbox_device_id_on_update(self):
        """Update rows must include netbox_device_id in extra_data equal to matched device PK."""
        device = self._make_existing_device(serial="SN-ID", asset_tag="AT-ID")
        result = self._call_preview("existing-server", serial="SN-ID", asset_tag="NEW-TAG")
        self.assertEqual(result.action, "update")
        self.assertIn("netbox_device_id", result.extra_data)
        self.assertEqual(result.extra_data["netbox_device_id"], device.pk)

    def test_netbox_device_id_absent_on_skip_row(self):
        """netbox_device_id must NOT be present on skip rows (update_existing=False)."""
        self.profile.adapter_config["update_existing"] = False
        self.profile.save()
        self._make_existing_device(serial="SN-SKIP", asset_tag=None)
        result = self._call_preview("existing-server", serial="SN-SKIP", asset_tag=None)
        self.assertEqual(result.action, "skip")
        self.assertNotIn("netbox_device_id", result.extra_data)

    def test_netbox_device_id_absent_on_create_row(self):
        """netbox_device_id must NOT be present on create rows (no matching device)."""
        result = self._call_preview("brand-new-device-xyz", serial="SN-NEW-XYZ", asset_tag=None)
        self.assertEqual(result.action, "create")
        self.assertNotIn("netbox_device_id", result.extra_data)


class NormalizeForCompareTest(TestCase):
    """Tests for _normalize_for_compare helper."""

    def test_integer_string_unchanged(self):
        self.assertEqual(_normalize_for_compare("35"), "35")

    def test_float_whole_number_normalized(self):
        """35.0 → '35'"""
        self.assertEqual(_normalize_for_compare("35.0"), "35")

    def test_float_whole_number_direct(self):
        """float(35.0) → '35'"""
        self.assertEqual(_normalize_for_compare(35.0), "35")

    def test_float_with_fraction_unchanged(self):
        """1.5 stays '1.5'"""
        self.assertEqual(_normalize_for_compare(1.5), "1.5")

    def test_none_returns_empty(self):
        self.assertEqual(_normalize_for_compare(None), "")

    def test_non_numeric_string_unchanged(self):
        self.assertEqual(_normalize_for_compare("ABC-123"), "ABC-123")

    def test_zero(self):
        self.assertEqual(_normalize_for_compare(0), "0")

    def test_zero_float(self):
        self.assertEqual(_normalize_for_compare(0.0), "0")

    def test_infinity_string_returns_stripped_value(self):
        for value in ("inf", "Infinity", "-inf"):
            with self.subTest(value=value):
                self.assertEqual(_normalize_for_compare(value), value.strip())


class ExistingRackPreviewActionTest(TestCase):
    """A rack row reports what it writes, so a row that writes nothing is not an update."""

    def setUp(self):
        """Create the profile, the site, and the rack the source row names."""
        from dcim.models import Rack, Site

        self.profile = _make_profile("RackNoopProfile")
        self.site = Site.objects.create(name="Rack Noop Site", slug="rack-noop-site")
        self.rack = Rack.objects.create(name="T1", site=self.site, u_height=42, serial="RACK-SERIAL")

    def _preview(self, **overrides):
        """Run one rack row through the preview and return its result row."""
        row = {"_row_number": 2, "source_id": "261988", "rack_name": "T1", "device_class": "Cabinet"}
        row.update(overrides)
        result = run_import([row], self.profile, {"site": self.site}, dry_run=True)
        return next(r for r in result.rows if r.object_type == "rack")

    def test_a_row_that_changes_nothing_is_not_reported_as_an_update(self):
        """The row repeats the rack's stored height and serial, so the import would write nothing."""
        preview = self._preview(u_height=42, serial="RACK-SERIAL")

        self.assertEqual(preview.action, "skip")
        self.assertIn("changes nothing", preview.detail)

    def test_a_row_that_changes_a_field_is_still_an_update(self):
        """A different height is a real write, so the row keeps reporting it."""
        preview = self._preview(u_height=47, serial="RACK-SERIAL")

        self.assertEqual(preview.action, "update")
        self.assertEqual(preview.detail, "Rack 'T1' already exists")

    def test_a_rack_row_the_profile_never_updates_is_not_a_no_op(self):
        """`update_existing` off is a different answer from a row that had nothing to write."""
        self.profile.adapter_config = {**self.profile.adapter_config, "update_existing": False}
        self.profile.save()

        preview = self._preview(u_height=42, serial="RACK-SERIAL")

        self.assertEqual(preview.action, "skip")
        self.assertFalse(preview.extra_data.get("writes_nothing"), preview.extra_data)
        self.assertIn("update_existing=False", preview.detail)

    def test_the_import_agrees_with_the_preview_about_a_row_that_writes_nothing(self):
        """The execute guard compares the writer's action to the previewed one, so they must agree."""
        from netbox_data_import.views import _import_intents

        row = {
            "_row_number": 2,
            "source_id": "261988",
            "rack_name": "T1",
            "device_class": "Cabinet",
            "u_height": 42,
            "serial": "RACK-SERIAL",
        }
        preview = run_import([dict(row)], self.profile, {"site": self.site}, dry_run=True)

        written = run_import(
            [dict(row)],
            self.profile,
            {"site": self.site},
            dry_run=False,
            expected_intents=_import_intents(preview),
        )

        rack_row = next(r for r in written.rows if r.object_type == "rack")
        self.assertNotEqual(rack_row.action, "error", rack_row.detail)
        self.assertEqual(rack_row.action, "skip", rack_row.detail)


class RackExpandedInTheSameBatchTest(TestCase):
    """A device may take the space the same batch adds to its rack."""

    def setUp(self):
        """Create the rack the batch grows and the role its new device needs."""
        from dcim.models import DeviceRole, Rack, Site

        self.profile = _make_profile("RackGrowthProfile")
        self.site = Site.objects.create(name="Rack Growth Site", slug="rack-growth-site")
        DeviceRole.objects.create(name="Server", slug="server")
        self.rack = Rack.objects.create(name="T1", site=self.site, u_height=20)

    def _rows(self):
        """Return a batch that raises the rack's height, then fills the space it added."""
        return [
            {
                "_row_number": 2,
                "source_id": "RACK-1",
                "rack_name": "T1",
                "device_class": "Cabinet",
                "u_height": 42,
            },
            {
                "_row_number": 3,
                "source_id": "DEV-1",
                "rack_name": "T1",
                "device_name": "grown-device",
                "device_class": "Server",
                "u_position": 30,
                "face": "front",
                "make": "GrowthMfg",
                "model": "GrowthModel",
            },
        ]

    def _device_row(self, result):
        """Return the one device row of *result*."""
        return next(row for row in result.rows if row.object_type == "device")

    def test_the_preview_places_the_device_in_the_space_the_batch_adds(self):
        """U30 is outside the stored 20U rack and inside the 42U this batch leaves behind."""
        result = run_import(self._rows(), self.profile, {"site": self.site}, dry_run=True)

        device_row = self._device_row(result)
        self.assertNotEqual(device_row.action, "error", device_row.detail)
        self.assertEqual(device_row.action, "create", device_row.detail)

    def test_the_import_agrees_with_the_preview(self):
        """A preview that errors and an import that writes disagree, and the batch rolls back."""
        from dcim.models import Device

        from netbox_data_import.views import _import_intents

        preview = run_import([dict(row) for row in self._rows()], self.profile, {"site": self.site}, dry_run=True)

        written = run_import(
            [dict(row) for row in self._rows()],
            self.profile,
            {"site": self.site},
            dry_run=False,
            expected_intents=_import_intents(preview),
        )

        device_row = self._device_row(written)
        self.assertNotEqual(device_row.action, "error", device_row.detail)
        self.rack.refresh_from_db()
        self.assertEqual(self.rack.u_height, 42)
        self.assertEqual(Device.objects.get(name="grown-device").position, 30)


class MatchedDevicePreviewActionTest(TestCase):
    """A matched device row reports what it writes, so a row that writes nothing is not an update."""

    def setUp(self):
        """Create the site, roles, and profile the workbook rows need."""
        from dcim.models import DeviceRole, Site

        self.site = Site.objects.create(name="Device Noop Site", slug="device-noop-site")
        DeviceRole.objects.create(name="Server", slug="server")
        DeviceRole.objects.create(name="Network Switch", slug="network-switch")
        self.profile = _make_profile("DeviceNoopProfile")

    def _rows(self):
        """Return the workbook's parsed rows."""
        with open(FIXTURE_PATH, "rb") as f:
            return parse_file(f, self.profile)

    def _run(self, rows, *, dry_run, expected_intents=None):
        """Run one import pass over *rows*."""
        return run_import(
            [dict(r) for r in rows],
            self.profile,
            {"site": self.site},
            dry_run=dry_run,
            expected_intents=expected_intents,
        )

    def _device_row(self, result, name):
        """Return the result row for one device by name."""
        return next(r for r in result.rows if r.object_type == "device" and r.name == name)

    def test_a_second_preview_of_an_imported_row_writes_nothing(self):
        """Re-previewing a file already imported is the case an operator hits every day."""
        rows = self._rows()
        first = self._run(rows, dry_run=False)
        self.assertEqual([r.detail for r in first.rows if r.action == "error"], [])

        preview = self._run(rows, dry_run=True)

        device_row = self._device_row(preview, "server-01")
        self.assertEqual(device_row.action, "skip", device_row.detail)
        self.assertIn("writes nothing", device_row.detail)

    def test_a_changed_field_is_still_an_update(self):
        """A serial the device does not carry is a real write, so the row keeps reporting it."""
        rows = self._rows()
        self._run(rows, dry_run=False)
        for row in rows:
            if row.get("device_name") == "server-01":
                row["serial"] = "CHANGED-SERIAL"

        preview = self._run(rows, dry_run=True)

        device_row = self._device_row(preview, "server-01")
        self.assertEqual(device_row.action, "update", device_row.detail)

    def test_a_row_that_still_has_to_record_its_import_state_is_an_update(self):
        """The import record is what a later run reads, so restoring it is a write like any other."""
        from netbox_data_import.models import DeviceImportSource

        rows = self._rows()
        self._run(rows, dry_run=False)
        DeviceImportSource.objects.filter(device__name="server-01").delete()

        preview = self._run(rows, dry_run=True)

        device_row = self._device_row(preview, "server-01")
        self.assertEqual(device_row.action, "update", device_row.detail)

    def test_a_row_that_writes_nothing_still_names_the_device_it_matched(self):
        """The execute guard only compares an object id when the preview recorded one."""
        rows = self._rows()
        self._run(rows, dry_run=False)

        preview = self._run(rows, dry_run=True)

        device_row = self._device_row(preview, "server-01")
        self.assertEqual(device_row.action, "skip", device_row.detail)
        self.assertTrue(device_row.extra_data.get("netbox_device_id"), device_row.extra_data)

    def test_a_zero_u_device_still_carrying_a_position_is_an_update(self):
        """A zero-U type has its position cleared too, and an equal position compares as settled."""
        from dcim.models import Device, DeviceType

        rows = self._rows()
        self._run(rows, dry_run=False)
        device = Device.objects.get(name="server-01")
        DeviceType.objects.filter(pk=device.device_type_id).update(u_height=0)
        Device.objects.filter(pk=device.pk).update(position=None, face="")
        Device.objects.filter(pk=device.pk).update(position=10)

        preview = self._run(rows, dry_run=True)

        device_row = self._device_row(preview, "server-01")
        self.assertEqual(device_row.action, "update", device_row.detail)

    def test_a_zero_u_device_still_carrying_a_face_is_an_update(self):
        """The writer clears face for a zero-U type, and a blank source face is never compared."""
        from dcim.models import Device, DeviceType

        rows = self._rows()
        self._run(rows, dry_run=False)
        device = Device.objects.get(name="server-01")
        DeviceType.objects.filter(pk=device.device_type_id).update(u_height=0)
        Device.objects.filter(pk=device.pk).update(face="front")
        for row in rows:
            row.pop("face", None)

        preview = self._run(rows, dry_run=True)

        device_row = self._device_row(preview, "server-01")
        self.assertEqual(device_row.action, "update", device_row.detail)

    def test_a_row_that_still_has_to_bind_its_device_is_an_update(self):
        """The binding is what matches this row next time, so creating it is a write."""
        rows = self._rows()
        self._run(rows, dry_run=False)
        self.profile.device_matches.all().delete()

        preview = self._run(rows, dry_run=True)

        device_row = self._device_row(preview, "server-01")
        self.assertEqual(device_row.action, "update", device_row.detail)

    def test_the_import_agrees_with_the_preview_about_a_row_that_writes_nothing(self):
        """The execute guard compares the writer's action to the previewed one, so they must agree."""
        from netbox_data_import.views import _import_intents

        rows = self._rows()
        self._run(rows, dry_run=False)
        preview = self._run(rows, dry_run=True)

        written = self._run(rows, dry_run=False, expected_intents=_import_intents(preview))

        device_row = self._device_row(written, "server-01")
        self.assertNotEqual(device_row.action, "error", device_row.detail)
        self.assertEqual(device_row.action, "skip", device_row.detail)


class DuplicateSerialReportTest(TestCase):
    """A refused row has to say which other row holds the serial, or nothing can be done about it."""

    def setUp(self):
        """Create the site and role two source rows sharing one serial need."""
        from dcim.models import DeviceRole, Site

        self.profile = _make_profile("DuplicateSerialProfile")
        self.site = Site.objects.create(name="Dup Serial Site", slug="dup-serial-site")
        DeviceRole.objects.create(name="Server", slug="server")

    def _rows(self):
        """Return two devices the source file gives the same serial."""
        return [
            {
                "_row_number": 4,
                "source_id": "DUP-SRC-A",
                "device_name": "PROD-A",
                "device_class": "Server",
                "serial": "SHARED-SERIAL-1",
                "make": "DupMfg",
                "model": "DupModel",
            },
            {
                "_row_number": 9,
                "source_id": "DUP-SRC-B",
                "device_name": "PROD-B",
                "device_class": "Server",
                "serial": "SHARED-SERIAL-1",
                "make": "DupMfg",
                "model": "DupModel",
            },
        ]

    def _device_rows(self):
        """Return the device rows of one preview, keyed by source ID."""
        result = run_import(self._rows(), self.profile, {"site": self.site}, dry_run=True)
        return {row.source_id: row for row in result.rows if row.object_type == "device"}

    def test_each_refused_row_names_the_other_row_holding_the_serial(self):
        """The operator has to find the other row to decide which one gives the serial up."""
        rows = self._device_rows()

        self.assertEqual(rows["DUP-SRC-A"].action, "error")
        self.assertIn("row 9", rows["DUP-SRC-A"].detail)
        self.assertIn("row 4", rows["DUP-SRC-B"].detail)

    def test_each_refused_row_carries_the_other_rows_for_the_template(self):
        """The preview offers an action per row, so it needs the collision as data, not prose."""
        row = self._device_rows()["DUP-SRC-A"]

        self.assertEqual(row.extra_data.get("identity_conflict"), "duplicate_serial")
        self.assertEqual(row.extra_data.get("duplicate_serial"), "SHARED-SERIAL-1")
        self.assertEqual(row.extra_data.get("duplicate_serial_rows"), [9])

    def test_ignoring_the_serial_on_one_row_releases_both(self):
        """One row gives the serial up, and the other keeps it, so the import can go ahead."""
        from netbox_data_import.models import SourceResolution

        SourceResolution.objects.create(
            profile=self.profile,
            source_id="DUP-SRC-A",
            source_column="serial",
            original_value="SHARED-SERIAL-1",
            resolved_fields={"serial": ""},
        )

        from netbox_data_import.engine import derive_effective_rows

        rows = self._rows()
        derived = derive_effective_rows(rows, self.profile)
        result = run_import(derived, self.profile, {"site": self.site}, dry_run=True)

        by_source = {row.source_id: row for row in result.rows if row.object_type == "device"}
        self.assertNotEqual(by_source["DUP-SRC-A"].action, "error", by_source["DUP-SRC-A"].detail)
        self.assertNotEqual(by_source["DUP-SRC-B"].action, "error", by_source["DUP-SRC-B"].detail)


class IpAlreadyAssignedTest(TestCase):
    """The writer resolves an address to one IPAddress, so only a unique address is settled."""

    def setUp(self):
        """Create a device whose interface carries the address it uses as its primary IPv4."""
        from dcim.models import Device, DeviceRole, DeviceType, Interface, Manufacturer, Site
        from ipam.models import IPAddress

        site = Site.objects.create(name="Ip Site", slug="ip-site")
        manufacturer = Manufacturer.objects.create(name="IpMfg", slug="ip-mfg")
        device_type = DeviceType.objects.create(manufacturer=manufacturer, model="IpModel", slug="ip-model")
        role = DeviceRole.objects.create(name="IpRole", slug="ip-role")
        self.device = Device.objects.create(name="ip-device", site=site, device_type=device_type, role=role)
        # The writer only assigns through an interface, so the settled state has to carry one.
        self.interface = Interface.objects.create(device=self.device, name="eth0", type="1000base-t")
        self.address = IPAddress.objects.create(address="10.0.0.5/24", assigned_object=self.interface)
        Device.objects.filter(pk=self.device.pk).update(primary_ip4=self.address)
        self.device.refresh_from_db()

    def test_the_address_the_device_carries_is_settled(self):
        """One address, one IPAddress: the writer can only resolve to what the device already has."""
        self.assertTrue(_ip_already_assigned(self.device, "primary_ip4", "10.0.0.5/24"))

    def test_a_second_row_holding_the_same_address_is_not_settled(self):
        """The writer filters by address and VRF, so it could resolve to the other object."""
        from ipam.models import IPAddress

        IPAddress.objects.create(address="10.0.0.5/24")

        self.assertFalse(_ip_already_assigned(self.device, "primary_ip4", "10.0.0.5/24"))

    def test_a_different_address_is_not_settled(self):
        """A row naming another address is a real write."""
        self.assertFalse(_ip_already_assigned(self.device, "primary_ip4", "10.0.0.6/24"))

    def test_an_address_no_interface_carries_is_not_settled(self):
        """Without an interface the writer cannot assign, so it records the address as unassigned."""
        self.address.assigned_object = None
        self.address.save()
        self.device.refresh_from_db()

        self.assertFalse(_ip_already_assigned(self.device, "primary_ip4", "10.0.0.5/24"))

    def test_an_interface_in_another_vrf_is_not_settled(self):
        """The writer resolves the address by the interface's VRF, which would miss this object."""
        from ipam.models import VRF

        self.interface.vrf = VRF.objects.create(name="IpVrf")
        self.interface.save()
        self.device.refresh_from_db()

        self.assertFalse(_ip_already_assigned(self.device, "primary_ip4", "10.0.0.5/24"))
