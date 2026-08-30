# SPDX-License-Identifier: Apache-2.0
# Copyright (C) 2025 Marcin Zieba <marcinpsk@gmail.com>
"""Integration tests for native NetBox contact synchronization."""

from io import BytesIO
from pathlib import Path
import re
import subprocess
from unittest import skipUnless

from django.contrib.auth import get_user_model
from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import ValidationError
from django.db import connection
from django.test import SimpleTestCase, TestCase
from django.test.utils import CaptureQueriesContext
from dcim.models import Device, DeviceRole, DeviceType, Manufacturer, Site
from tenancy.models import Contact, ContactAssignment, ContactRole

from netbox_data_import.adapter_config import interpreter_config_for
from netbox_data_import.adapters import FlatWorkbookAdapter
from netbox_data_import.contact_resolution import (
    ContactResolutionRequired,
    ContactReview,
    PrimaryContactResolver,
)
from netbox_data_import.models import ClassRoleMapping, ColumnMapping, ImportProfile, stored_import_source
from netbox_data_import.object_permissions import ObjectPermissionDenied
from netbox_data_import.tests.helpers import set_import_source, user_with_object_permission


LOCAL_EXAMPLE_PATH = Path(__file__).resolve().parents[3] / "libre" / "example.xlsx"


def _interpret(workbook, profile):
    """Interpret workbook bytes through the registered flat-workbook adapter."""
    return list(FlatWorkbookAdapter.interpret(workbook.read(), interpreter_config_for(profile)).rows)


class ContactMappingWorkbookTest(TestCase):
    """Exercise Contact column mappings with a portable workbook."""

    @staticmethod
    def _workbook():
        """Return one in-memory workbook with the supported Contact columns."""
        import openpyxl

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        worksheet.append(["Primary Contact", "Contact", "Contact Number", "Owner"])
        worksheet.append(
            [
                "primary.contact@example.invalid",
                "contact@example.invalid",
                "+1 202-555-0100",
                "Example Owner",
            ]
        )
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    def test_primary_contact_column_maps_to_native_contact_target(self):
        """A Primary Contact cell parses as a native Contact value."""
        profile = ImportProfile.objects.create(name="Portable Contact Mapping", adapter_config={"sheet_name": "Data"})
        ColumnMapping.objects.create(
            profile=profile,
            source_column="Primary Contact",
            target_field="primary_contact",
        )

        rows = _interpret(self._workbook(), profile)

        self.assertEqual(rows[0]["primary_contact"], "primary.contact@example.invalid")
        self.assertNotIn("_extra_columns", rows[0])

    def test_contact_candidate_columns_are_collected_per_row(self):
        """Every configured candidate column remains selectable per row."""
        profile = ImportProfile.objects.create(
            name="Portable Contact Candidates", adapter_config={"sheet_name": "Data"}
        )
        candidate_columns = {"Primary Contact", "Contact", "Contact Number", "Owner"}
        ColumnMapping.objects.bulk_create(
            [
                ColumnMapping(
                    profile=profile,
                    source_column=source_column,
                    target_field="candidate:contact",
                )
                for source_column in candidate_columns
            ]
        )

        rows = _interpret(self._workbook(), profile)

        self.assertEqual(set(rows[0]["_candidate_values"]["contact"]), candidate_columns)
        self.assertNotIn("candidate:contact", rows[0])


@skipUnless(LOCAL_EXAMPLE_PATH.is_file(), "Local LibreNMS example workbook is not available")
class LocalExamplePrivacyTest(SimpleTestCase):
    """Ensure the operator's local workbook does not leak into tracked files."""

    def test_private_workbook_values_are_not_embedded_in_repository(self):
        """Tracked files must not contain values copied from private workbook columns."""
        import openpyxl
        import tomllib

        private_headers = {
            "Asset Tag",
            "Asset_Tag",
            "Asset_Tag_Archived",
            "Audit Notes",
            "CANS",
            "City",
            "Company",
            "Contact",
            "Contact Number",
            "Country",
            "Department",
            "Department Director",
            "Description",
            "Dir. Department",
            "Equipment Notes",
            "Express Service Code",
            "Hostname",
            "IDRAC Default Password",
            "IDRAC MAC Address",
            "IP Address (IPv4)",
            "Id",
            "JIRA ID",
            "Location",
            "MAC Address",
            "Management IP Address",
            "Name",
            "Owner",
            "Primary Contact",
            "Project",
            "Purchase Order",
            "Purchase Price",
            "RACK",
            "Rack",
            "Room",
            "Serial Number",
            "Service Provider",
            "Service Tag",
            "SolarWinds ID",
            "VP Department",
            "Wave ID",
        }
        repository_root = Path(__file__).resolve().parents[2]
        with (repository_root / "pyproject.toml").open("rb") as metadata_file:
            project_metadata = tomllib.load(metadata_file)["project"]
        ignored_values = {"n/a", "none", "unknown", "yes", "no"}
        ignored_values.update(
            str(value).strip().casefold() for author in project_metadata.get("authors", []) for value in author.values()
        )
        synthetic_workbook = openpyxl.load_workbook(
            Path(__file__).resolve().parent / "fixtures" / "sample_cans.xlsx",
            read_only=True,
            data_only=True,
        )
        ignored_values.update(
            str(cell.value).strip().casefold()
            for synthetic_sheet in synthetic_workbook.worksheets
            for row in synthetic_sheet.iter_rows()
            for cell in row
            if cell.value is not None
        )
        workbook = openpyxl.load_workbook(LOCAL_EXAMPLE_PATH, read_only=True, data_only=True)
        worksheet = workbook["Data"]
        headers = {cell.column: str(cell.value).strip() for cell in worksheet[1] if cell.value is not None}
        private_values = {}
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if cell.value is None:
                    continue
                header = headers.get(cell.column)
                if header not in private_headers:
                    continue
                value = str(cell.value).strip()
                if len(value) < 4 or "\n" in value or value.casefold() in ignored_values:
                    continue
                if value.isdigit() and len(value) < 6:
                    continue
                private_values.setdefault(value, (header, cell.coordinate))

        token_pattern = re.compile(r"[A-Za-z0-9][A-Za-z0-9@._:/+-]{3,}")
        quoted_pattern = re.compile(r""""([^"\n]{4,})"|'([^'\n]{4,})' """, re.VERBOSE)
        tracked_paths = (
            subprocess.check_output(
                ["git", "-c", f"safe.directory={repository_root}", "ls-files", "-z"],
                cwd=repository_root,
            )
            .decode()
            .split("\0")
        )
        matches = []
        for relative_path in tracked_paths:
            if not relative_path:
                continue
            path = repository_root / relative_path
            try:
                lines = path.read_text().splitlines()
            except (OSError, UnicodeDecodeError):
                continue
            for line_number, line in enumerate(lines, 1):
                lowered_line = line.casefold()
                if (
                    "copyright" in lowered_line
                    or "author_email" in lowered_line
                    or "author =" in lowered_line
                    or "authors =" in lowered_line
                ):
                    continue
                candidates = set(token_pattern.findall(line))
                candidates.update(value.strip() for match in quoted_pattern.findall(line) for value in match if value)
                for value in candidates & private_values.keys():
                    header, coordinate = private_values[value]
                    matches.append(f"{relative_path}:{line_number} matches {header} at Data!{coordinate}")

        self.assertFalse(matches, "Private workbook values are tracked:\n" + "\n".join(matches))


@skipUnless(LOCAL_EXAMPLE_PATH.is_file(), "Local LibreNMS example workbook is not available")
class LocalExampleContactMappingTest(TestCase):
    """Check the operator's local workbook without recording its values."""

    def test_primary_contact_column_maps_to_native_contact_target(self):
        """Populated Primary Contact cells parse as primary_contact values."""
        profile = ImportProfile.objects.create(name="Local Contact Mapping", adapter_config={"sheet_name": "Data"})
        ColumnMapping.objects.create(
            profile=profile,
            source_column="Primary Contact",
            target_field="primary_contact",
        )

        with LOCAL_EXAMPLE_PATH.open("rb") as workbook:
            rows = _interpret(workbook, profile)

        contact_rows = [row for row in rows if row.get("primary_contact")]
        self.assertTrue(contact_rows)
        self.assertTrue(all("_extra_columns" not in row for row in contact_rows))

    def test_configured_contact_candidate_columns_are_collected_per_row(self):
        """Configured columns provide candidate values without asserting their meaning."""
        profile = ImportProfile.objects.create(name="Local Contact Candidates", adapter_config={"sheet_name": "Data"})
        candidate_columns = {"Primary Contact", "Contact", "Contact Number", "Owner"}
        ColumnMapping.objects.bulk_create(
            [
                ColumnMapping(
                    profile=profile,
                    source_column=source_column,
                    target_field="candidate:contact",
                )
                for source_column in candidate_columns
            ]
        )

        with LOCAL_EXAMPLE_PATH.open("rb") as workbook:
            rows = _interpret(workbook, profile)

        candidate_rows = [row for row in rows if row.get("_candidate_values", {}).get("contact")]
        self.assertTrue(candidate_rows)
        self.assertTrue(all(set(row["_candidate_values"]["contact"]) <= candidate_columns for row in candidate_rows))
        self.assertTrue(all("candidate:contact" not in row for row in candidate_rows))


class NativeContactResolverTest(TestCase):
    """Exercise Contact review and apply through the interface the Device module consumes."""

    def setUp(self):
        """Create one imported Device and a profile with native Contact policy."""
        self.site = Site.objects.create(name="Resolver Site", slug="resolver-site")
        manufacturer = Manufacturer.objects.create(name="Resolver Vendor", slug="resolver-vendor")
        device_type = DeviceType.objects.create(
            manufacturer=manufacturer,
            model="Resolver Model",
            slug="resolver-model",
        )
        role = DeviceRole.objects.create(name="Resolver Device", slug="resolver-device")
        self.contact_role = ContactRole.objects.create(name="Resolver Primary", slug="resolver-primary")
        self.profile = ImportProfile.objects.create(
            name="Resolver Profile",
            adapter_config={
                "primary_contact_role": self.contact_role.name,
                "primary_contact_lookup_field": "email",
                "update_existing": True,
            },
        )
        ClassRoleMapping.objects.create(profile=self.profile, source_class="Server", role_slug=role.slug)
        self.device = Device.objects.create(
            name="resolver-device",
            site=self.site,
            device_type=device_type,
            role=role,
        )
        set_import_source(
            self.device,
            self.profile,
            "CONTACT-1",
            extra_columns={"depth": 750, "primary_contact": "legacy@example.invalid"},
        )

    def _row(self, **values):
        """Return one source row with optional Contact decisions."""
        row = {"source_id": "CONTACT-1"}
        row.update(values)
        return row

    def _resolved_row(self, contact_id=None, **field_values):
        """Return one source row carrying a saved Contact decision."""
        return self._row(
            contact_resolution_applied=True,
            contact_field_sources={},
            contact_field_values=field_values,
            contact_id=contact_id,
        )

    def _assignment(self, contact, priority="primary"):
        """Create one assignment for this Device and configured role."""
        return ContactAssignment.objects.create(
            object_type=ContentType.objects.get_for_model(self.device),
            object_id=self.device.pk,
            contact=contact,
            role=self.contact_role,
            priority=priority,
        )

    def test_legacy_contact_is_migrated_to_a_native_assignment(self):
        """Review and apply create one Contact and remove only the legacy JSON key."""
        review = PrimaryContactResolver.review(self.device, self._row(), self.profile)

        plan = PrimaryContactResolver.apply(self.device, self.profile, review)

        contact = Contact.objects.get(email="legacy@example.invalid")
        assignment = ContactAssignment.objects.get(contact=contact, object_id=self.device.pk)
        self.assertEqual(plan["contact_action"], "create")
        self.assertEqual(assignment.priority, "primary")
        self.assertEqual(stored_import_source(self.device).extra_columns, {"depth": 750})

    def test_an_explicit_no_contact_decision_removes_legacy_json_only(self):
        """A reviewed no-contact choice leaves no assignment and still finishes migration."""
        review = ContactReview(None, {}, None, {}, None)

        self.assertIsNone(PrimaryContactResolver.apply(self.device, self.profile, review))
        self.assertFalse(ContactAssignment.objects.exists())
        self.assertEqual(stored_import_source(self.device).extra_columns, {"depth": 750})

    def test_candidates_require_a_decision_and_offer_a_visible_suggestion(self):
        """Unresolved candidates stop planning and suggest only the visible exact Contact."""
        contact = Contact.objects.create(name="Resolver Person", email="person@example.invalid")
        set_import_source(self.device, self.profile, "CONTACT-1", extra_columns={"depth": 750})
        row = self._row(_candidate_values={"contact": {"Owner": contact.name, "Email": contact.email}})

        with self.assertRaises(ContactResolutionRequired) as raised:
            PrimaryContactResolver.review(self.device, row, self.profile)

        self.assertEqual(raised.exception.candidate_values, row["_candidate_values"]["contact"])
        self.assertEqual(raised.exception.suggestion["id"], contact.pk)

    def test_candidate_columns_are_loaded_from_provenance_and_removed_from_extra_data(self):
        """Configured candidate columns become selection values, not retained provenance."""
        ColumnMapping.objects.create(
            profile=self.profile,
            source_column="Candidate Email",
            target_field="candidate:contact",
        )
        set_import_source(
            self.device,
            self.profile,
            "CONTACT-1",
            extra_columns={"Candidate Email": "candidate@example.invalid"},
        )
        row = self._row(
            contact_resolution_applied=True,
            contact_field_sources={"name": "Candidate Email", "email": "Candidate Email"},
        )

        review = PrimaryContactResolver.review(self.device, row, self.profile)

        self.assertEqual(review.candidate_values, {"Candidate Email": "candidate@example.invalid"})
        self.assertEqual(review.extra_columns, {})
        self.assertEqual(review.plan["assignment_action"], "create")

    def test_an_empty_preloaded_candidate_map_skips_the_profile_query(self):
        """A batch with no candidate mappings does not query them again for each Device."""
        with CaptureQueriesContext(connection) as queries:
            PrimaryContactResolver.review(
                self.device,
                self._row(),
                self.profile,
                candidate_source_columns={},
            )

        self.assertFalse(
            any("netbox_data_import_columnmapping" in query["sql"].lower() for query in queries.captured_queries)
        )

    def test_invalid_candidate_decision_preserves_candidates_and_validation_message(self):
        """A bad saved choice remains editable and keeps its precise validation reason."""
        row = self._row(
            _candidate_values={"contact": {"Owner": "Resolver Person"}},
            contact_resolution_applied=True,
            contact_field_sources={"name": "Owner", "email": "Missing"},
        )

        with self.assertRaises(ContactResolutionRequired) as raised:
            PrimaryContactResolver.review(self.device, row, self.profile)

        self.assertIn("has no candidate value", str(raised.exception))
        self.assertEqual(raised.exception.candidate_values, {"Owner": "Resolver Person"})

    def test_literal_validation_without_candidates_is_not_rewritten(self):
        """A literal-only invalid email raises the native validator message."""
        row = self._row(
            contact_resolution_applied=True,
            contact_field_sources={},
            contact_field_values={"name": "Resolver Person", "email": "not-an-email"},
        )

        with self.assertRaisesMessage(ValidationError, "valid email"):
            PrimaryContactResolver.review(self.device, row, self.profile)

    def test_suggestions_respect_lookup_priority_visibility_and_ambiguity(self):
        """Email wins over a name match, while hidden or ambiguous contacts produce no suggestion."""
        by_email = Contact.objects.create(name="Email Person", email="preferred@example.invalid")
        Contact.objects.create(name="Name Person", email="")
        suggestion = PrimaryContactResolver.suggest(
            {"Email": by_email.email, "Owner": "Name Person"},
            self.profile,
        )
        self.assertEqual(suggestion["id"], by_email.pk)

        Contact.objects.create(name="Shared Person", email="first@example.invalid")
        Contact.objects.create(name="Shared Person", email="second@example.invalid")
        self.assertIsNone(PrimaryContactResolver.suggest({"Owner": "Shared Person"}, self.profile))
        self.assertIsNone(PrimaryContactResolver.suggest({"Blank": ""}, self.profile))

        actor = get_user_model().objects.create_user(username="resolver-hidden-contact")
        self.assertIsNone(PrimaryContactResolver.suggest({"Email": by_email.email}, self.profile, actor))

    def test_selected_contact_must_still_exist_be_visible_and_keep_its_identity(self):
        """A stored Contact ID cannot bypass fresh identity and visibility checks."""
        contact = Contact.objects.create(name="Selected Person", email="selected@example.invalid")
        missing_id = contact.pk
        contact.delete()
        with self.assertRaisesMessage(ValidationError, "no longer exists"):
            PrimaryContactResolver.review(self.device, self._resolved_row(contact_id=missing_id), self.profile)

        replacement = Contact.objects.create(name="Selected Person", email="selected@example.invalid")
        with self.assertRaisesMessage(ValidationError, "no longer has the chosen email"):
            PrimaryContactResolver.review(
                self.device,
                self._resolved_row(
                    contact_id=replacement.pk,
                    name=replacement.name,
                    email="changed@example.invalid",
                ),
                self.profile,
            )

        actor = get_user_model().objects.create_user(username="resolver-contact-reader")
        with self.assertRaisesMessage(ObjectPermissionDenied, "tenancy.view_contact"):
            PrimaryContactResolver.review(
                self.device,
                self._resolved_row(contact_id=replacement.pk),
                self.profile,
                actor,
            )

    def test_contact_creation_requires_lookup_contact_and_assignment_permissions(self):
        """Planning refuses each missing permission before any related row is written."""
        row = self._resolved_row(name="New Person", email="new@example.invalid")
        actor = get_user_model().objects.create_user(username="resolver-contact-writer")

        with self.assertRaisesMessage(ObjectPermissionDenied, "tenancy.add_contact"):
            PrimaryContactResolver.review(self.device, row, self.profile, actor)

        actor = user_with_object_permission("resolver-assignment-writer", [(Contact, ("add",), {})])
        with self.assertRaisesMessage(ObjectPermissionDenied, "tenancy.add_contactassignment"):
            PrimaryContactResolver.review(self.device, row, self.profile, actor)

    def test_contact_role_and_lookup_value_are_required(self):
        """Contact planning fails fast when profile policy cannot identify an assignment."""
        self.profile.adapter_config["primary_contact_role"] = None
        self.profile.save(update_fields=["adapter_config"])
        with self.assertRaisesMessage(ValidationError, "Select a primary contact role"):
            PrimaryContactResolver.review(
                self.device,
                self._resolved_row(name="Incomplete Person", email="incomplete@example.invalid"),
                self.profile,
            )

        self.profile.adapter_config["primary_contact_role"] = self.contact_role.name
        self.profile.save(update_fields=["adapter_config"])
        with self.assertRaisesMessage(ValidationError, "Contact email lookup field"):
            PrimaryContactResolver.review(
                self.device,
                self._resolved_row(name="Incomplete Person"),
                self.profile,
            )

    def test_deleted_role_is_refused_when_apply_rechecks_the_review(self):
        """Apply does not trust the role cached during review."""
        review = PrimaryContactResolver.review(
            self.device,
            self._resolved_row(name="Late Person", email="late@example.invalid"),
            self.profile,
        )
        self.contact_role.delete()

        with self.assertRaisesMessage(ValidationError, "no longer exists"):
            PrimaryContactResolver.apply(self.device, self.profile, review)

    def test_assignment_actions_replace_demote_promote_and_reuse_without_duplicates(self):
        """Each existing-assignment shape converges on one primary for the selected Contact."""
        previous = Contact.objects.create(name="Previous Person", email="previous@example.invalid")
        selected = Contact.objects.create(name="Selected Person", email="selected@example.invalid")
        previous_assignment = self._assignment(previous)

        replace = PrimaryContactResolver.review(
            self.device,
            self._row(primary_contact=selected.email),
            self.profile,
        )
        self.assertEqual(replace.plan["assignment_action"], "replace")
        PrimaryContactResolver.apply(self.device, self.profile, replace)
        previous_assignment.refresh_from_db()
        self.assertEqual(previous_assignment.contact, selected)

        previous_assignment.contact = previous
        previous_assignment.save(update_fields=["contact"])
        selected_assignment = self._assignment(selected, priority="secondary")
        promote = PrimaryContactResolver.review(
            self.device,
            self._row(primary_contact=selected.email),
            self.profile,
        )
        self.assertEqual(promote.plan["assignment_action"], "demote_and_promote")
        PrimaryContactResolver.apply(self.device, self.profile, promote)
        previous_assignment.refresh_from_db()
        selected_assignment.refresh_from_db()
        self.assertEqual(previous_assignment.priority, "secondary")
        self.assertEqual(selected_assignment.priority, "primary")

        selected_assignment.priority = "secondary"
        selected_assignment.save(update_fields=["priority"])
        previous_assignment.delete()
        promote_only = PrimaryContactResolver.review(
            self.device,
            self._row(primary_contact=selected.email),
            self.profile,
        )
        self.assertEqual(promote_only.plan["assignment_action"], "promote")
        PrimaryContactResolver.apply(self.device, self.profile, promote_only)
        settled = PrimaryContactResolver.review(
            self.device,
            self._row(primary_contact=selected.email),
            self.profile,
        )
        self.assertEqual(settled.plan["assignment_action"], "unchanged")

    def test_multiple_primary_assignments_and_duplicate_contacts_are_refused(self):
        """Ambiguous contact and assignment identities never choose an arbitrary row."""
        first = Contact.objects.create(name="First Person", email="shared@example.invalid")
        Contact.objects.create(name="Second Person", email="SHARED@example.invalid")
        with self.assertRaisesMessage(ValidationError, "More than one contact"):
            PrimaryContactResolver.review(
                self.device,
                self._resolved_row(name="New Person", email=first.email),
                self.profile,
            )

        Contact.objects.all().delete()
        for index in range(2):
            contact = Contact.objects.create(
                name=f"Primary Person {index}",
                email=f"primary-{index}@example.invalid",
            )
            self._assignment(contact)
        with self.assertRaisesMessage(ValidationError, "More than one primary assignment"):
            PrimaryContactResolver.review(
                self.device,
                self._resolved_row(name="New Person", email="new-primary@example.invalid"),
                self.profile,
            )
