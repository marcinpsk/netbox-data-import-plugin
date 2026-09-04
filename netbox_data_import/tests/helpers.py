# SPDX-License-Identifier: Apache-2.0
# Copyright (C) 2025 Marcin Zieba <marcinpsk@gmail.com>
"""Shared test helpers for netbox_data_import tests."""

import os
import re
from unittest import TestCase
from contextlib import contextmanager
from queue import Queue
from threading import Thread
from time import monotonic, sleep

from django.db import connections

FIXTURE_PATH = os.path.join(os.path.dirname(__file__), "fixtures", "sample_cans.xlsx")


def workbook_bytes(headers, rows, *, sheet_name="Data") -> bytes:
    """Return one in-memory workbook containing *headers* and *rows*."""
    from io import BytesIO

    import openpyxl
    from openpyxl.worksheet.worksheet import Worksheet

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    if not isinstance(worksheet, Worksheet):
        worksheet = workbook.create_sheet()
    worksheet.title = sheet_name
    worksheet.append(list(headers))
    for row in rows:
        worksheet.append(list(row))
    content = BytesIO()
    workbook.save(content)
    return content.getvalue()


TRACE_PATH_HEADER = (
    "Port",
    "PortClass",
    "Cards",
    "Device",
    "UPos",
    "Rack",
    "Location",
    "CableClass",
    "Port",
    "PortClass",
    "Cards",
    "Device",
    "UPos",
    "Rack",
    "Location",
)
TRACE_LIST_HEADER = ("Location", "Rack", "UPos", "Device", "Cards", "Port", "PortClass", "Cable")


def trace_termination(device, cards, port, port_class):
    """Return one compact termination tuple for an in-memory trace workbook."""
    return device, cards, port, port_class


def trace_endpoint_line(termination):
    """Render one endpoint line in the source format."""
    device, cards, port, port_class = termination
    parts = [device]
    if cards:
        parts.append(cards)
    parts.append(f"{port} ({port_class})")
    return " > ".join(parts)


def trace_segment(left, cable_class, right, corroboration=("", "", "")):
    """Render one Segment Evidence row in the source column order."""
    left_device, left_cards, left_port, left_class = left
    right_device, right_cards, right_port, right_class = right
    return (
        left_port,
        left_class,
        left_cards,
        left_device,
        *corroboration,
        cable_class,
        right_port,
        right_class,
        right_cards,
        right_device,
        *corroboration,
    )


def trace_visit(termination):
    """Render one Trace List visit row."""
    device, cards, port, port_class = termination
    return "", "", "", device, cards, port, port_class, "Ignored"


def add_trace_sheet(book, name, header, blocks, export_timestamp):
    """Add one trace sheet with the supplied source blocks."""
    sheet = book.create_sheet(name)
    sheet.append(("Executed", export_timestamp))
    sheet.append(())
    for from_line, to_line, rows in blocks:
        sheet.append(("From", from_line))
        sheet.append(("To", to_line))
        sheet.append(header)
        for row in rows:
            sheet.append(row)
    return sheet


def trace_workbook_bytes(
    *,
    path_blocks=(),
    list_blocks=(),
    include_path=True,
    include_list=False,
    export_timestamp="2026-08-31 12:00:00+00:00",
) -> bytes:
    """Build trace workbook bytes with the fixed trace sheet names."""
    from io import BytesIO

    import openpyxl
    from openpyxl.worksheet.worksheet import Worksheet

    book = openpyxl.Workbook()
    active = book.active
    if isinstance(active, Worksheet):
        book.remove(active)
    if include_path:
        add_trace_sheet(book, "Trace From To", TRACE_PATH_HEADER, path_blocks, export_timestamp)
    if include_list:
        add_trace_sheet(book, "Trace List", TRACE_LIST_HEADER, list_blocks, export_timestamp)
    buffer = BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def store_workbook_document(profile, headers, rows, uploaded_by, filename, *, sheet_name="Data"):
    """Store one generated workbook and return its Source Document."""
    from netbox_data_import.models import SourceDocument

    return SourceDocument.store(
        profile=profile,
        content=workbook_bytes(headers, rows, sheet_name=sheet_name),
        filename=filename,
        uploaded_by=uploaded_by,
    )


def plan_source_rows(rows, profile, site, *, actor=None, location=None, tenant=None):
    """Plan canonical flat-source rows through the registered Target Module interfaces."""
    from netbox_data_import import catalog, target_modules
    from netbox_data_import.adapters import FlatWorkbookAdapter, SourceBatch
    from netbox_data_import.netbox_reader import NetBoxReader
    from netbox_data_import.plan import ImportPlan
    from netbox_data_import.review_workspace import ReviewWorkspace
    from netbox_data_import.source_resolution import derive_effective_rows

    reader = NetBoxReader.for_actor(actor) if actor is not None else NetBoxReader.unrestricted()
    reader = reader.for_target(site=site, location=location, tenant=tenant)
    batch = SourceBatch(
        output_kinds=FlatWorkbookAdapter.output_kinds,
        rows=tuple(derive_effective_rows(list(rows), profile)),
    )
    units = []
    for declaration in catalog.TARGET_MODULES:
        if declaration.consumes & batch.output_kinds:
            units.extend(target_modules.runtime_for(declaration.key).plan(batch, profile, catalog.CATALOG, reader))
    plan = ImportPlan(
        units=tuple(units),
        source_fingerprint="0" * 64,
        profile_fingerprint=profile.planning_fingerprint,
        actor=str(actor.pk) if actor is not None else "test-unrestricted",
        planning_context={
            "site_id": site.pk,
            "location_id": location.pk if location is not None else None,
            "tenant_id": tenant.pk if tenant is not None else None,
        },
    )
    return ReviewWorkspace(plan)


def apply_source_rows(rows, profile, site, *, actor=None, location=None, tenant=None):
    """Apply canonical rows through Target Module runtimes and return their accepted workspace."""
    from django.db import transaction

    from netbox_data_import import target_modules
    from netbox_data_import.netbox_reader import NetBoxReader
    from netbox_data_import.plan import executable_units, merge_changes
    from netbox_data_import.target_modules import ExecutionContext

    workspace = plan_source_rows(
        rows,
        profile,
        site,
        actor=actor,
        location=location,
        tenant=tenant,
    )
    reader = NetBoxReader.for_actor(actor) if actor is not None else NetBoxReader.unrestricted()
    reader = reader.for_target(site=site, location=location, tenant=tenant)
    context = ExecutionContext(actor=actor, reader=reader, profile=profile)
    with transaction.atomic():
        for change in merge_changes(executable_units(workspace.plan.units)):
            target_modules.runtime_for(change.target_module).apply(change, context)
    return workspace


@contextmanager
def run_on_separate_connection(target):
    """Run *target* in a thread with a fresh database connection."""
    errors = Queue()

    def runner():
        connections["default"].close()
        try:
            target()
        except BaseException as exc:
            errors.put(exc)
        finally:
            connections["default"].close()

    thread = Thread(target=runner, daemon=True)
    thread.start()
    try:
        yield
    finally:
        thread.join(timeout=10)
        if thread.is_alive():
            raise AssertionError("The separate database connection did not finish.")
        if not errors.empty():
            raise errors.get()


def user_with_object_permission(username, grants):
    """Create a user holding one real ObjectPermission per model grant."""
    from django.contrib.auth import get_user_model
    from django.contrib.contenttypes.models import ContentType
    from users.models import ObjectPermission

    user = get_user_model().objects.create_user(username=username, password="testpass")
    for model, actions, constraints in grants:
        permission = ObjectPermission.objects.create(
            name=f"{username} {model.__name__} {'-'.join(actions)}",
            actions=list(actions),
            constraints=constraints,
        )
        permission.object_types.add(ContentType.objects.get_for_model(model))
        permission.users.add(user)
    return user


def make_dcim_objects(name_prefix=""):
    """Create a Site, Manufacturer, Device Type, and Device Role with one prefix."""
    from dcim.models import DeviceRole, DeviceType, Manufacturer, Site

    slug_prefix = name_prefix.lower()
    site = Site.objects.create(name=f"{name_prefix}Site", slug=f"{slug_prefix}site")
    manufacturer = Manufacturer.objects.create(name=f"{name_prefix}Mfg", slug=f"{slug_prefix}mfg")
    device_type = DeviceType.objects.create(
        manufacturer=manufacturer,
        model=f"{name_prefix}Model",
        # The importer derives a Device Type slug from make and model, so the fixture must match it.
        slug=f"{slug_prefix}mfg-{slug_prefix}model",
        u_height=1,
    )
    role = DeviceRole.objects.create(name=f"{name_prefix}Role", slug=f"{slug_prefix}role")
    return site, manufacturer, device_type, role


def setup_preview_with_device_matches(client, profile):
    """Populate a preview with two persisted device matches."""
    from dcim.models import Device

    from netbox_data_import.import_engine import ImportEngine
    from netbox_data_import.models import DeviceExistingMatch, SourceDocument
    from netbox_data_import.preview_row_actions import (
        PREVIEW_USE_MATERIALIZED_ONCE_SESSION_KEY,
        record_recalculated_preview,
    )
    from netbox_data_import.review_workspace import ReviewWorkspace

    site, _manufacturer, device_type, role = make_dcim_objects("Match")

    device1 = Device.objects.create(name="device-a", site=site, device_type=device_type, role=role)
    device2 = Device.objects.create(name="device-b", site=site, device_type=device_type, role=role)

    with open(FIXTURE_PATH, "rb") as f:
        content = f.read()
    user = client.session.get("_auth_user_id")
    from django.contrib.auth import get_user_model

    actor = get_user_model().objects.get(pk=user)
    document = SourceDocument.store(
        profile=profile,
        content=content,
        filename="sample_cans.xlsx",
        uploaded_by=actor,
    )
    planning_context = {"site_id": site.pk, "location_id": None, "tenant_id": None}
    result = ReviewWorkspace(ImportEngine.plan(profile, document, actor, planning_context))

    device_rows = [row for row in result.units if row.object_type == "device" and row.source_id]
    if len(device_rows) > 0:
        DeviceExistingMatch.objects.create(
            profile=profile,
            source_id=device_rows[0].source_id,
            source_asset_tag=device_rows[0].extra_data.get("asset_tag", "asset_a"),
            netbox_device_id=device1.id,
            device_name=device1.name,
        )
    if len(device_rows) > 1:
        DeviceExistingMatch.objects.create(
            profile=profile,
            source_id=device_rows[1].source_id,
            source_asset_tag=device_rows[1].extra_data.get("asset_tag", "asset_b"),
            netbox_device_id=device2.id,
            device_name=device2.name,
        )

    plan = ImportEngine.plan(profile, document, actor, planning_context)
    result = ReviewWorkspace(plan)
    session = client.session
    record_recalculated_preview(session, plan)
    session["import_rows"] = result.source_rows
    session["import_context"] = {
        "profile_id": profile.pk,
        "site_id": site.pk,
        "location_id": None,
        "tenant_id": None,
        "filename": "sample_cans.xlsx",
        "source_document_id": document.pk,
    }
    session["import_preview_pending"] = True
    session[PREVIEW_USE_MATERIALIZED_ONCE_SESSION_KEY] = True
    session.save()
    return site, device1, device2, device_rows


def set_import_source(device, profile, source_id="", extra_columns=None, unassigned_ips=None):
    """Store the import record the engine writes for one device."""
    from netbox_data_import.models import DeviceImportSource

    record, _ = DeviceImportSource.objects.update_or_create(
        device=device,
        defaults={
            "profile": profile,
            "source_id": source_id,
            "extra_columns": extra_columns or {},
            "unassigned_ips": unassigned_ips or {},
        },
    )
    return record


def client_with_object_permission(username, model, *, granted, actions=("view",)):
    """Return a logged-in client, failing here so no test can pass on a login redirect."""
    from django.test import Client

    password = "testpass"
    grants = [(model, actions, None)] if granted else []
    user_with_object_permission(username, grants)
    client = Client()
    if not client.login(username=username, password=password):
        raise AssertionError(f"the test client could not log in as '{username}'")
    return client


def wait_until_a_lock_is_blocked(test, timeout=10):
    """Block until another backend is waiting for a lock this connection holds."""
    from django.db import connection

    # pg_stat_activity is cached for the whole transaction; pg_locks reads live lock-manager state.
    query = "SELECT count(*) FROM pg_locks WHERE NOT granted AND pg_backend_pid() = ANY(pg_blocking_pids(pid))"
    deadline = monotonic() + timeout
    while monotonic() < deadline:
        with connection.cursor() as cursor:
            cursor.execute(query)
            if cursor.fetchone()[0]:
                return
        sleep(0.05)
    test.fail("No other backend started waiting for a lock this connection holds.")


def action_link_tag(html: str, href: str) -> str:
    """Return the opening anchor tag that targets *href*, so a test can read its attributes."""
    match = re.search(rf'<a[^>]*href="{re.escape(href)}"[^>]*>', html)
    return match.group(0) if match else ""


def assert_action_link_is_named(test: TestCase, html: str, href: str, name: str) -> None:
    """Fail unless the icon-only action link at *href* announces *name* to a screen reader."""
    tag = action_link_tag(html, href)
    test.assertNotEqual(tag, "", f"No action link renders for {href}.")
    match = re.search(r'aria-label="([^"]*)"', tag)
    if match is None:
        test.fail(f"The action link for {href} has no accessible name: {tag}")
    test.assertIn(name, match.group(1))
