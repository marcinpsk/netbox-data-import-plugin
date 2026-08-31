# SPDX-License-Identifier: Apache-2.0
# SPDX-FileCopyrightText: 2026 Marcin Zieba <marcinpsk@gmail.com>
"""Interpret flat workbooks without accessing the ORM."""

from __future__ import annotations

from typing import Any

from dataclasses import dataclass, field
from io import BytesIO
from time import monotonic

import openpyxl
import regex

from .adapters import SourceUnreadable
from .catalog import CANDIDATE_TARGET_PREFIX
from .values import comparison_key

EXTRA_JSON_PREFIX = "extra_json:"
_MAX_UNUSED_SAMPLES = 5
_TRANSFORM_REGEX_MATCH_TIMEOUT_SECONDS = 0.5
_TRANSFORM_REGEX_WORKBOOK_TIMEOUT_SECONDS = 1.0


@dataclass(frozen=True)
class TransformRule:
    """One regex that splits a source column into up to two Target Fields."""

    source_column: str
    pattern: str
    group_1_target: str = ""
    group_2_target: str = ""


@dataclass(frozen=True)
class FlatWorkbookConfig:
    """Everything the flat adapter needs to interpret one workbook."""

    sheet_name: str
    column_map: dict[str, tuple[str, ...]] = field(default_factory=dict)
    transform_rules: tuple[TransformRule, ...] = ()
    capture_extra_data: bool = False


@dataclass
class _TransformRegexBudget:
    """Bound cumulative regex execution across one workbook."""

    remaining_seconds: float = _TRANSFORM_REGEX_WORKBOOK_TIMEOUT_SECONDS

    def fullmatch(self, pattern: str, text: str):
        """Match with the smaller of the per-match and remaining workbook budgets."""
        if self.remaining_seconds <= 0:
            raise TimeoutError
        started_at = monotonic()
        try:
            return regex.fullmatch(
                pattern,
                text,
                timeout=min(_TRANSFORM_REGEX_MATCH_TIMEOUT_SECONDS, self.remaining_seconds),
            )
        finally:
            self.remaining_seconds -= monotonic() - started_at


def _text(value) -> str:
    """Return the trimmed text of a cell value, empty for None."""
    return "" if value is None else str(value).strip()


def _header_index_map(sheet) -> dict[str, int]:
    """Map each header name in row 1 to its column index; the first of a duplicate wins."""
    headers: dict[str, int] = {}
    for index, cell in enumerate(sheet[1]):
        if cell.value is not None:
            name = str(cell.value).strip()
            if name not in headers:
                headers[name] = index
    return headers


def _cell(raw_row, index) -> object:
    """Return the raw cell at *index*, or None when the row is short."""
    return raw_row[index] if index is not None and index < len(raw_row) else None


def _merge_row_values(row_number: int, raw_row, headers: dict[str, int], column_map) -> dict[str, Any]:
    """Build one row dict, recording a conflict when two source columns disagree."""
    row: dict[str, Any] = {"_row_number": row_number}
    for target_field, source_columns in column_map.items():
        values: dict[str, Any] = {}
        for source_column in source_columns:
            value = _cell(raw_row, headers.get(source_column))
            if isinstance(value, str):
                value = value.strip()
            if value is not None and str(value).strip():
                values[source_column] = value

        if not values:
            continue
        if target_field.startswith(CANDIDATE_TARGET_PREFIX):
            candidate_target = target_field.removeprefix(CANDIDATE_TARGET_PREFIX)
            row.setdefault("_candidate_values", {})[candidate_target] = {
                source_column: str(value) for source_column, value in values.items()
            }
            continue
        if len({comparison_key(target_field, value) for value in values.values()}) == 1:
            row[target_field] = next(iter(values.values()))
        else:
            row[target_field] = None
            row.setdefault("_conflicts", {})[target_field] = {
                source_column: str(value) for source_column, value in values.items()
            }
    return row


def promote_extra_json_fields(row: dict) -> None:
    """Move every `extra_json:<name>` entry into the row's captured extra columns."""
    for key in [key for key in list(row) if isinstance(key, str) and key.startswith(EXTRA_JSON_PREFIX)]:
        value = row.pop(key)
        if value not in (None, ""):
            row.setdefault("_extra_columns", {})[key[len(EXTRA_JSON_PREFIX) :]] = value


def _apply_transform_rules(
    row: dict,
    raw_row,
    headers: dict[str, int],
    rules,
    budget: _TransformRegexBudget,
) -> None:
    """Apply each transform rule in place, refusing invalid or over-budget patterns."""
    for rule in rules:
        raw_value = _cell(raw_row, headers.get(rule.source_column))
        if raw_value is None:
            continue
        text = str(raw_value).strip()
        try:
            match = budget.fullmatch(rule.pattern, text)
        except TimeoutError as exc:
            raise SourceUnreadable(
                f"Regex pattern in transform rule for column '{rule.source_column}' timed out."
            ) from exc
        except regex.error as exc:
            raise SourceUnreadable(
                f"Invalid regex pattern '{rule.pattern}' in transform rule for column "
                f"'{rule.source_column}' (value: {text!r}): {exc}"
            ) from exc
        if match is None:
            continue
        if rule.group_1_target and len(match.groups()) >= 1:
            row[rule.group_1_target] = match.group(1)
        if rule.group_2_target and len(match.groups()) >= 2:
            row[rule.group_2_target] = match.group(2)


def _collect_unmapped_values(raw_row, headers, unmapped_columns, unused_stats, keep_stats, capture) -> dict[str, str]:
    """Return this row's unmapped values and add them to *unused_stats* in place."""
    extra: dict[str, str] = {}
    for column in unmapped_columns:
        text = _text(_cell(raw_row, headers[column]))
        if not text:
            continue
        if capture:
            extra[column] = text
        if keep_stats:
            entry = unused_stats.setdefault(column, {"count": 0, "samples": []})
            entry["count"] += 1
            if len(entry["samples"]) < _MAX_UNUSED_SAMPLES:
                entry["samples"].append(text)
    return extra


def _open_sheet(content: bytes, sheet_name: str):
    """Return the named worksheet, or say which sheets the file actually has."""
    try:
        book = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise SourceUnreadable(f"Cannot open Excel file: {exc}") from exc
    if sheet_name not in book.sheetnames:
        available = ", ".join(book.sheetnames)
        raise SourceUnreadable(f"Sheet '{sheet_name}' not found. Available sheets: {available}")
    return book[sheet_name]


def interpret(content: bytes, config: FlatWorkbookConfig, *, collect_unused: bool = False):
    """Return the rows one workbook carries, and the unmapped-column tally when asked for it."""
    sheet = _open_sheet(content, config.sheet_name)
    headers = _header_index_map(sheet)
    mapped = {column for columns in config.column_map.values() for column in columns}
    transformed = {rule.source_column for rule in config.transform_rules}
    consumed_columns = mapped | transformed
    unmapped_columns = [column for column in headers if column not in consumed_columns]
    unused_stats: dict[str, dict] = {}
    transform_budget = _TransformRegexBudget()

    rows = []
    for row_number, raw_row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(value is None for value in raw_row):
            continue
        row = _merge_row_values(row_number, raw_row, headers, config.column_map)
        _apply_transform_rules(row, raw_row, headers, config.transform_rules, transform_budget)
        promote_extra_json_fields(row)
        if collect_unused or config.capture_extra_data:
            extra = _collect_unmapped_values(
                raw_row, headers, unmapped_columns, unused_stats, collect_unused, config.capture_extra_data
            )
            if extra:
                row.setdefault("_extra_columns", {}).update(extra)
        rows.append(row)

    return rows, unused_stats


__all__ = (
    "EXTRA_JSON_PREFIX",
    "FlatWorkbookConfig",
    "TransformRule",
    "interpret",
    "promote_extra_json_fields",
)
